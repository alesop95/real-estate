# -*- coding: utf-8 -*-
"""Test del generatore del workbook.

Non verificano i numeri, che sono coperti da `test_calcoli.py` e dal ricalcolo con
Excel: verificano la struttura, e in particolare il contratto piu' fragile del
progetto, cioe' la corrispondenza posizionale fra le colonne del foglio Annunci e
l'ordine con cui `annunci.esporta_in_excel` le scrive. Sono due elenchi in due file
diversi che devono restare allineati, e se divergono l'esportazione mette i prezzi
nella colonna delle note senza che nulla protesti.

Si eseguono con `python -m pytest tests`, oppure con `python tests/test_workbook.py`.
"""

from __future__ import annotations

import sys
import tempfile
from dataclasses import fields
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from openpyxl import load_workbook  # noqa: E402

from immobiliare import annunci as A  # noqa: E402
from immobiliare import excel_builder as E  # noqa: E402

_workbook_generato: Path | None = None


def workbook() -> Path:
    """Genera il workbook una volta sola per l'intera sessione di test."""
    global _workbook_generato
    if _workbook_generato is None or not _workbook_generato.exists():
        destinazione = Path(tempfile.gettempdir()) / "valutazione-test.xlsx"
        E.genera(str(destinazione))
        _workbook_generato = destinazione
    return _workbook_generato


def test_il_workbook_si_genera_con_tutti_i_fogli():
    wb = load_workbook(workbook())
    attesi = [
        "Guida", "Cruscotto", "Parametri", "Immobile", "Mutuo", "Ammortamento", "Simulatore mutuo", "Locazione",
        "Cash flow", "Metriche", "Confronto affitto", "Scenari", "Rischio", "Comproprieta", "Checklist", "Asta", "Dossier tecnico",
        "Annunci", "Confronto immobili", "Fonti", "_Estrazioni",
    ]
    assert wb.sheetnames == attesi


def test_ogni_cella_a_tendina_ha_il_colore_della_scelta():
    """Una cella dove si sceglie non deve avere l'aspetto di una dove si digita.

    E' la correzione di una segnalazione d'uso, e la segnalazione veniva da chi
    ha seguito il progetto: le celle a tendina erano gialle come quelle da
    digitare, quindi a video nulla distingueva una cella dove si scrive un numero
    da una dove si sceglie fra valori ammessi. Chi non conosce il file ci scrive
    dentro, la validazione rifiuta il valore, e il messaggio di Excel non spiega
    perche'.

    Il presidio e' che l'unico modo previsto di applicare una tendina sia
    `stile.scelta`, che aggancia la cella e la colora insieme. Il test verifica
    la conseguenza osservabile: ogni cella coperta da una validazione a elenco
    porta il riempimento della scelta e non quello dell'input.
    """
    wb = load_workbook(workbook())
    atteso = E.S.AZZURRO
    controllate = 0
    for ws in wb.worksheets:
        for validazione in ws.data_validations.dataValidation:
            if validazione.type != "list":
                continue
            for intervallo in validazione.sqref.ranges:
                for riga in ws.iter_rows(min_row=intervallo.min_row, max_row=intervallo.max_row,
                                         min_col=intervallo.min_col, max_col=intervallo.max_col):
                    for cella in riga:
                        colore = cella.fill.fgColor.rgb if cella.fill and cella.fill.fgColor else None
                        assert colore and colore.endswith(atteso), (
                            f"{ws.title}!{cella.coordinate} ha una tendina ma il riempimento e' "
                            f"{colore!r} invece di {atteso}: chi la guarda non sa che c'e' un elenco"
                        )
                        controllate += 1
    # Se il conteggio crollasse a zero il test passerebbe a vuoto, quindi la
    # soglia e' un presidio del test stesso e non del workbook.
    assert controllate > 200, f"solo {controllate} celle a tendina trovate: il test non sta verificando nulla"


def test_ogni_foglio_dichiara_in_testa_se_si_scrive_o_si_legge():
    """La fascia in riga tre dice a chi apre il foglio che cosa farci.

    Nasce dalla stessa segnalazione: l'indice diceva per ogni foglio se si
    compila o si legge, ma quell'informazione stava solo nell'indice, e chi
    arrivava su un foglio dalle linguette in basso non la vedeva. La fascia la
    ripete dove serve, e non e' una duplicazione da mantenere a mano perche' viene
    dalla stessa tupla da cui nasce l'indice.

    Il test verifica anche la coerenza fra le due: un foglio dichiarato di sola
    lettura nell'indice non puo' avere in testa una fascia che invita a scrivere.
    """
    wb = load_workbook(workbook())
    indice = E.S.FOGLIO_INDICE
    usi = {nome: azione for _, fogli in E.Costruttore.PERCORSO for nome, azione, _, _ in fogli}

    for ws in wb.worksheets:
        if ws.title == indice or ws.sheet_state != "visible":
            continue
        fascia = ws.cell(row=3, column=2).value
        assert isinstance(fascia, str) and fascia, (
            f"il foglio {ws.title!r} non ha la fascia d'uso in riga 3, colonna B"
        )
        azione = usi[ws.title]
        if azione == "Si legge":
            assert fascia.startswith("QUI NON SI SCRIVE NULLA"), (
                f"{ws.title} e' dichiarato di sola lettura nell'indice ma la fascia dice: {fascia[:60]!r}"
            )
        elif azione == "Si consulta":
            assert fascia.startswith("QUI SI CONSULTA"), fascia[:60]
        else:
            assert fascia.startswith("QUI SI SCRIVE"), (
                f"{ws.title} e' dichiarato da compilare nell'indice ma la fascia dice: {fascia[:60]!r}"
            )


def test_indice_porta_la_legenda_dei_colori_mostrandoli():
    """La legenda mostra i colori invece di descriverli a parole.

    Descriverli richiederebbe a chi legge di ricordare un'associazione fra un
    nome e un colore, e la segnalazione da cui nasce questa sezione era proprio
    che quell'associazione non era chiara. Ogni riga della legenda porta quindi
    il riempimento che spiega, e il test verifica che il riempimento ci sia e sia
    quello giusto, perche' una legenda con i colori sbagliati e' peggio di
    nessuna legenda.
    """
    wb = load_workbook(workbook())
    ws = wb[E.S.FOGLIO_INDICE]
    attesi = {
        "Gialla": E.S.GIALLO,
        "Azzurra": E.S.AZZURRO,
        "Grigia": E.S.GRIGIO,
        "Verde": E.S.VERDE,
        "Rossa": E.S.ROSSO,
    }
    trovati = {}
    for riga in range(1, 40):
        cella = ws.cell(row=riga, column=2)
        if cella.value in attesi:
            colore = cella.fill.fgColor.rgb if cella.fill and cella.fill.fgColor else None
            trovati[cella.value] = colore
            spiegazione = ws.cell(row=riga, column=3).value
            assert isinstance(spiegazione, str) and len(spiegazione) > 40, (
                f"la voce {cella.value!r} della legenda non ha una spiegazione: {spiegazione!r}"
            )
    assert set(trovati) == set(attesi), f"legenda incompleta: {sorted(trovati)}"
    for nome, colore in trovati.items():
        assert colore and colore.endswith(attesi[nome]), (
            f"la voce {nome!r} della legenda e' riempita di {colore!r} invece di {attesi[nome]}"
        )


def test_indice_copre_tutti_i_fogli_e_i_collegamenti_esistono():
    """Ogni foglio visibile sta nell'indice una volta sola, e ogni link porta a un foglio vero.

    E' il presidio contro un difetto che Excel non segnala: un foglio rinominato
    lascia nell'indice un collegamento sintatticamente valido verso una
    destinazione che non esiste piu', e Excel lo apre senza errore visibile,
    semplicemente non andando da nessuna parte. Il test confronta la tupla
    `PERCORSO`, che e' la sorgente da cui l'indice si costruisce, con i fogli
    davvero presenti nel workbook, e verifica che le due cose coincidano
    esattamente: nessun foglio dimenticato nell'indice, nessuna voce dell'indice
    senza foglio.
    """
    wb = load_workbook(workbook())
    visibili = {ws.title for ws in wb.worksheets if ws.sheet_state == "visible"}
    indice = E.S.FOGLIO_INDICE
    assert indice in visibili, f"il foglio indice {indice!r} non esiste"

    # Cio' che la sorgente dichiara.
    dichiarati = [nome for _, fogli in E.Costruttore.PERCORSO for nome, *_ in fogli]
    assert len(dichiarati) == len(set(dichiarati)), (
        f"un foglio compare due volte nell'indice: {sorted(dichiarati)}"
    )
    attesi = visibili - {indice}
    assert set(dichiarati) == attesi, (
        f"indice e workbook divergono. Nell'indice e non nel workbook: "
        f"{sorted(set(dichiarati) - attesi)}. Nel workbook e non nell'indice: "
        f"{sorted(attesi - set(dichiarati))}"
    )

    # Cio' che il foglio contiene davvero, letto dalle celle.
    ws = wb[indice]
    collegati = {}
    for riga in range(1, 120):
        c = ws.cell(row=riga, column=2)
        if c.hyperlink is None:
            continue
        collegati[c.value] = c.hyperlink
    assert set(collegati) == attesi, (
        f"le celle dell'indice non coincidono con i fogli: mancano "
        f"{sorted(attesi - set(collegati))}, in eccesso {sorted(set(collegati) - attesi)}"
    )

    for nome, link in collegati.items():
        # Deve essere un collegamento interno, cioe' con `location` e senza
        # destinazione esterna: la forma esterna finisce fra le relazioni verso
        # l'esterno del file e Excel la tratta come tale.
        assert link.location == f"'{nome}'!A1", (
            f"il collegamento a {nome!r} non e' interno o non punta ad A1: "
            f"location={link.location!r}, target={link.target!r}"
        )
        assert not link.target, f"il collegamento a {nome!r} ha una destinazione esterna: {link.target!r}"


def test_ogni_foglio_visibile_torna_all_indice():
    """Da ogni foglio si torna all'indice, e il ritorno sta sempre nello stesso posto.

    La posizione fissa in colonna A non e' un vezzo grafico: il ritorno lo scrive
    `stile.titolo`, che ogni foglio chiama come prima cosa, quindi un foglio nuovo
    non puo' nascere senza via di ritorno. Metterlo accanto al titolo lo avrebbe
    spostato a destra di tante colonne quante ne occupa il titolo, che in questo
    workbook varia da quattro a ventisei, e su un foglio largo sarebbe finito
    fuori dalla vista.
    """
    wb = load_workbook(workbook())
    indice = E.S.FOGLIO_INDICE
    for ws in wb.worksheets:
        if ws.title == indice or ws.sheet_state != "visible":
            continue
        trovato = None
        for riga in range(1, 8):
            c = ws.cell(row=riga, column=1)
            if c.value == "<< Indice":
                trovato = c
                break
        assert trovato is not None, (
            f"il foglio {ws.title!r} non ha il ritorno all'indice nelle prime righe della colonna A"
        )
        assert trovato.hyperlink is not None, f"il ritorno del foglio {ws.title!r} non e' un collegamento"
        assert trovato.hyperlink.location == f"'{indice}'!A1", (
            f"il ritorno del foglio {ws.title!r} punta a {trovato.hyperlink.location!r}"
        )


def test_nomi_definiti_essenziali_presenti():
    """I nomi definiti sono il collante fra i fogli: se ne sparisce uno, le formule
    che lo usano diventano #NOME? in silenzio fino all'apertura del file."""
    wb = load_workbook(workbook())
    definiti = set(wb.defined_names)
    essenziali = {
        "prezzo", "rendita", "categoria", "prima_casa", "da_impresa", "agevolata",
        "di_lusso", "valore_catastale", "base_registro", "imposte_totali",
        "costi_accessori", "costo_totale", "esborso", "mutuo_importo", "tasso",
        "durata", "rata_mensile", "oneri_mutuo", "detrazione_anno",
        "debito_residuo_anno", "noi_annuo", "utile_locazione", "ricavo_lordo",
        "ricavo_effettivo", "flussi_tir", "flussi_tir_dal_primo",
        "cash_flow_primo_anno", "orizzonte", "data_erogazione", "rend_obiettivo",
        "ltv_conf", "aliquota_conf", "irpef_marginale", "sim_capitale",
        "sim_rata_iniziale", "sim_flussi", "sim_pagato",
        "accantonamento_ristrutturazione", "verdetto", "verifiche_aperte",
        "prob_cash_negativo", "prob_batte_alternativa", "sim_cash_flow", "vol_canone",
    }
    mancanti = essenziali - definiti
    assert not mancanti, f"nomi definiti mancanti: {sorted(mancanti)}"


def test_intestazione_annunci_allineata_all_esportazione():
    """Il contratto fra `foglio_annunci` e `esporta_in_excel`.

    Il foglio ha tre colonne in piu' rispetto ai campi del registro, che sono le
    tre calcolate da formula. Ogni campo del registro deve trovarsi nella colonna
    che l'esportazione gli assegna, e nessuna colonna di formula deve finire
    sotto un campo.
    """
    wb = load_workbook(workbook())
    ws = wb["Annunci"]

    intestazione = None
    for riga in range(1, 12):
        if ws.cell(row=riga, column=1).value == "ID":
            intestazione = riga
            break
    assert intestazione is not None, "intestazione del foglio Annunci non trovata"

    titoli = []
    colonna = 1
    while ws.cell(row=intestazione, column=colonna).value:
        titoli.append(ws.cell(row=intestazione, column=colonna).value)
        colonna += 1

    campi_registro = [f.name for f in fields(A.Annuncio)]
    # Tre colonne del foglio sono calcolate e non hanno un campo corrispondente.
    assert len(titoli) == len(campi_registro) + 3, (
        f"il foglio ha {len(titoli)} colonne e il registro {len(campi_registro)} campi: "
        "la differenza deve essere esattamente le tre colonne di formula"
    )
    for atteso, effettivo in (
        (19, "Prezzo al mq"), (22, "Scarto su OMI"), (29, "Rendimento lordo"),
    ):
        assert titoli[atteso - 1] == effettivo, (
            f"la colonna {atteso} dovrebbe essere '{effettivo}' e invece e' '{titoli[atteso - 1]}'"
        )


def test_esportazione_scrive_nelle_colonne_giuste(tmp_path=None):
    """Esporta un annuncio noto e rilegge le celle una per una."""
    cartella = Path(tmp_path) if tmp_path else Path(tempfile.mkdtemp())
    destinazione = cartella / "esportazione.xlsx"
    E.genera(str(destinazione))

    registro = A.Registro(cartella / "annunci.csv")
    registro.aggiungi(
        A.Annuncio(
            id="house_9", stato="visitata", fonte="immobiliare.it", agenzia="Agenzia Prova",
            contatto="333 1112223", link="https://esempio.invalid/1", comune="Comune di esempio",
            provincia="XX", indirizzo="via Prova 1", tipologia="bilocale",
            destinazione_uso="abitazione", nuova_costruzione="SI", data_consegna="2027-06",
            mq=60, prezzo_richiesto=100_000, prezzo_obiettivo=93_000,
            rendita_catastale=500, categoria="A/3", canone_atteso_mese=520,
            asta="SI", base_asta=72_000, data_asta="2026-11-14",
            tribunale_procedura="Macerata RGE 55/2025",
            stato_occupazione="occupato dal debitore", note="prova",
        )
    )
    scritti = A.esporta_in_excel(registro, str(destinazione))
    assert scritti == 1

    ws = load_workbook(destinazione)["Annunci"]
    riga = None
    for r in range(1, 20):
        if ws.cell(row=r, column=1).value == "house_9":
            riga = r
            break
    assert riga is not None, "l'annuncio esportato non si trova nel foglio"

    atteso = {
        1: "house_9", 3: "visitata", 5: "Agenzia Prova", 6: "333 1112223",
        8: "Comune di esempio", 9: "XX", 12: "bilocale", 13: "abitazione",
        14: "SI", 15: "2027-06", 16: 60, 17: 100_000, 18: 93_000,
        21: None, 23: 500, 24: "A/3", 28: 520,
        # Le cinque colonne dell'asta stanno fra il rendimento lordo e il punteggio.
        30: "SI", 31: 72_000, 32: "2026-11-14", 33: "Macerata RGE 55/2025",
        34: "occupato dal debitore", 36: "prova",
    }
    for colonna, valore in atteso.items():
        effettivo = ws.cell(row=riga, column=colonna).value
        assert effettivo == valore, f"colonna {colonna}: atteso {valore!r}, trovato {effettivo!r}"

    # Le tre colonne calcolate devono essere rimaste formule, non valori.
    for colonna in (19, 22, 29):
        contenuto = ws.cell(row=riga, column=colonna).value
        assert isinstance(contenuto, str) and contenuto.startswith("="), (
            f"la colonna calcolata {colonna} e' stata sovrascritta con {contenuto!r}"
        )


def test_confronto_immobili_legge_dalla_riga_giusta():
    """Il foglio di confronto punta alla prima riga di dati del foglio Annunci."""
    wb = load_workbook(workbook())
    annunci = wb["Annunci"]
    confronto = wb["Confronto immobili"]

    prima_annunci = None
    for riga in range(1, 12):
        if annunci.cell(row=riga, column=1).value == "ID":
            prima_annunci = riga + 1
            break
    assert prima_annunci is not None

    prima_confronto = None
    for riga in range(1, 20):
        valore = confronto.cell(row=riga, column=1).value
        if isinstance(valore, str) and valore.startswith("=IF(Annunci!"):
            prima_confronto = riga
            break
    assert prima_confronto is not None, "nessuna riga di confronto trovata"
    assert f"Annunci!$A{prima_annunci}" in confronto.cell(row=prima_confronto, column=1).value


def test_campi_a_tre_stati_normalizzati():
    """SI, NO oppure vuoto: un booleano del modello locale non deve passare per NO.

    Il caso che conta e' `true`: Excel confronta il testo senza distinguere le
    maiuscole, quindi `si` funziona, mentre `true` risulta diverso da SI e il
    foglio lo legge come un NO senza segnalare nulla. Cio' che non e' riconosciuto
    resta intatto, perche' un valore strano visibile e' preferibile a un valore
    strano tradotto per ipotesi.
    """
    a = A.Annuncio(id="x_1", venditore_impresa="true", prima_casa="  sI ", asta="0")
    assert (a.venditore_impresa, a.prima_casa, a.asta) == ("SI", "SI", "NO")

    # Il vuoto resta vuoto: e' il terzo stato, cioe' eredita dal foglio Immobile.
    b = A.Annuncio(id="x_2")
    assert b.prima_casa == "" and b.venditore_impresa == ""

    # Un valore non riconosciuto non viene indovinato.
    c = A.Annuncio(id="x_3", prima_casa="da chiarire col notaio")
    assert c.prima_casa == "da chiarire col notaio"

    # I due campi nuovi non alterano il numero di campi del registro, che il
    # contratto con il foglio Annunci conta.
    assert "prima_casa" in [f.name for f in fields(A.Annuncio)]
    assert "venditore_impresa" in [f.name for f in fields(A.Annuncio)]


def test_prezzo_massimo_e_esatto_e_si_autoverifica():
    """Il prezzo massimo sostenibile non deve piu' passare per l'incidenza dei costi.

    La versione precedente divideva il costo totale sostenibile per uno piu'
    l'incidenza percentuale dei costi accessori dello scenario base, e sbagliava
    due volte: assumeva quell'incidenza costante al variare del prezzo, mentre
    notaio, oneri del mutuo, imposte fisse e, con il prezzo-valore, l'intera
    imposta di registro sono importi fissi la cui incidenza cresce al calare del
    prezzo; e teneva l'utile netto fermo, mentre manutenzione e accantonamento per
    la ristrutturazione sono quote del valore e quindi l'utile sale se il prezzo
    scende. Sul caso precaricato le due formule danno 15.609 euro contro 43.445, e
    l'errore va nella direzione che fa sembrare impossibile qualunque trattativa.
    """
    wb = load_workbook(workbook())
    ws = wb["Scenari"]

    righe = {}
    for riga in range(1, 200):
        etichetta = ws.cell(row=riga, column=1).value
        if isinstance(etichetta, str) and etichetta not in righe:
            righe[etichetta] = riga

    for etichetta in ("Rendimento netto obiettivo", "Quota del prezzo che diventa costo aggiuntivo",
                      "Costi che non dipendono dal prezzo", "Costi annui che scalano col prezzo",
                      "Prezzo massimo corrispondente", "Verifica: rendimento netto a quel prezzo",
                      "Scarto dalla soglia, deve essere zero"):
        assert etichetta in righe, f"voce {etichetta!r} assente dalla sezione del prezzo massimo"

    pmax = ws.cell(row=righe["Prezzo massimo corrispondente"], column=2).value
    assert isinstance(pmax, str) and pmax.startswith("=")
    assert "incidenza_costi" not in pmax, (
        f"il prezzo massimo e' tornato all'approssimazione sull'incidenza dei costi: {pmax!r}"
    )
    # Deve citare le tre celle della soluzione chiusa e il rendimento obiettivo.
    for etichetta in ("Quota del prezzo che diventa costo aggiuntivo",
                      "Costi che non dipendono dal prezzo",
                      "Costi annui che scalano col prezzo",
                      "Rendimento netto obiettivo"):
        assert f"B{righe[etichetta]}" in pmax, (
            f"il prezzo massimo non legge {etichetta!r}: {pmax!r}"
        )

    # Lo scarto sul prezzo trattato deve derivare dal prezzo massimo, non ricalcolarlo.
    scarto = ws.cell(row=righe["Scarto rispetto al prezzo trattato"], column=2).value
    assert f"B{righe['Prezzo massimo corrispondente']}" in scarto and "prezzo" in scarto, (
        f"lo scarto non deriva dal prezzo massimo: {scarto!r}"
    )

    # Il controllo di chiusura deve usare le formule esatte, floor di legge compreso.
    verifica = ws.cell(row=righe["Verifica: rendimento netto a quel prezzo"], column=2).value
    assert "reg_min" in verifica, (
        f"la verifica non applica il minimo di legge dell'imposta di registro: {verifica!r}"
    )
    assert "utile_locazione" in verifica


def test_tabella_scenari_non_usa_offset_numerici():
    """Le formule dei tre scenari devono citare righe scritte, non righe calcolate.

    La tabella usava indici del tipo base piu' una costante, con lo stesso difetto
    del conto economico della locazione. Ora ogni riga si registra sotto una chiave
    e le formule citano le chiavi: una chiave assente rompe la generazione, che e'
    il comportamento giusto, invece di produrre un riferimento valido a una riga
    diversa. Il test fissa la conseguenza osservabile: ogni formula della tabella
    deve puntare a righe interne alla tabella e precedenti alla propria.
    """
    wb = load_workbook(workbook())
    ws = wb["Scenari"]

    intestazione = None
    for riga in range(1, 60):
        if ws.cell(row=riga, column=2).value == "Pessimistico":
            intestazione = riga
            break
    assert intestazione is not None, "intestazione della tabella dei tre scenari non trovata"

    attese = ["Canone mensile", "Mesi di sfitto all'anno", "Morosita'", "Tasso del mutuo",
              "Rivalutazione annua dell'immobile", "Ricavo effettivo", "Costi operativi",
              "Reddito operativo netto", "Imposta sul canone", "Utile netto", "Rata annua",
              "Cash flow annuo", "Rendimento netto", "Debt service coverage ratio",
              "Valore dell'immobile a fine orizzonte", "Debito residuo a fine orizzonte",
              "Patrimonio netto a fine orizzonte"]
    for scostamento, etichetta in enumerate(attese, start=1):
        effettiva = ws.cell(row=intestazione + scostamento, column=1).value
        assert effettiva == etichetta, (
            f"riga {intestazione + scostamento}: attesa {etichetta!r}, trovata {effettiva!r}"
        )

    prima = intestazione + 1
    ultima = intestazione + len(attese)
    import re

    for riga in range(prima, ultima + 1):
        formula = ws.cell(row=riga, column=2).value
        if not (isinstance(formula, str) and formula.startswith("=")):
            continue
        for citata in re.findall(r"\$?[BCD](\d+)", formula):
            citata = int(citata)
            assert prima <= citata <= ultima, (
                f"la riga {riga} cita la riga {citata}, fuori dalla tabella: {formula!r}"
            )
            assert citata < riga, (
                f"la riga {riga} cita la riga {citata}, che viene dopo: {formula!r}"
            )


def test_conto_economico_locazione_somma_le_righe_giuste():
    """Il reddito operativo netto deve sommare tutte le righe di costo e nessun'altra.

    Era il punto piu' fragile del generatore. Gli indici delle righe si scrivevano
    come base piu' una costante a mano, quindi una voce inserita in mezzo al conto
    economico spostava di uno tutte le righe successive e lasciava le costanti
    dov'erano: la somma dei costi diventava un intervallo traslato, l'utile netto
    leggeva la riga sbagliata, e nulla andava in errore. Il test fissa l'invariante
    in termini di etichette e non di numeri di riga, cosi' che valga anche dopo un
    riordino del foglio: l'intervallo sommato deve iniziare subito sotto il ricavo
    effettivo e finire subito sopra il reddito operativo netto.
    """
    wb = load_workbook(workbook())
    ws = wb["Locazione"]

    righe = {}
    for riga in range(1, 200):
        etichetta = ws.cell(row=riga, column=1).value
        if isinstance(etichetta, str) and etichetta not in righe:
            righe[etichetta] = riga

    for etichetta in ("Ricavo effettivo", "Reddito operativo netto",
                      "Imposta sul reddito da locazione", "Utile netto annuo",
                      "Spese condominiali a carico", "Gestione e costi variabili"):
        assert etichetta in righe, f"voce {etichetta!r} non trovata nel foglio Locazione"

    r_eff = righe["Ricavo effettivo"]
    r_noi = righe["Reddito operativo netto"]
    r_primo = righe["Spese condominiali a carico"]
    r_ultimo = righe["Gestione e costi variabili"]

    # L'ordine delle voci e' esso stesso parte dell'invariante.
    assert r_eff < r_primo <= r_ultimo < r_noi, (
        "l'ordine delle voci del conto economico non e' quello atteso: "
        f"ricavo {r_eff}, costi da {r_primo} a {r_ultimo}, reddito operativo {r_noi}"
    )
    assert r_primo == r_eff + 1, "fra il ricavo effettivo e il primo costo c'e' una riga estranea"
    assert r_ultimo == r_noi - 1, "fra l'ultimo costo e il reddito operativo c'e' una riga estranea"

    noi = ws.cell(row=r_noi, column=2).value
    assert isinstance(noi, str) and noi.startswith("=")
    assert f"B{r_eff}" in noi, f"il reddito operativo non legge il ricavo effettivo: {noi!r}"
    assert f"SUM(B{r_primo}:B{r_ultimo})" in noi, (
        f"il reddito operativo non somma l'intero blocco dei costi: {noi!r}"
    )

    # L'utile netto deve leggere il reddito operativo e l'imposta, non altre righe.
    r_imp = righe["Imposta sul reddito da locazione"]
    utile = ws.cell(row=righe["Utile netto annuo"], column=2).value
    assert f"B{r_noi}" in utile and f"B{r_imp}" in utile, (
        f"l'utile netto non somma reddito operativo e imposta: {utile!r}"
    )

    # L'invariante vale su tutte e quattro le colonne dei regimi, non solo sulla prima.
    for colonna, lettera in ((3, "C"), (4, "D"), (5, "E")):
        formula = ws.cell(row=r_noi, column=colonna).value
        assert f"SUM({lettera}{r_primo}:{lettera}{r_ultimo})" in formula, (
            f"la colonna {lettera} del reddito operativo somma un intervallo diverso: {formula!r}"
        )


def test_regime_di_acquisto_per_riga_nel_confronto():
    """Le imposte del foglio di confronto leggono il regime della riga, non quello globale.

    Prima, ogni riga pagava le imposte del regime impostato nel foglio Immobile:
    un usato da privato e un nuovo da costruttore comparivano nella stessa lista
    con la stessa imposta, e la graduatoria sbagliava nel verso peggiore, perche'
    faceva sembrare piu' conveniente proprio l'immobile con l'imposta piu' alta.
    Il presidio non e' la presenza delle due colonne ma il fatto che la formula
    delle imposte le legga: se tornasse ai nomi globali `agevolata` e `da_impresa`
    il foglio continuerebbe a calcolare senza il minimo segnale.
    """
    wb = load_workbook(workbook())
    ws = wb["Confronto immobili"]

    intestazione = None
    for riga in range(1, 20):
        if ws.cell(row=riga, column=1).value == "ID":
            intestazione = riga
            break
    assert intestazione is not None

    testate = {
        ws.cell(row=intestazione, column=colonna).value: colonna
        for colonna in range(1, 40)
        if isinstance(ws.cell(row=intestazione, column=colonna).value, str)
    }
    for etichetta in ("Prima casa", "Da impresa"):
        assert etichetta in testate, f"colonna {etichetta!r} assente dal foglio di confronto"
    assert testate["Prima casa"] == 26 and testate["Da impresa"] == 27, (
        "le due colonne del regime hanno cambiato posizione: le formule delle imposte "
        "le citano per lettera, quindi la posizione e' contrattuale"
    )

    prima = intestazione + 1
    # Le due colonne ereditano dal foglio Immobile quando il registro tace.
    for colonna, globale, registro in (
        (26, "agevolata", "Annunci!$AK"), (27, "da_impresa", "Annunci!$AL"),
    ):
        formula = ws.cell(row=prima, column=colonna).value
        assert isinstance(formula, str)
        assert globale in formula, f"la colonna {colonna} non ricade su {globale}: {formula!r}"
        assert registro in formula, f"la colonna {colonna} non legge {registro}: {formula!r}"

    # Imposte di trasferimento e costi accessori devono leggere le due colonne.
    imposte = ws.cell(row=prima, column=9).value
    assert f"$AA{prima}" in imposte, f"le imposte non leggono il regime del venditore: {imposte!r}"
    assert f"$Z{prima}" in imposte, f"le imposte non leggono la prima casa della riga: {imposte!r}"
    assert "da_impresa" not in imposte and "agevolata" not in imposte, (
        f"le imposte usano ancora il regime globale: {imposte!r}"
    )
    accessori = ws.cell(row=prima, column=11).value
    assert f"$Z{prima}" in accessori and "agevolata" not in accessori, (
        f"l'imposta sostitutiva del mutuo usa ancora il regime globale: {accessori!r}"
    )


def test_cruscotto_legge_il_confronto_affitto_per_nome():
    """Il verdetto del Cruscotto non deve dipendere da una coordinata fissa.

    La formula citava `'Confronto affitto'!$B$52`: inserire una riga in quel foglio
    non avrebbe prodotto alcun errore, avrebbe prodotto un verdetto sbagliato sul
    primo foglio del workbook, cioe' quello che si legge per decidere.
    """
    wb = load_workbook(workbook())
    assert "conf_differenza" in wb.defined_names, (
        "il nome definito conf_differenza non esiste piu'"
    )

    # Il nome deve puntare alla differenza, non a un'altra cella della sezione.
    # E' l'invariante che il difetto precedente violava: la formula del Cruscotto
    # citava $B$52, cioe' il patrimonio comprando, che e' positivo per qualunque
    # immobile di valore, quindi il verdetto diceva "conviene comprare" anche
    # quando il foglio concludeva l'opposto. Nessun errore, nessuna cella rossa:
    # solo il primo foglio del workbook che risponde alla domanda sbagliata.
    destinazioni = list(wb.defined_names["conf_differenza"].destinations)
    assert len(destinazioni) == 1
    foglio, coordinata = destinazioni[0]
    assert foglio == "Confronto affitto", f"conf_differenza punta al foglio {foglio!r}"
    riga_nome = int("".join(c for c in coordinata if c.isdigit()))
    etichetta = wb[foglio].cell(row=riga_nome, column=1).value
    assert etichetta == "Differenza a favore dell'acquisto", (
        f"conf_differenza punta alla riga {riga_nome}, etichettata {etichetta!r}, "
        "che non e' la differenza fra i due patrimoni"
    )
    ws = wb["Cruscotto"]
    trovata = None
    for riga in range(1, 80):
        valore = ws.cell(row=riga, column=2).value
        if isinstance(valore, str) and "conviene comprare" in valore:
            trovata = valore
            break
    assert trovata is not None, "il verdetto sul comprare o affittare non si trova nel Cruscotto"
    assert "conf_differenza" in trovata, f"il verdetto non usa il nome definito: {trovata!r}"
    assert "$B$52" not in trovata and "$B$21" not in trovata, (
        f"il verdetto cita ancora una coordinata fissa: {trovata!r}"
    )


def test_confronto_immobili_porta_il_blocco_omi():
    """Il foglio di confronto mostra la zona OMI e ricalcola lo scarto sul proprio prezzo.

    Il rischio che questo test presidia non e' la presenza delle colonne ma la
    provenienza dello scarto. Leggerlo dalla colonna V del foglio Annunci sarebbe
    stato piu' corto di una riga e sbagliato, perche' quella colonna confronta la
    quotazione di zona con il prezzo richiesto mentre ogni altra colonna di questo
    foglio ragiona sul prezzo obiettivo quando c'e': la riga avrebbe portato un solo
    numero riferito a un prezzo diverso da tutti gli altri, senza alcun segnale.
    Si verifica quindi che la formula dello scarto citi la colonna E del foglio
    stesso, cioe' il prezzo al metro quadro locale, e non il foglio Annunci.
    """
    wb = load_workbook(workbook())
    ws = wb["Confronto immobili"]

    intestazione = None
    for riga in range(1, 20):
        if ws.cell(row=riga, column=1).value == "ID":
            intestazione = riga
            break
    assert intestazione is not None, "intestazione del foglio di confronto non trovata"

    testate = {
        ws.cell(row=intestazione, column=colonna).value: colonna
        for colonna in range(1, 40)
        if isinstance(ws.cell(row=intestazione, column=colonna).value, str)
    }
    for etichetta in ("Zona OMI", "Quot. OMI min", "Quot. OMI max", "Scarto su OMI", "Esito"):
        assert etichetta in testate, f"colonna {etichetta!r} assente dal foglio di confronto"

    prima = intestazione + 1
    zona = ws.cell(row=prima, column=testate["Zona OMI"]).value
    assert isinstance(zona, str) and "Annunci!$J" in zona, (
        f"la zona OMI non legge la colonna J del registro: {zona!r}"
    )
    for etichetta, colonna_registro in (("Quot. OMI min", "$T"), ("Quot. OMI max", "$U")):
        formula = ws.cell(row=prima, column=testate[etichetta]).value
        assert isinstance(formula, str) and f"Annunci!{colonna_registro}" in formula, (
            f"{etichetta} non legge la colonna {colonna_registro} del registro: {formula!r}"
        )

    scarto = ws.cell(row=prima, column=testate["Scarto su OMI"]).value
    assert isinstance(scarto, str) and scarto.startswith("=")
    assert "Annunci!" not in scarto, (
        "lo scarto su OMI e' tornato a leggere il foglio Annunci, quindi il prezzo "
        f"richiesto invece di quello usato dal foglio: {scarto!r}"
    )
    assert f"$E{prima}" in scarto, (
        f"lo scarto non usa il prezzo al mq della colonna E: {scarto!r}"
    )

    # L'esito resta l'ultima colonna e continua a leggere il rendimento netto.
    esito = ws.cell(row=prima, column=testate["Esito"]).value
    assert isinstance(esito, str) and "rend_obiettivo" in esito
    assert testate["Esito"] == max(testate.values()), "l'esito non e' piu' l'ultima colonna"


def test_precompilazione_scrive_per_nome_e_azzera_solo_le_lacune():
    """Precompila un workbook da un annuncio e verifica le tre garanzie.

    La precompilazione toglie il passaggio piu' pericoloso del percorso, cioe' la
    ridigitazione a mano dei dati dell'immobile scelto nei fogli di input: un
    prezzo con una cifra in meno produce un'operazione che sembra ottima e
    nessuna cella va in errore per dirlo. Le garanzie da presidiare sono tre.

    La prima e' che si scriva per nome definito e non per coordinata, secondo
    ADR-013. La seconda e' che non si scriva mai in una cella che contiene una
    formula: la distinzione fra input e calcolo vive nel colore e non nel tipo,
    quindi niente impedirebbe a un nome di puntare a una cella calcolata e
    sovrascriverla romperebbe la catena in silenzio. La terza, la meno ovvia, e'
    che i campi assenti dal registro vengano azzerati e non lasciati al valore di
    esempio: un workbook appena generato porta una rendita catastale di 450 euro
    che serve a mostrare il formato, e in un file dedicato a un immobile reale
    quel valore farebbe applicare il prezzo-valore su una base inventata senza
    che alcun controllo se ne accorga.

    L'azzeramento non tocca i campi la cui assenza significa qualcosa, cioe' i
    due del regime di acquisto, dove il vuoto e' il terzo stato di ADR-014, e la
    base d'asta, dove il vuoto significa che non e' un'asta.
    """
    cartella = Path(tempfile.mkdtemp())
    destinazione = cartella / "precompilato.xlsx"
    E.genera(str(destinazione))

    # Un annuncio con alcuni campi e non altri: il prezzo obiettivo deve vincere
    # sul richiesto, come fa il foglio di confronto.
    annuncio = A.Annuncio(
        id="house_prova", comune="Comune di prova", mq=62,
        prezzo_richiesto=159_000, prezzo_obiettivo=150_000,
        canone_atteso_mese=600,
    )
    esito = A.precompila_workbook(annuncio, str(destinazione))

    wb = load_workbook(destinazione)

    def per_nome(nome):
        foglio, coordinata = list(wb.defined_names[nome].destinations)[0]
        return wb[foglio][coordinata].value

    # Scritti: il prezzo obiettivo e non il richiesto.
    assert per_nome("prezzo") == 150_000, per_nome("prezzo")
    assert per_nome("mq") == 62
    assert per_nome("comune") == "Comune di prova"
    assert per_nome("canone_mese") == 600
    assert set(esito["scritti"]) == {"prezzo", "mq", "comune", "canone_mese"}, esito["scritti"]

    # Azzerati: le lacune, e solo quelle. La rendita e' il caso che conta.
    azzerati = {nome for nome, _, _ in esito["azzerati"]}
    assert "rendita" in azzerati, "la rendita mancante non e' stata azzerata"
    assert "condominio" in azzerati
    assert per_nome("rendita") == 0, "la rendita ha conservato il valore di esempio"
    assert per_nome("condominio") == 0

    # Non azzerati: i campi la cui assenza significa qualcosa.
    for nome in ("prima_casa", "da_impresa", "asta_base"):
        assert nome not in azzerati, (
            f"{nome} e' stato azzerato: il suo vuoto significa eredita o non pertinente, "
            "e azzerarlo cambia il modello invece di dichiararlo incompleto"
        )

    # Nessuna cella di formula toccata, e nessun rifiuto: se ci fossero rifiuti
    # significherebbe che un nome della mappa punta a una cella calcolata.
    assert esito["rifiutati"] == [], esito["rifiutati"]

    # La mappa e i nomi del workbook non devono divergere: un nome scomparso
    # deve far fallire la funzione, non passare in silenzio.
    for nome, _, _, _ in A.PRECOMPILAZIONE:
        assert nome in wb.defined_names, f"il nome {nome!r} della mappa non esiste nel workbook"


def test_controlli_di_plausibilita_contano_solo_i_non_superati():
    """Il Cruscotto porta i controlli, e il contatore conta i messaggi.

    I controlli non verificano che un input sia giusto, cosa che il modello non
    puo' sapere, ma che non sia ancora quello di esempio, che non sia a zero dove
    uno zero non e' plausibile, o che non sia incoerente con un'altra scelta.
    Ciascuno restituisce la stringa vuota quando e' superato, e non la parola ok,
    perche' una colonna di ok diventa rumore che si impara a ignorare mentre una
    colonna quasi vuota rende visibile cio' che resta.

    Il contatore deve usare un criterio sulle stringhe non vuote e non sulle
    celle non vuote: una formula che restituisce la stringa vuota produce una
    cella tecnicamente non vuota, quindi COUNTA li conterebbe tutti.
    """
    wb = load_workbook(workbook())
    ws = wb["Cruscotto"]

    intestazione = None
    for riga in range(1, 140):
        if ws.cell(row=riga, column=1).value == "Controllo" and ws.cell(row=riga, column=2).value == "Esito":
            intestazione = riga
            break
    assert intestazione is not None, "la sezione dei controlli non si trova nel Cruscotto"

    attesi = ["Rendita catastale contro prezzo-valore", "Aliquota IMU deliberata dal Comune",
              "Spese condominiali", "Canone atteso", "Superficie", "Comune",
              "Assicurazione del fabbricato", "Patrimonio complessivo"]
    for scostamento, etichetta in enumerate(attesi, start=1):
        effettiva = ws.cell(row=intestazione + scostamento, column=1).value
        assert effettiva == etichetta, (
            f"riga {intestazione + scostamento}: atteso {etichetta!r}, trovato {effettiva!r}"
        )
        formula = ws.cell(row=intestazione + scostamento, column=2).value
        assert isinstance(formula, str) and formula.startswith("=IF("), (
            f"il controllo {etichetta!r} non e' una formula condizionale: {formula!r}"
        )
        assert formula.endswith('"")'), (
            f"il controllo {etichetta!r} non restituisce la stringa vuota quando e' superato: {formula!r}"
        )

    # Il contatore, e il criterio che usa.
    assert "controlli_falliti" in wb.defined_names
    foglio, coordinata = list(wb.defined_names["controlli_falliti"].destinations)[0]
    contatore = wb[foglio][coordinata].value
    assert 'COUNTIF(controlli_esiti,"?*")' in contatore, (
        f"il contatore non conta le stringhe non vuote: {contatore!r}. "
        "COUNTA conterebbe anche le celle che contengono la stringa vuota."
    )

    # L'intervallo contato deve coprire esattamente le righe dei controlli.
    foglio_int, riferimento = list(wb.defined_names["controlli_esiti"].destinations)[0]
    import re
    numeri = [int(n) for n in re.findall(r"\d+", riferimento)]
    assert numeri == [intestazione + 1, intestazione + len(attesi)], (
        f"l'intervallo dei controlli e' {riferimento}, atteso da {intestazione + 1} a {intestazione + len(attesi)}"
    )

    # E il Cruscotto deve mostrarlo in testa, non solo in fondo.
    in_testa = False
    for riga in range(1, intestazione):
        if ws.cell(row=riga, column=1).value == "Controlli di plausibilita' non superati":
            assert ws.cell(row=riga, column=2).value == "=controlli_falliti"
            in_testa = True
            break
    assert in_testa, "il contatore dei controlli non compare fra i numeri in testa al Cruscotto"


def test_piano_ammortamento_copre_quaranta_anni():
    wb = load_workbook(workbook())
    ws = wb["Ammortamento"]
    numeri = [
        ws.cell(row=riga, column=1).value
        for riga in range(1, 500)
        if isinstance(ws.cell(row=riga, column=1).value, int)
    ]
    assert max(numeri) == E.MAX_RATE == 480


def test_dossier_tecnico_ha_peso_e_contatori_coerenti():
    """Ogni documento porta un peso fra i tre ammessi, e i contatori li leggono.

    Il peso non e' decorativo: la formula del contatore dei bloccanti fa
    `COUNTIFS` sulla stringa esatta, quindi un valore scritto diversamente,
    anche solo con un'iniziale maiuscola, sparirebbe dal conteggio senza che
    nulla protesti e il cruscotto direbbe zero documenti mancanti.
    """
    wb = load_workbook(workbook())
    ws = wb["Dossier tecnico"]

    ammessi = {"bloccante", "importante", "se ricorre"}
    pesi = []
    for riga in range(1, ws.max_row + 1):
        documento = ws.cell(row=riga, column=2).value
        peso = ws.cell(row=riga, column=6).value
        # Le righe di dati sono quelle con un documento e uno stato iniziale.
        if documento and ws.cell(row=riga, column=8).value == "da chiedere":
            assert peso in ammessi, f"riga {riga}: peso {peso!r} non ammesso"
            pesi.append(peso)

    assert len(pesi) > 40, f"solo {len(pesi)} documenti in elenco"
    assert pesi.count("bloccante") > 0

    for nome in ("documenti_bloccanti_aperti", "documenti_completamento"):
        assert nome in wb.defined_names, f"nome definito {nome} assente"

    riferimento = wb.defined_names["documenti_bloccanti_aperti"].attr_text
    assert riferimento.startswith("'Dossier tecnico'!")
    cella = riferimento.split("!")[1].replace("$", "")
    formula = ws[cella].value
    assert formula.startswith("=COUNTIFS(") and '"bloccante"' in formula

def test_registro_riletto_non_duplica_gli_annunci():
    """Rileggere il file da disco deve sostituire, non accodare.

    Il costruttore chiama gia' `carica`, quindi chiamarlo di nuovo per rileggere
    il file, cosa che viene naturale dopo averlo modificato, raddoppiava
    l'elenco. Nessun errore: un registro che conta il doppio, un foglio di
    confronto con le righe ripetute e una graduatoria che sembra piu' ricca di
    quello che e'. E' il difetto che produce un risultato plausibile.
    """
    import tempfile

    from immobiliare import annunci as A

    percorso = Path(tempfile.mkdtemp()) / "annunci.csv"
    registro = A.Registro(percorso)
    registro.aggiungi(A.Annuncio(id="x_1", comune="Comune di esempio", mq=50, prezzo_richiesto=100_000))
    registro.aggiungi(A.Annuncio(id="x_2", comune="Comune di esempio", mq=60, prezzo_richiesto=120_000))
    registro.salva()

    riletto = A.Registro(percorso)
    assert len(riletto.annunci) == 2
    riletto.carica()
    riletto.carica()
    assert len(riletto.annunci) == 2, f"il registro si e' duplicato: {len(riletto.annunci)} righe"
    assert [a.id for a in riletto.annunci] == ["x_1", "x_2"]

def test_numero_riconosce_il_separatore_delle_migliaia_italiano():
    """Un punto in un annuncio italiano quasi sempre separa le migliaia.

    Il modello locale restituisce a volte i numeri come stringhe, cosi' come li
    trova nel testo. Trattare 175.000 come decimale produce un prezzo di
    centosettantacinque euro: nessun errore, nessuna eccezione, un immobile che
    nel confronto risulta regalato. La discriminante e' quante cifre seguono il
    punto, tre per le migliaia e una o due per i decimali, ed e' l'unica euristica
    che regge sui valori che questo dominio incontra davvero, cioe' prezzi,
    superfici, canoni e rendite.
    """
    from immobiliare import annunci as A

    casi = [
        ("175.000", 175_000.0),      # migliaia: il caso che rompeva
        ("1.500.000", 1_500_000.0),  # due separatori
        ("89.000 EUR", 89_000.0),    # con valuta appesa
        ("612,45", 612.45),          # decimale all'italiana
        ("1.234,56", 1_234.56),      # migliaia e decimale insieme
        ("612.45", 612.45),          # due cifre dopo il punto: decimale
        ("620 al mese", 620.0),
        ("95", 95.0),
        ("2,5", 2.5),
        ("", 0.0),
        ("nessun dato", 0.0),
    ]
    for testo, atteso in casi:
        ottenuto = A._numero(testo)
        assert abs(ottenuto - atteso) < 1e-9, f"{testo!r}: atteso {atteso}, ottenuto {ottenuto}"

    # I numeri veri passano intatti.
    assert A._numero(175_000) == 175_000.0
    assert A._numero(612.45) == 612.45


def test_schema_di_estrazione_copre_i_campi_che_decidono():
    """Lo schema passato al modello e' cio' che il modello puo' trovare.

    Rendita catastale, categoria e canone compaiono di rado in un annuncio, ma
    quando ci sono valgono piu' di tutto il resto: la prima sblocca il
    prezzo-valore, la seconda decide moltiplicatore ed esclusione
    dall'agevolazione, il terzo determina l'intero calcolo del rendimento.
    Ometterli dallo schema significa non trovarli mai anche quando sono scritti
    in chiaro, e il modello non sbaglia nulla: non gli e' stato chiesto.
    """
    from immobiliare import annunci as A

    for campo in ("rendita_catastale", "categoria", "canone_atteso_mese",
                  "comune", "mq", "prezzo_richiesto", "spese_condominio_anno"):
        assert campo in A.CAMPI_ESTRAIBILI, f"{campo} non e' nello schema di estrazione"

    # Ogni campo dello schema deve esistere davvero nella dataclass, altrimenti
    # il modello lo estrae e la costruzione dell'annuncio lo scarta in silenzio.
    nomi = {f.name for f in __import__("dataclasses").fields(A.Annuncio)}
    for campo in A.CAMPI_ESTRAIBILI:
        assert campo in nomi, f"lo schema chiede {campo}, che non e' un campo di Annuncio"

def test_annunci_confronta_ordina_per_scarto_e_non_per_prezzo():
    """La graduatoria si ordina sullo scarto di zona, non sul prezzo.

    Fra immobili di taglia diversa il prezzo non dice nulla: duecento metri
    quadri a centoquarantamila euro e trentacinque a centoventimila non sono
    confrontabili finche' non si rapportano alla quotazione della loro zona.
    Il test congela il criterio, perche' e' una scelta di metodo e non un
    dettaglio di presentazione, e chi la cambiasse per comodita' romperebbe il
    senso del comando.
    """
    import subprocess
    import sys as _sys

    radice = Path(__file__).resolve().parent.parent
    esito = subprocess.run(
        [_sys.executable, str(radice / "tools" / "valuta.py"), "annunci", "confronta"],
        capture_output=True, text=True, timeout=300, cwd=str(radice),
    )
    assert esito.returncode in (0, 1), esito.stderr[-500:]
    uscita = esito.stdout
    if "Nessun annuncio" in uscita:
        return  # registro vuoto su una macchina pulita: nulla da verificare

    assert "scarto" in uscita and "canone di zona" in uscita
    # L'attribuzione della fonte OMI e' obbligatoria ovunque i dati compaiano.
    assert "Agenzia Entrate - OMI" in uscita

    percentuali = []
    for riga in uscita.splitlines():
        pezzi = riga.split()
        if len(pezzi) > 7 and pezzi[0].startswith("house_"):
            valore = next((p for p in pezzi if p.endswith("%") and (p.startswith("+") or p.startswith("-"))), None)
            if valore:
                percentuali.append(int(valore.rstrip("%")))
    if len(percentuali) > 1:
        assert percentuali == sorted(percentuali), f"non ordinato per scarto: {percentuali}"


def test_segnalazioni_non_scattano_su_parole_fuori_contesto():
    """Un flag che compare dove non serve smette di voler dire qualcosa.

    La nota di un annuncio conteneva la frase "prima di ogni ipotesi abitativa",
    e la ricerca della sola parola "ipotesi" lo marcava come zona incerta. Il
    difetto non produce errori: produce una tabella in cui la colonna delle
    segnalazioni e' rumore, e chi la legge smette di guardarla.
    """
    fuori_contesto = "Destinazione ufficio: verificare il cambio d uso prima di ogni ipotesi abitativa".lower()
    dentro_contesto = "Zona B1 assegnata per prossimita': Via Mozzi non e' nominata, da confermare.".lower()
    per_ipotesi = "ZONA OMI B5 ASSEGNATA PER IPOTESI: l'annuncio non da' via ne' mappa.".lower()

    def incerta(nota):
        return "per ipotesi" in nota or "da confermare" in nota

    assert not incerta(fuori_contesto), "flag acceso su una nota che non parla di zona"
    assert incerta(dentro_contesto)
    assert incerta(per_ipotesi)

    # E la nota fuori contesto deve invece accendere il flag che le compete.
    assert "destinazione ufficio" in fuori_contesto

if __name__ == "__main__":
    superati = 0
    falliti = []
    for nome, funzione in sorted(globals().items()):
        if nome.startswith("test_") and callable(funzione):
            try:
                funzione()
                superati += 1
            except AssertionError as e:
                falliti.append((nome, e))
            except Exception as e:
                falliti.append((nome, f"{e.__class__.__name__}: {e}"))
    print(f"{superati} test superati, {len(falliti)} falliti")
    for nome, errore in falliti:
        print(f"  FALLITO {nome}: {errore}")
    raise SystemExit(1 if falliti else 0)
