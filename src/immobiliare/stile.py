# -*- coding: utf-8 -*-
"""Stili condivisi del workbook.

Tenere gli stili in un modulo a parte evita di ripetere le stesse definizioni in
ogni foglio e rende il cambio di palette un'operazione in un punto solo.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.utils import get_column_letter

# Palette: blu profondo per le intestazioni, giallo tenue per cio' che si edita,
# verde tenue per i risultati, grigio per il calcolato non editabile.
BLU = "1F3864"
BLU_CHIARO = "2E5395"
GIALLO = "FFF2CC"
AZZURRO = "DDEBF7"
VERDE = "E2EFDA"
GRIGIO = "F2F2F2"
ROSSO = "FCE4EC"
BIANCO = "FFFFFF"

TITOLO = Font(name="Calibri", size=16, bold=True, color=BLU)
SOTTOTITOLO = Font(name="Calibri", size=11, italic=True, color="595959")
SEZIONE = Font(name="Calibri", size=12, bold=True, color=BIANCO)
ETICHETTA = Font(name="Calibri", size=11)
ETICHETTA_BOLD = Font(name="Calibri", size=11, bold=True)
VALORE = Font(name="Calibri", size=11)
NOTA = Font(name="Calibri", size=9, italic=True, color="7F7F7F")
LINK = Font(name="Calibri", size=11, color="0563C1", underline="single")
KPI = Font(name="Calibri", size=14, bold=True, color=BLU)

FILL_SEZIONE = PatternFill("solid", fgColor=BLU)
FILL_SEZIONE_2 = PatternFill("solid", fgColor=BLU_CHIARO)
FILL_INPUT = PatternFill("solid", fgColor=GIALLO)
FILL_SCELTA = PatternFill("solid", fgColor=AZZURRO)
FILL_RISULTATO = PatternFill("solid", fgColor=VERDE)
FILL_CALCOLO = PatternFill("solid", fgColor=GRIGIO)
FILL_ATTENZIONE = PatternFill("solid", fgColor=ROSSO)

_sottile = Side(style="thin", color="BFBFBF")
BORDO = Border(left=_sottile, right=_sottile, top=_sottile, bottom=_sottile)

EURO = '#,##0 "€"'
EURO_DEC = '#,##0.00 "€"'
PERC = "0.00%"
PERC_1 = "0.0%"
NUMERO = "#,##0"
NUMERO_DEC = "#,##0.00"
DATA = "dd/mm/yyyy"

SINISTRA = Alignment(horizontal="left", vertical="center", wrap_text=True)
DESTRA = Alignment(horizontal="right", vertical="center")
CENTRO = Alignment(horizontal="center", vertical="center", wrap_text=True)


FOGLIO_INDICE = "Guida"
"""Nome del foglio che fa da indice. Vive qui perche' `titolo` scrive il ritorno."""

USI: dict = {}
"""Che cosa si fa in ciascun foglio, per nome del foglio.

Lo riempie il costruttore prima di creare i fogli, dalla stessa tupla da cui
nasce l'indice, e lo legge `titolo` per scrivere la fascia d'uso in testa a ogni
foglio. La forma e' un dizionario di modulo e non un parametro di `titolo` per
una ragione pratica: `titolo` e' chiamata da venti metodi diversi, e passare
l'informazione a ciascuno avrebbe richiesto venti modifiche con venti occasioni
di dimenticarne una, mentre il registro fa comparire la fascia su ogni foglio che
sia dichiarato nell'indice, e su nessun altro.
"""


def scelta(validazione, cella):
    """Aggancia una cella a un elenco di valori ammessi e la colora di conseguenza.

    E' l'unico modo previsto di applicare una tendina, e la ragione e' una
    segnalazione d'uso: prima le celle da scegliere erano gialle come quelle da
    digitare, quindi a video non c'era nulla che distinguesse una cella dove si
    scrive un numero da una dove si sceglie fra valori ammessi. Chi non conosce il
    file ci scriveva dentro, la validazione rifiutava il valore e il messaggio di
    Excel non spiegava perche'. Passando sempre da qui, una cella con tendina non
    puo' avere il colore sbagliato.
    """
    validazione.add(cella)
    cella.fill = FILL_SCELTA
    return cella


def collegamento(ws, riga: int, colonna: int, testo: str, foglio: str, cella: str = "A1"):
    """Scrive una cella che porta a un altro foglio dello stesso workbook.

    Il collegamento va costruito come `Hyperlink` con `location` e non assegnando
    una stringa a `cell.hyperlink`, perche' quella scorciatoia lo registra come
    destinazione esterna: nel file finisce fra le relazioni verso l'esterno, e a
    seconda della versione Excel lo apre chiedendo conferma o lo segnala come
    riferimento non attendibile. Un collegamento interno non ha una destinazione
    esterna, ha una posizione dentro il file, ed e' esattamente cio' che
    `location` esprime.

    Il nome del foglio va fra apici singoli perche' diversi fogli di questo
    workbook hanno uno spazio nel nome, e senza apici la destinazione non si
    risolve. E si punta a una cella e non al foglio nudo, perche' un collegamento
    senza cella lascia il foglio dove era stato lasciato scrollato, che su un
    piano di ammortamento di 480 righe significa arrivarci in fondo.
    """
    c = ws.cell(row=riga, column=colonna, value=testo)
    c.font = LINK
    c.alignment = SINISTRA
    c.hyperlink = Hyperlink(ref=c.coordinate, location=f"'{foglio}'!{cella}")
    return c


def titolo(ws, riga: int, testo: str, sottotitolo: str = "", larghezza: int = 8) -> int:
    """Scrive il titolo del foglio e restituisce la prima riga libera.

    Scrive anche, appena a destra del titolo, il collegamento di ritorno
    all'indice. Sta qui e non nei singoli fogli per una ragione di manutenzione:
    ogni foglio del workbook comincia chiamando questa funzione, quindi metterlo
    qui garantisce che nessun foglio nuovo nasca senza via di ritorno. Un
    workbook di venti fogli in cui la navigazione dipende dalla linguetta in
    basso e' navigabile solo da chi lo ha costruito.
    """
    c = ws.cell(row=riga, column=1, value=testo)
    c.font = TITOLO
    ws.merge_cells(start_row=riga, start_column=1, end_row=riga, end_column=larghezza)
    ws.row_dimensions[riga].height = 24
    riga += 1
    if sottotitolo:
        c = ws.cell(row=riga, column=1, value=sottotitolo)
        c.font = SOTTOTITOLO
        c.alignment = SINISTRA
        ws.merge_cells(start_row=riga, start_column=1, end_row=riga, end_column=larghezza)
        ws.row_dimensions[riga].height = 30
        riga += 1
    # La riga che questa funzione lasciava vuota fra il titolo e il primo
    # contenuto porta due cose: il ritorno all'indice in colonna A, e da colonna B
    # la fascia che dice se in questo foglio si scrive o si legge. La posizione e'
    # la stessa su tutti i fogli: metterle accanto al titolo le avrebbe spostate a
    # destra di tante colonne quante ne occupa il titolo, che varia da quattro a
    # ventisei, e su un foglio largo sarebbero finite fuori dalla vista.
    if ws.title != FOGLIO_INDICE:
        collegamento(ws, riga, 1, "<< Indice", FOGLIO_INDICE)
        uso = USI.get(ws.title)
        if uso:
            azione, quando, esito = uso
            if azione == "Si legge":
                testo = ("QUI NON SI SCRIVE NULLA: tutto quello che vedi e' calcolato negli altri fogli. "
                         f"Si apre {quando.lower()}. Ne esce: {esito.lower()}.")
                riempimento = FILL_CALCOLO
            elif azione == "Si consulta":
                testo = ("QUI SI CONSULTA: si apre per capire da dove viene un numero, e si modifica solo "
                         f"per un aggiornamento normativo. Si apre {quando.lower()}. Contiene: {esito.lower()}.")
                riempimento = FILL_CALCOLO
            else:
                testo = ("QUI SI SCRIVE: compila le celle GIALLE, scegli dall'elenco nelle celle AZZURRE, "
                         "non toccare le grigie e le verdi che sono calcolate. "
                         f"Si apre {quando.lower()}. Ne esce: {esito.lower()}.")
                riempimento = FILL_INPUT
            c = ws.cell(row=riga, column=2, value=testo)
            c.font = ETICHETTA_BOLD
            c.alignment = SINISTRA
            for col in range(2, max(larghezza, 3) + 1):
                ws.cell(row=riga, column=col).fill = riempimento
                ws.cell(row=riga, column=col).border = BORDO
            ws.merge_cells(start_row=riga, start_column=2, end_row=riga, end_column=max(larghezza, 3))
            ws.row_dimensions[riga].height = 30
    return riga + 1


def sezione(ws, riga: int, testo: str, larghezza: int = 8, secondaria: bool = False) -> int:
    """Barra di sezione a tutta larghezza. Restituisce la riga successiva."""
    for col in range(1, larghezza + 1):
        c = ws.cell(row=riga, column=col)
        c.fill = FILL_SEZIONE_2 if secondaria else FILL_SEZIONE
        c.border = BORDO
    c = ws.cell(row=riga, column=1, value=testo)
    c.font = SEZIONE
    c.alignment = SINISTRA
    ws.row_dimensions[riga].height = 20
    return riga + 1


def campo(
    ws,
    riga: int,
    etichetta: str,
    valore=None,
    formato: str = None,
    nota: str = "",
    input_utente: bool = False,
    risultato: bool = False,
    colonna_valore: int = 2,
    colonna_nota: int = 3,
) -> int:
    """Riga a tre colonne: etichetta, valore, nota. Restituisce la riga successiva."""
    e = ws.cell(row=riga, column=1, value=etichetta)
    e.font = ETICHETTA_BOLD if risultato else ETICHETTA
    e.alignment = SINISTRA

    v = ws.cell(row=riga, column=colonna_valore, value=valore)
    v.font = VALORE
    v.border = BORDO
    v.alignment = DESTRA
    if formato:
        v.number_format = formato
    if input_utente:
        v.fill = FILL_INPUT
    elif risultato:
        v.fill = FILL_RISULTATO
        v.font = ETICHETTA_BOLD
    else:
        v.fill = FILL_CALCOLO

    if nota:
        n = ws.cell(row=riga, column=colonna_nota, value=nota)
        n.font = NOTA
        n.alignment = SINISTRA
    return riga + 1


def intestazioni(ws, riga: int, etichette, larghezze=None, colonna_iniziale: int = 1) -> int:
    """Riga di intestazione di una tabella."""
    for i, testo in enumerate(etichette):
        c = ws.cell(row=riga, column=colonna_iniziale + i, value=testo)
        c.font = SEZIONE
        c.fill = FILL_SEZIONE_2
        c.alignment = CENTRO
        c.border = BORDO
        if larghezze:
            ws.column_dimensions[get_column_letter(colonna_iniziale + i)].width = larghezze[i]
    ws.row_dimensions[riga].height = 30
    return riga + 1


def larghezze_colonne(ws, mappa: dict) -> None:
    """Imposta le larghezze da un dizionario lettera-larghezza."""
    for lettera, larghezza in mappa.items():
        ws.column_dimensions[lettera].width = larghezza


def nota_riga(ws, riga: int, testo: str, larghezza: int = 8) -> int:
    """Riga di nota esplicativa a tutta larghezza."""
    c = ws.cell(row=riga, column=1, value=testo)
    c.font = NOTA
    c.alignment = SINISTRA
    ws.merge_cells(start_row=riga, start_column=1, end_row=riga, end_column=larghezza)
    ws.row_dimensions[riga].height = max(15, 13 * (len(testo) // 110 + 1))
    return riga + 1
