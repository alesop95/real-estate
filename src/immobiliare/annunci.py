# -*- coding: utf-8 -*-
"""Registro degli annunci in valutazione, con acquisizione rispettosa delle regole.

Il modulo tiene un archivio CSV di annunci e sa arricchirlo in tre modi, in ordine
di preferenza decrescente. Il primo e' l'inserimento manuale, che non tocca nessun
sito. Il secondo e' l'incolla del testo dell'annuncio gia' aperto nel browser, che
viene strutturato in locale da un modello linguistico sulla rete di casa, quindi
senza che il contenuto esca dalla macchina. Il terzo, disattivato per impostazione
predefinita, e' il prelievo diretto della pagina, consentito solo quando il file
`robots.txt` del sito lo permette per il percorso richiesto.

Sul terzo punto la posizione del modulo e' deliberatamente restrittiva. Il
`robots.txt` viene letto e rispettato senza eccezioni, la frequenza e' limitata a
una richiesta ogni pochi secondi, lo user agent dichiara chi e' e a cosa serve, e
non esiste alcun meccanismo per aggirare protezioni anti bot: se un sito risponde
con un blocco, la risposta corretta e' fermarsi, non travestirsi. I dati raccolti
sono i soli attributi economici e tecnici dell'immobile, mai i dati di contatto
delle persone fisiche, che sono dati personali e il cui trattamento massivo
richiederebbe una base giuridica che qui non esiste.
"""

from __future__ import annotations

import csv
import json
import os
import time
import urllib.error
import urllib.parse
import urllib.request
import urllib.robotparser
from dataclasses import asdict, dataclass, field, fields
from datetime import date, datetime
from pathlib import Path

USER_AGENT = (
    "valutazione-immobiliare/1.0 (uso personale, non commerciale; "
    "rispetta robots.txt; contatto: proprietario dell'archivio locale)"
)
INTERVALLO_MINIMO_SECONDI = 5.0
TIMEOUT_SECONDI = 20

_ultima_richiesta: dict[str, float] = {}


@dataclass
class Annuncio:
    """Una riga del registro. I campi calcolati restano vuoti e li calcola Excel."""

    id: str = ""
    data: str = ""
    stato: str = "da contattare"
    fonte: str = ""
    agenzia: str = ""
    contatto: str = ""
    """Riferimento dell'agenzia o del venditore, inserito a mano.

    Non e' un dato raccolto dal prelievo automatico, che per scelta non estrae
    recapiti di persone fisiche: e' il contatto con cui si sta trattando, che chi
    compra annota comunque e che serve a sapere chi richiamare. La differenza fra
    annotare un recapito ricevuto e raccoglierne a strascico e' esattamente la
    differenza fra un'agenda e una banca dati.
    """
    link: str = ""
    comune: str = ""
    provincia: str = ""
    zona_omi: str = ""
    indirizzo: str = ""
    tipologia: str = ""
    destinazione_uso: str = ""
    """Classificazione ministeriale, cioe' la destinazione d'uso dichiarata.

    Un monolocale accatastato come ufficio non e' un'abitazione: cambia le imposte,
    la possibilita' di prendervi la residenza e quella di locarlo a uso abitativo,
    e il cambio di destinazione ha un costo che va messo nel prezzo.
    """
    nuova_costruzione: str = "NO"
    data_consegna: str = ""
    """Per le nuove costruzioni: data prevista di consegna, o 'pronto' se ultimato."""
    mq: float = 0.0
    prezzo_richiesto: float = 0.0
    prezzo_obiettivo: float = 0.0
    quotazione_omi_min: float = 0.0
    quotazione_omi_max: float = 0.0
    rendita_catastale: float = 0.0
    categoria: str = ""
    piano: str = ""
    classe_energetica: str = ""
    spese_condominio_anno: float = 0.0
    canone_atteso_mese: float = 0.0
    asta: str = "NO"
    """SI se l'immobile viene da una vendita giudiziaria.

    Cambia tutto cio' che sta a valle: non c'e' garanzia per i vizi, non c'e'
    provvigione di agenzia, il prezzo non e' trattabile ma si costruisce per
    rilanci, i gravami si cancellano per decreto e i tempi sono dettati dal
    tribunale. Un'asta valutata con il modello del libero mercato da' un numero
    che sembra ottimo e non lo e'.
    """
    base_asta: float = 0.0
    """Prezzo base d'asta. Non e' il prezzo che si paghera': e' il punto di partenza."""
    data_asta: str = ""
    tribunale_procedura: str = ""
    """Tribunale e numero di procedura, per esempio Macerata RGE 123/2024."""
    stato_occupazione: str = ""
    """Libero, occupato dal debitore, locato con contratto opponibile, occupato senza titolo.

    E' la variabile che separa un'asta conveniente da un contenzioso: l'articolo
    2923 del codice civile rende opponibile all'acquirente la locazione con data
    certa anteriore al pignoramento, e l'articolo 560 del codice di procedura
    lascia il debitore e i familiari conviventi nel possesso fino al decreto di
    trasferimento.
    """
    punteggio: int = 0
    """Priorita' da 0 a 10, assegnata a mano da chi valuta.

    Non e' un punteggio calcolato: il modello ha gia' metriche per quello, e
    sovrapporne una sintetica servirebbe solo a nascondere il ragionamento. E'
    l'ordine in cui si vogliono guardare gli immobili, che dipende anche da cose
    che il modello non sa, come la vicinanza a chi ci deve abitare o il fatto che
    il venditore abbia fretta. Dieci e' massima priorita', zero e' il valore di
    chi entra a registro senza averlo ancora deciso.
    """
    note: str = ""
    prima_casa: str = ""
    """SI, NO, oppure vuoto per ereditare il regime del foglio Immobile.

    Non e' una caratteristica dell'immobile ma della posizione di chi compra
    rispetto a quell'immobile: lo stesso appartamento e' prima casa per chi non
    ha altre abitazioni nel Comune e non lo e' per chi ce le ha, e la differenza
    vale qualche migliaio di euro di imposte. Sta nel registro perche' quando si
    confrontano piu' immobili la risposta puo' cambiare da riga a riga, tipico il
    caso di un immobile nel Comune di residenza accanto a uno fuori.

    Il vuoto e' un terzo stato e non un sinonimo di NO: significa che la riga non
    dichiara nulla e il foglio di confronto le applica il regime impostato nel
    foglio Immobile, cioe' esattamente il comportamento precedente. Un registro
    scritto prima che questo campo esistesse continua percio' a dare gli stessi
    numeri di prima.
    """
    venditore_impresa: str = ""
    """SI, NO, oppure vuoto per ereditare il regime del foglio Immobile.

    Discrimina l'acquisto soggetto a IVA da quello con l'imposta di registro, che
    e' il salto piu' grosso fra due righe della stessa lista: sullo stesso prezzo
    l'IVA si applica per intero mentre il registro con il prezzo-valore si applica
    al valore catastale, che di norma e' una frazione. Confrontare un usato da
    privato e un nuovo da costruttore senza questa distinzione produce una
    graduatoria sbagliata nel verso peggiore, perche' fa sembrare piu' conveniente
    proprio l'immobile che porta l'imposta piu' alta.
    """

    CAMPI_SI_NO = ("asta", "nuova_costruzione", "prima_casa", "venditore_impresa")
    """I campi che il workbook confronta con la stringa SI.

    Excel confronta il testo senza distinguere le maiuscole, quindi una cella che
    contiene si minuscolo si comporta correttamente. Non si comporta correttamente
    una cella che contiene true, yes oppure 1, che e' esattamente cio' che un
    modello linguistico restituisce quando gli si chiede un booleano: il foglio la
    legge come diversa da SI, cioe' come un NO, senza segnalare nulla. La
    normalizzazione esiste per questo, e lascia intatto quello che non riconosce,
    perche' un valore strano che resta visibile e' preferibile a un valore strano
    tradotto per ipotesi in una delle due risposte.
    """

    def __post_init__(self) -> None:
        if not self.data:
            self.data = date.today().isoformat()
        affermativi = {"si", "s", "yes", "y", "true", "vero", "1"}
        negativi = {"no", "n", "false", "falso", "0"}
        for campo in self.CAMPI_SI_NO:
            valore = str(getattr(self, campo) or "").strip()
            if not valore:
                setattr(self, campo, "")
            elif valore.casefold() in affermativi:
                setattr(self, campo, "SI")
            elif valore.casefold() in negativi:
                setattr(self, campo, "NO")

    @property
    def prezzo_mq(self) -> float:
        return self.prezzo_richiesto / self.mq if self.mq else 0.0

    @property
    def rendimento_lordo(self) -> float:
        return (
            self.canone_atteso_mese * 12 / self.prezzo_richiesto
            if self.prezzo_richiesto
            else 0.0
        )

    @property
    def scarto_su_omi(self) -> float:
        """Scarto del prezzo al metro quadro rispetto alla media della zona OMI."""
        if not (self.quotazione_omi_min and self.quotazione_omi_max and self.prezzo_mq):
            return 0.0
        media = (self.quotazione_omi_min + self.quotazione_omi_max) / 2
        return self.prezzo_mq / media - 1


class Registro:
    """Archivio CSV degli annunci, con lettura e scrittura idempotenti."""

    def __init__(self, percorso: str | Path = "data/annunci.csv") -> None:
        self.percorso = Path(percorso)
        self.percorso.parent.mkdir(parents=True, exist_ok=True)
        self.annunci: list[Annuncio] = []
        if self.percorso.exists():
            self.carica()

    @property
    def colonne(self) -> list[str]:
        return [f.name for f in fields(Annuncio)]

    def carica(self) -> None:
        # La lista si azzera prima di leggere. Il costruttore chiama gia' questo
        # metodo, quindi chi lo richiama per rileggere il file da disco, cosa
        # legittima e naturale, si ritroverebbe ogni annuncio due volte: nessun
        # errore, solo un registro che conta il doppio e un confronto fra immobili
        # con le righe duplicate.
        self.annunci = []
        # Il punto e virgola e' il separatore che Excel italiano apre senza chiedere
        # nulla, ed e' quindi quello con cui il file viene scritto. In lettura si
        # riconosce comunque anche la virgola, per i file arrivati da altrove.
        with self.percorso.open("r", encoding="utf-8-sig", newline="") as f:
            prima = f.readline()
            f.seek(0)
            delimitatore = ";" if prima.count(";") >= prima.count(",") else ","
            for riga in csv.DictReader(f, delimiter=delimitatore):
                pulita = {}
                for campo in fields(Annuncio):
                    valore = riga.get(campo.name, "")
                    if campo.type in ("float", float):
                        try:
                            pulita[campo.name] = float(str(valore).replace(",", ".") or 0)
                        except ValueError:
                            pulita[campo.name] = 0.0
                    elif campo.type in ("int", int):
                        try:
                            pulita[campo.name] = int(float(valore or 0))
                        except ValueError:
                            pulita[campo.name] = 0
                    else:
                        pulita[campo.name] = valore or ""
                self.annunci.append(Annuncio(**pulita))

    def salva(self) -> None:
        with self.percorso.open("w", encoding="utf-8-sig", newline="") as f:
            scrittore = csv.DictWriter(f, fieldnames=self.colonne, delimiter=";")
            scrittore.writeheader()
            for a in self.annunci:
                scrittore.writerow(asdict(a))

    def prossimo_id(self) -> str:
        numeri = []
        for a in self.annunci:
            if a.id.startswith("house_"):
                coda = a.id.removeprefix("house_")
                if coda.isdigit():
                    numeri.append(int(coda))
        return f"house_{max(numeri, default=0) + 1}"

    def aggiungi(self, annuncio: Annuncio) -> Annuncio:
        """Aggiunge, assegnando un identificativo se manca e rifiutando i doppioni."""
        if not annuncio.id:
            annuncio.id = self.prossimo_id()
        if annuncio.link:
            esistente = self.trova_per_link(annuncio.link)
            if esistente:
                raise ValueError(f"annuncio gia' presente con id {esistente.id}")
        self.annunci.append(annuncio)
        return annuncio

    def trova_per_link(self, link: str) -> Annuncio | None:
        normalizzato = link.split("?")[0].rstrip("/")
        for a in self.annunci:
            if a.link.split("?")[0].rstrip("/") == normalizzato:
                return a
        return None

    def trova(self, identificativo: str) -> Annuncio | None:
        for a in self.annunci:
            if a.id == identificativo:
                return a
        return None

    def rimuovi(self, identificativo: str) -> bool:
        prima = len(self.annunci)
        self.annunci = [a for a in self.annunci if a.id != identificativo]
        return len(self.annunci) < prima

    def ordina_per_convenienza(self) -> list[Annuncio]:
        """Ordina per rendimento lordo decrescente, a parita' di dati disponibili."""
        return sorted(self.annunci, key=lambda a: a.rendimento_lordo, reverse=True)


# ---------------------------------------------------------------------------
# Acquisizione rispettosa
# ---------------------------------------------------------------------------

class ProbitaRifiutata(Exception):
    """Il sito, tramite robots.txt, non consente il prelievo di quel percorso."""


def robots_consente(url: str, user_agent: str = USER_AGENT) -> tuple[bool, str]:
    """Verifica il robots.txt del sito per il percorso richiesto.

    Restituisce la coppia consenso e motivazione. Se il robots.txt non e'
    raggiungibile la risposta e' negativa: in assenza di un permesso esplicito il
    comportamento prudente e' astenersi, non presumere.
    """
    pezzi = urllib.parse.urlparse(url)
    if pezzi.scheme not in ("http", "https"):
        return False, "schema non supportato"
    robots = f"{pezzi.scheme}://{pezzi.netloc}/robots.txt"
    parser = urllib.robotparser.RobotFileParser()
    parser.set_url(robots)
    try:
        richiesta = urllib.request.Request(robots, headers={"User-Agent": user_agent})
        with urllib.request.urlopen(richiesta, timeout=TIMEOUT_SECONDI) as risposta:
            testo = risposta.read().decode("utf-8", errors="replace")
        parser.parse(testo.splitlines())
    except Exception as e:  # rete assente, 404, 403, qualunque cosa
        return False, f"robots.txt non leggibile ({e.__class__.__name__}): ci si astiene"
    if parser.can_fetch(user_agent, url):
        return True, "consentito dal robots.txt"
    return False, "escluso dal robots.txt del sito"


def _attendi_il_turno(dominio: str) -> None:
    """Impone l'intervallo minimo fra due richieste allo stesso dominio."""
    ultimo = _ultima_richiesta.get(dominio)
    if ultimo is not None:
        trascorso = time.monotonic() - ultimo
        if trascorso < INTERVALLO_MINIMO_SECONDI:
            time.sleep(INTERVALLO_MINIMO_SECONDI - trascorso)
    _ultima_richiesta[dominio] = time.monotonic()


class PrelievoBloccato(RuntimeError):
    """Il sito ha risposto con un blocco, e la risposta corretta e' fermarsi.

    E' distinta da `ProbitaRifiutata`, che riguarda il permesso dichiarato nel
    `robots.txt`, perche' i due casi si risolvono diversamente. Qui il permesso
    c'era e il server ha comunque negato, tipicamente con una protezione anti
    bot: non c'e' nulla da correggere nella richiesta, e insistere significherebbe
    aggirare la protezione, che il progetto non fa.
    """


def scarica_pagina(url: str, forza: bool = False) -> str:
    """Preleva una pagina solo se il robots.txt lo consente.

    Il parametro `forza` non aggira il robots.txt: serve soltanto a saltare il
    controllo quando l'URL e' di un sito proprio o di un'agenzia che ha dato
    autorizzazione esplicita, e resta responsabilita' di chi lo usa.

    Un blocco del server non e' un errore da riprovare. I portali maggiori
    dichiarano nel `robots.txt` percorsi consentiti e poi rispondono comunque
    403 alle richieste che non arrivano da un browser: il permesso c'era, il
    server ha negato lo stesso. L'eccezione dedicata serve a far arrivare a chi
    usa lo strumento la sola informazione utile, cioe' quali sono le due vie
    alternative, invece di un errore HTTP grezzo che sembra un guasto.
    """
    if not forza:
        consentito, motivo = robots_consente(url)
        if not consentito:
            raise ProbitaRifiutata(f"{url}: {motivo}")
    dominio = urllib.parse.urlparse(url).netloc
    _attendi_il_turno(dominio)
    richiesta = urllib.request.Request(
        url,
        headers={
            "User-Agent": USER_AGENT,
            "Accept": "text/html,application/xhtml+xml",
            "Accept-Language": "it-IT,it;q=0.9",
        },
    )
    try:
        with urllib.request.urlopen(richiesta, timeout=TIMEOUT_SECONDI) as risposta:
            grezzo = risposta.read()
            codifica = risposta.headers.get_content_charset() or "utf-8"
    except urllib.error.HTTPError as e:
        if e.code in (401, 403, 405, 406, 429, 503):
            raise PrelievoBloccato(
                f"{dominio} ha risposto {e.code}: il prelievo automatico e' bloccato "
                "dalla protezione anti bot del sito, anche se il robots.txt consente "
                "il percorso. Non si insiste: si incolla il testo dell'annuncio in un "
                "file e si usa `annunci importa --file`, oppure si inseriscono i campi "
                "a mano con `annunci aggiungi`."
            ) from e
        raise PrelievoBloccato(f"{dominio} ha risposto {e.code}: {e.reason}") from e
    except (urllib.error.URLError, TimeoutError, OSError) as e:
        raise PrelievoBloccato(f"{dominio} non raggiungibile: {e}") from e
    return grezzo.decode(codifica, errors="replace")


def testo_da_html(html: str) -> str:
    """Estrae il testo visibile, senza dipendenze esterne.

    Non e' un parser HTML completo e non pretende di esserlo: serve solo a ridurre
    una pagina a un blocco di testo che il modello locale possa strutturare.
    """
    import re

    html = re.sub(r"(?is)<(script|style|noscript)[^>]*>.*?</\1>", " ", html)
    html = re.sub(r"(?s)<!--.*?-->", " ", html)
    html = re.sub(r"(?i)<br\s*/?>|</p>|</div>|</li>|</tr>", "\n", html)
    testo = re.sub(r"(?s)<[^>]+>", " ", html)
    import html as modulo_html

    testo = modulo_html.unescape(testo)
    righe = [r.strip() for r in testo.splitlines()]
    return "\n".join(r for r in righe if r)


CAMPI_BLOCCANTI = (
    # campo, che cosa blocca la sua assenza, come si ottiene
    ("mq", "prezzo al metro quadro, quindi lo scarto sulla zona e la graduatoria",
     "dall'annuncio, verificando se e' commerciale o calpestabile"),
    ("prezzo_richiesto", "qualunque calcolo: senza prezzo non c'e' operazione",
     "dall'annuncio"),
    ("rendita_catastale", "il prezzo-valore, cioe' la leva fiscale piu' grossa dell'operazione",
     "si chiede all'agenzia con la visura, oppure la si legge in una visura propria"),
    ("zona_omi", "lo scarto sulla quotazione di zona: senza, si usa la forbice dell'intero Comune",
     'python tools/valuta.py omi zone --comune "..."'),
    ("canone_atteso_mese", "tutto il conto economico della locazione e il rendimento",
     "dagli annunci di affitto comparabili, o dalle quotazioni OMI di locazione"),
    ("spese_condominio_anno", "i costi operativi, quindi il reddito operativo netto",
     "dal consuntivo condominiale degli ultimi due esercizi, non dalla stima"),
    ("categoria", "il moltiplicatore catastale e l'esclusione dall'agevolazione",
     "dalla visura catastale"),
    ("comune", "l'aggancio alle quotazioni OMI e la delibera IMU da cercare",
     "dall'annuncio"),
)
"""I campi la cui assenza blocca un calcolo, in ordine di importanza decrescente.

Non e' l'elenco dei campi vuoti, che sarebbe inutile su un registro di
trentacinque colonne dove molti campi sono facoltativi per costruzione: e'
l'elenco di quelli che, mancando, rendono muto un pezzo del modello, e per
ciascuno dice quale pezzo e come si ottiene il dato.

La costante ha due consumatori, il comando che risponde alla domanda "che cosa
manca" e la scheda di trattativa che elenca cosa chiedere, ed e' la ragione per
cui vive qui accanto alla dataclass invece che dentro uno dei due: due copie di
un elenco divergono, e la loro divergenza non produce un errore ma una scheda
che dice di chiedere una cosa diversa da quella che il comando segnala.
"""


STATI_ANNUNCIO = (
    "da contattare",
    "contattato",
    "visita fissata",
    "visitata",
    "proposta fatta",
    "scartato",
    "acquistato",
)
"""Gli stati ammessi per un annuncio, in ordine di avanzamento della trattativa.

Sono una sorgente unica per due consumatori che prima ne tenevano due copie
divergenti: il menu a tendina della colonna Stato nel foglio Annunci e l'aiuto
dell'opzione `--stato` della riga di comando. Le due copie dicevano cose diverse,
la seconda offriva "contattata", "scartata" e "in trattativa", che nel foglio non
esistono, e la conseguenza era concreta: un valore scritto dalla riga di comando
finiva nella cella senza passare per la validazione, quindi restava li' senza
errore, ma il menu a tendina non lo conteneva e un filtro per stato non lo
trovava dove chi lo aveva scritto lo cercava.

L'ordine e' quello dell'avanzamento e non alfabetico, perche' e' l'ordine in cui
il menu li presenta e in cui una trattativa li attraversa.
"""


CAMPI_ESTRAIBILI = {
    "comune": "il Comune in cui si trova l'immobile",
    "indirizzo": "via e numero civico, o la zona se il civico non c'e'",
    "tipologia": "monolocale, bilocale, trilocale, quadrilocale, villetta, attico",
    "mq": "superficie in metri quadri, solo il numero",
    "prezzo_richiesto": "prezzo richiesto in euro, solo il numero senza separatori",
    "piano": "il piano dell'immobile",
    "classe_energetica": "la classe energetica, una lettera fra A4 e G",
    "spese_condominio_anno": "spese condominiali annue in euro, solo il numero",
    "nuova_costruzione": "SI se e' di nuova costruzione o in costruzione, altrimenti NO",
    # I tre campi seguenti compaiono raramente in un annuncio, ma quando ci sono
    # valgono piu' di tutti gli altri messi insieme: la rendita catastale sblocca
    # il prezzo-valore, che e' la leva fiscale piu' grossa dell'operazione; la
    # categoria decide il moltiplicatore e l'esclusione dall'agevolazione; il
    # canone gia' in essere determina l'intero calcolo del rendimento. Ometterli
    # dallo schema significa non trovarli mai, anche quando sono scritti in
    # chiaro nel testo, ed e' un'omissione che non produce alcun errore.
    "rendita_catastale": "rendita catastale in euro, solo il numero, se indicata",
    "categoria": "categoria catastale, per esempio A/2 o A/3, se indicata",
    "canone_atteso_mese": "canone di locazione mensile in euro se l'immobile e' gia' locato o se l'annuncio ne indica uno, solo il numero",
    "provincia": "sigla della provincia, due lettere",
    "destinazione_uso": "abitazione, ufficio, negozio, box",
    "data_consegna": "data prevista di consegna se in costruzione, altrimenti vuoto",
    # Il regime del venditore si estrae, quello dell'acquirente no: che la vendita
    # sia diretta dal costruttore o soggetta a IVA sta scritto nell'annuncio, mentre
    # essere o non essere in prima casa dipende da chi compra e il testo non lo sa.
    "venditore_impresa": "SI se la vendita e' diretta dal costruttore o da un'impresa, o se l'annuncio indica il prezzo soggetto a IVA, altrimenti vuoto",
    "note": "una sintesi in una riga delle caratteristiche rilevanti",
}


def _testo_per_il_modello(testo: str, limite: int = 24_000) -> str:
    """Riduce il testo al limite conservando la coda, non solo la testa.

    Chi copia un annuncio dal browser prende spesso l'intera pagina, menu e piede
    compresi, e su un portale sono decine di migliaia di caratteri. Il troncamento
    ingenuo taglia la fine, ma e' proprio in fondo che sta la tabella delle
    caratteristiche, cioe' spese condominiali, classe energetica, rendita e
    categoria catastale: i campi che valgono piu' di tutti gli altri messi insieme.
    Tagliare la coda significa perderli senza che nulla lo segnali.

    Si tiene quindi la testa, dove stanno titolo e prezzo, e la coda, dove stanno i
    dati tabellari, dichiarando l'omissione nel mezzo perche' il modello sappia che
    il testo non e' continuo e non provi a ricucirlo.
    """
    if len(testo) <= limite:
        return testo
    meta = limite // 2
    return testo[:meta] + "\n\n[...parte centrale omessa per lunghezza...]\n\n" + testo[-meta:]


def struttura_con_modello_locale(testo: str, url: str = "", modello: str = "") -> dict:
    """Struttura il testo di un annuncio con il modello linguistico locale.

    Il testo non lascia la rete locale: la richiesta va all'istanza Ollama
    configurata, quindi nessun contenuto viene inviato a servizi esterni.
    """
    from .llm_locale import ClienteLocale, LlmNonDisponibile

    descrizione = "\n".join(f"- {k}: {v}" for k, v in CAMPI_ESTRAIBILI.items())
    prompt = (
        "Estrai i dati di questo annuncio immobiliare italiano e rispondi solo con un "
        "oggetto JSON, senza testo prima o dopo, senza blocchi di codice.\n\n"
        f"Campi richiesti:\n{descrizione}\n\n"
        "Regole: se un dato non compare nel testo usa la stringa vuota per i campi "
        "testuali e 0 per quelli numerici; non inventare nulla; non riportare nomi, "
        "numeri di telefono o indirizzi email di persone.\n\n"
        f"Testo dell'annuncio:\n{_testo_per_il_modello(testo)}"
    )
    cliente = ClienteLocale(modello=modello) if modello else ClienteLocale()
    risposta = cliente.completa(prompt, formato_json=True)
    dati = _json_dal_testo(risposta)
    for chiave in ("mq", "prezzo_richiesto", "spese_condominio_anno",
                   "rendita_catastale", "canone_atteso_mese"):
        if chiave in dati:
            dati[chiave] = _numero(dati[chiave])
    if url:
        dati["link"] = url
        dati["fonte"] = urllib.parse.urlparse(url).netloc
    return dati


def _json_dal_testo(testo: str) -> dict:
    """Estrae il primo oggetto JSON da una risposta, tollerando testo attorno."""
    testo = testo.strip()
    inizio = testo.find("{")
    fine = testo.rfind("}")
    if inizio == -1 or fine == -1:
        raise ValueError(f"nessun JSON nella risposta del modello: {testo[:200]}")
    return json.loads(testo[inizio : fine + 1])


def _numero(valore) -> float:
    """Converte in numero una stringa che puo' contenere separatori e valuta."""
    if isinstance(valore, (int, float)):
        return float(valore)
    testo = str(valore)
    tenuto = "".join(c for c in testo if c.isdigit() or c in ".,")
    if not tenuto:
        return 0.0
    if "," in tenuto and "." in tenuto:
        tenuto = tenuto.replace(".", "").replace(",", ".")
    elif "," in tenuto:
        tenuto = tenuto.replace(",", ".")
    elif tenuto.count(".") > 1:
        tenuto = tenuto.replace(".", "")
    elif "." in tenuto:
        # Un solo punto e nessuna virgola: in un annuncio italiano quasi sempre
        # e' il separatore delle migliaia, non il decimale. La discriminante e'
        # quante cifre lo seguono: tre sono migliaia, 175.000 vale
        # centosettantacinquemila; una o due sono decimali, 612.45 vale
        # seicentododici e quarantacinque. Senza questa distinzione un prezzo
        # letto dal modello come stringa diventa 175 euro, e il modello non
        # sbaglia nulla: sbaglia chi lo converte.
        intero, _, coda = tenuto.rpartition(".")
        if len(coda) == 3 and intero:
            tenuto = intero.replace(".", "") + coda
    try:
        return float(tenuto)
    except ValueError:
        return 0.0


# I campi del registro che alimentano una cella di input del workbook, mappati
# sul nome definito della cella. La mappa e' l'intero contratto di questa
# funzione, ed e' scritta per nome e non per coordinata per la ragione di
# ADR-013: un nome inesistente fallisce a voce alta, una coordinata sbagliata
# scrive un prezzo in una cella di manutenzione senza che nulla protesti.
#
# Le tre voci con `trasformazione` esistono perche' il registro e il workbook
# esprimono la stessa cosa in unita' diverse: il registro tiene il canone al
# mese e le spese condominiali all'anno, il workbook vuole il canone al mese e
# le spese all'anno, quindi la conversione e' l'identita'; il prezzo, invece, va
# scelto fra obiettivo e richiesto, e la scelta e' la stessa che fa il foglio di
# confronto, cosi' che i due non divergano.
PRECOMPILAZIONE = (
    # nome definito, campo del registro, formato di stampa, significato dell'assenza
    ("prezzo", "prezzo_da_usare", "euro", "da_chiedere"),
    ("rendita", "rendita_catastale", "euro", "da_chiedere"),
    ("categoria", "categoria", "testo", "da_chiedere"),
    ("mq", "mq", "numero", "da_chiedere"),
    ("comune", "comune", "testo", "da_chiedere"),
    ("canone_mese", "canone_atteso_mese", "euro", "da_chiedere"),
    ("condominio", "spese_condominio_anno", "euro", "da_chiedere"),
    # Le tre voci seguenti hanno un'assenza che non e' una lacuna. Il vuoto dei
    # due campi del regime di acquisto e' il terzo stato di ADR-014, che
    # significa eredita dal foglio Immobile, e non e' un dato da chiedere:
    # `prima_casa` in particolare dipende da chi compra e non dall'immobile,
    # quindi nessuna agenzia lo sa. La base d'asta riguarda solo le vendite
    # giudiziarie, quindi su un immobile di libero mercato deve restare vuota.
    ("prima_casa", "prima_casa", "testo", "neutro"),
    ("da_impresa", "venditore_impresa", "testo", "neutro"),
    ("asta_base", "base_asta", "euro", "neutro"),
)


def precompila_workbook(
    annuncio: "Annuncio",
    percorso_workbook: str,
    azzera_assenti: bool = True,
) -> dict:
    """Scrive nelle celle di input del workbook i dati di un annuncio a registro.

    Toglie il passaggio piu' noioso e piu' pericoloso del percorso di lavoro.
    Scelto l'immobile dalla graduatoria, i suoi dati stavano nel registro e
    andavano ridigitati a mano nei fogli di input: un lavoro di due minuti che
    introduce l'unica classe di errore contro cui il modello non ha difese, cioe'
    la trascrizione. Un prezzo con una cifra in meno produce un'operazione che
    sembra ottima, e nessuna cella va in errore per dirlo.

    Due presidi rendono l'operazione sicura, e sono la ragione per cui questa
    funzione e' piu' lunga di un ciclo di assegnazioni.

    Il primo e' che si scrive per nome definito. Un nome che non esiste piu',
    perche' il generatore e' cambiato, fa fallire la funzione con un messaggio
    che dice quale nome manca; una coordinata sbagliata scriverebbe il valore in
    una cella qualunque.

    Il secondo e' che si rifiuta di scrivere in una cella che contiene una
    formula. Le celle di input del workbook sono gialle e le calcolate grigie,
    ma la distinzione vive nel colore e non nel tipo: niente impedisce a un nome
    di puntare a una cella calcolata, e sovrascriverla romperebbe la catena in
    silenzio. La funzione controlla il contenuto prima di scrivere e riporta le
    celle che ha saltato, invece di fidarsi della convenzione.

    Il terzo presidio riguarda cio' che il registro non ha, ed e' la ragione del
    parametro `azzera_assenti`. Un workbook appena generato porta valori di
    esempio in tutte le celle di input, che servono a mostrare il formato atteso
    e a far funzionare il modello a vuoto. In un file dedicato a un immobile
    reale quei valori diventano pericolosi: se il registro non ha la rendita
    catastale, la cella conserva i 450 euro dell'esempio, il modello applica il
    prezzo-valore su una base inventata, e i controlli di plausibilita' del
    Cruscotto non se ne accorgono perche' guardano se il valore e' zero, non se
    e' vero. Il risultato e' un numero plausibile calcolato su un dato falso, che
    e' precisamente la classe di difetto contro cui il resto del progetto e'
    costruito.

    Con `azzera_assenti` attivo, che e' il default, i campi mancanti vengono
    quindi azzerati invece di essere lasciati all'esempio. La conseguenza e'
    voluta e va capita: il foglio mostrera' un modello visibilmente incompleto,
    con rendimenti a zero e controlli non superati, invece di un modello
    apparentemente sano. Fra le due, la prima e' l'unica onesta.

    L'azzeramento riguarda solo i campi la cui assenza e' una lacuna. I due campi
    del regime di acquisto e la base d'asta hanno un'assenza che significa
    qualcosa, cioe' eredita dal foglio Immobile e non e' un'asta, e azzerarli
    cambierebbe il modello invece di dichiararlo incompleto.

    Restituisce un rapporto di cio' che ha scritto, cio' che ha azzerato, cio'
    che ha saltato e cio' che ha rifiutato perche' la cella era calcolata. Il rapporto e' il valore di
    ritorno e non un effetto collaterale stampato, cosi' che chi chiama decida
    come mostrarlo, e distingue due tipi di assenza: quella che e' una lacuna da
    colmare e quella che non lo e'. La distinzione non e' cosmetica: elencare
    `prima_casa` fra i campi da chiedere sarebbe fuorviante due volte, perche'
    il suo vuoto significa eredita dal foglio Immobile ed e' un comportamento
    corretto, e perche' quel dato dipende da chi compra e non dall'immobile,
    quindi non c'e' nessuno a cui chiederlo.
    """
    from openpyxl import load_workbook

    wb = load_workbook(percorso_workbook)
    scritti, assenti, rifiutati, azzerati = {}, [], [], []

    for nome, campo, formato, natura_assenza in PRECOMPILAZIONE:
        if nome not in wb.defined_names:
            raise ValueError(
                f"il nome definito {nome!r} non esiste nel workbook: il generatore e' "
                "cambiato e la mappa PRECOMPILAZIONE va aggiornata"
            )
        if campo == "prezzo_da_usare":
            valore = annuncio.prezzo_obiettivo or annuncio.prezzo_richiesto
        else:
            valore = getattr(annuncio, campo)

        assente = valore in (None, "", 0, 0.0)
        if assente:
            assenti.append((nome, campo, natura_assenza))
            if not (azzera_assenti and natura_assenza == "da_chiedere"):
                continue

        destinazioni = list(wb.defined_names[nome].destinations)
        if len(destinazioni) != 1:
            rifiutati.append((nome, "il nome non punta a una cella sola"))
            continue
        foglio, coordinata = destinazioni[0]
        cella = wb[foglio][coordinata]
        if isinstance(cella.value, str) and cella.value.startswith("="):
            rifiutati.append((nome, f"{foglio}!{coordinata} contiene una formula"))
            continue

        if assente:
            # Si azzera nel tipo della cella: una cella numerica va a zero, una
            # testuale alla stringa vuota. Scrivere zero in una cella di testo
            # farebbe comparire uno zero dove il controllo cerca il vuoto.
            precedente = cella.value
            cella.value = 0 if isinstance(precedente, (int, float)) else ""
            azzerati.append((nome, f"{foglio}!{coordinata}", precedente))
            continue

        cella.value = valore
        scritti[nome] = (f"{foglio}!{coordinata}", valore, formato)

    wb.save(percorso_workbook)
    return {
        "scritti": scritti,
        "assenti": assenti,
        "azzerati": azzerati,
        "rifiutati": rifiutati,
    }


def esporta_in_excel(registro: Registro, percorso_workbook: str) -> int:
    """Riversa il registro nel foglio Annunci del workbook, preservando le formule.

    Le tre colonne calcolate, prezzo al metro quadro, scarto su OMI e rendimento
    lordo, non vengono toccate: restano formule e si ricalcolano da sole.
    """
    from openpyxl import load_workbook

    wb = load_workbook(percorso_workbook)
    ws = wb["Annunci"]

    intestazione = None
    for riga in range(1, 12):
        if ws.cell(row=riga, column=1).value == "ID":
            intestazione = riga
            break
    if intestazione is None:
        raise ValueError("intestazione del foglio Annunci non trovata")

    # L'ordine deve corrispondere colonna per colonna a quello del foglio Annunci in
    # `excel_builder.foglio_annunci`. I `None` sono le colonne calcolate da formula,
    # che non si toccano mai perche' sovrascriverle le distruggerebbe.
    ordine = [
        "id", "data", "stato", "fonte", "agenzia", "contatto", "link", "comune",
        "provincia", "zona_omi", "indirizzo", "tipologia", "destinazione_uso",
        "nuova_costruzione", "data_consegna", "mq", "prezzo_richiesto",
        "prezzo_obiettivo", None, "quotazione_omi_min", "quotazione_omi_max", None,
        "rendita_catastale", "categoria", "piano", "classe_energetica",
        "spese_condominio_anno", "canone_atteso_mese", None,
        "asta", "base_asta", "data_asta", "tribunale_procedura", "stato_occupazione",
        "punteggio", "note", "prima_casa", "venditore_impresa",
    ]
    colonne_calcolate = {i for i, campo in enumerate(ordine, start=1) if campo is None}
    totale_colonne = len(ordine)

    riga = intestazione + 1
    scritti = 0
    for a in registro.annunci:
        for colonna, campo in enumerate(ordine, start=1):
            if campo is None:
                continue
            valore = getattr(a, campo)
            if campo == "data" and valore:
                try:
                    valore = datetime.fromisoformat(valore).date()
                except ValueError:
                    pass
            # L'assegnazione va fatta sull'attributo e non passando `value` a
            # `cell()`, perche' quella scorciatoia salta l'assegnazione quando il
            # valore e' None e lascerebbe in cella il contenuto precedente: un
            # campo azzerato non si ripulirebbe e l'annuncio esportato erediterebbe
            # in silenzio il dato di quello che occupava prima la riga.
            ws.cell(row=riga, column=colonna).value = valore if valore != 0 else None
        riga += 1
        scritti += 1

    # Ripulisce le righe rimaste da un'esportazione precedente piu' lunga.
    for residua in range(riga, intestazione + 201):
        for colonna in range(1, totale_colonne + 1):
            if colonna not in colonne_calcolate:
                ws.cell(row=residua, column=colonna).value = None

    wb.save(percorso_workbook)
    return scritti
