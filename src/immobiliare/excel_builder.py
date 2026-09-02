# -*- coding: utf-8 -*-
"""Generatore del workbook di valutazione, con formule vive.

Il workbook non e' un rapporto stampato: e' il modello stesso. Ogni numero
derivato e' una formula Excel, non un valore calcolato in Python e incollato, cosi'
che chi apre il file possa cambiare il prezzo o il tasso e vedere ricalcolare tutto
senza rieseguire nulla. Le celle gialle sono gli input, quelle grigie il calcolato,
quelle verdi i risultati di sintesi.

I riferimenti fra fogli passano per nomi definiti, registrati in `_nome`, cosi' che
le formule si leggano come `prezzo * reg_prima` invece che come `Immobile!$B$12`.
"""

from __future__ import annotations

from datetime import date

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from . import annunci as A
from . import parametri as P
from . import stile as S

MAX_RATE = 480          # 40 anni di rate mensili
ESTRAZIONI = 1000       # scenari della simulazione probabilistica
SEME_SIMULAZIONE = 20260831   # dichiarato, cosi' la simulazione e' riproducibile
MAX_ANNI = 40
ORIZZONTE_MAX = 40


class Costruttore:
    """Costruisce il workbook e tiene il registro dei nomi definiti."""

    def __init__(self) -> None:
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        self.nomi: dict[str, str] = {}

    # -- infrastruttura ----------------------------------------------------

    def nome(self, chiave: str, ws, cella: str) -> str:
        """Registra un nome definito che punta a una cella e lo restituisce."""
        riferimento = f"'{ws.title}'!${cella[0]}${cella[1:]}" if cella[1:].isdigit() else f"'{ws.title}'!{cella}"
        self.wb.defined_names.add(DefinedName(chiave, attr_text=riferimento))
        self.nomi[chiave] = chiave
        return chiave

    def nome_intervallo(self, chiave: str, ws, intervallo: str) -> str:
        self.wb.defined_names.add(DefinedName(chiave, attr_text=f"'{ws.title}'!{intervallo}"))
        self.nomi[chiave] = chiave
        return chiave

    def salva(self, percorso: str) -> None:
        self.wb.save(percorso)

    # -- fogli -------------------------------------------------------------

    def costruisci(self) -> None:
        # Il registro degli usi va riempito prima di creare i fogli, perche' e'
        # `stile.titolo` a leggerlo per scrivere la fascia in testa a ciascuno, ed
        # e' la prima cosa che ogni foglio chiama. La sorgente e' la stessa tupla
        # da cui nasce l'indice: cosi' la fascia di un foglio e la sua riga
        # nell'indice non possono dire cose diverse.
        S.USI = {
            nome: (azione, quando, esito)
            for _, fogli in self.PERCORSO
            for nome, azione, quando, esito in fogli
        }
        self.foglio_guida()
        self.foglio_cruscotto()
        self.foglio_parametri()
        self.foglio_immobile()
        self.foglio_mutuo()
        self.foglio_ammortamento()
        self.foglio_simulatore()
        self.foglio_locazione()
        self.foglio_cashflow()
        self.foglio_metriche()
        self.foglio_confronto()
        self.foglio_scenari()
        self.foglio_rischio()
        self.foglio_comproprieta()
        self.foglio_checklist()
        self.foglio_asta()
        self.foglio_dossier()
        # Il foglio Annunci va costruito prima del confronto, che ne legge le righe
        # a partire da `self.riga_annunci`.
        self.foglio_annunci()
        self.foglio_confronto_immobili()
        self.foglio_fonti()
        self.foglio_estrazioni()

    # ------------------------------------------------------------------ guida
    # L'indice dei fogli. La tupla e' la sorgente unica: la usa il foglio Guida per
    # costruire la tabella navigabile, e il test la confronta con i fogli davvero
    # presenti nel workbook. Il rischio che presidia non e' teorico: un foglio
    # rinominato lascerebbe un collegamento ipertestuale che Excel apre senza
    # errore visibile, portando su una destinazione che non esiste piu'.
    #
    # L'ordine e' quello di lettura consigliata e non quello delle linguette, che
    # coincide quasi sempre ma non del tutto: i due fogli di riferimento, Parametri
    # e Fonti, stanno fra le prime linguette per comodita' di consultazione e in
    # fondo al percorso, perche' si aprono quando serve e non all'inizio.
    PERCORSO = (
        ("Da dove si comincia", (
            ("Cruscotto", "Si legge", "Sempre, per primo e per ultimo",
             "I numeri che decidono, con accanto la soglia oltre la quale sono un problema, e il conto di cosa manca ancora"),
        )),
        ("I dati dell'operazione, cioe' le celle gialle", (
            ("Immobile", "Si compila", "Quando c'e' un immobile candidato",
             "Imposte di trasferimento, costo totale dell'operazione, cassa necessaria al rogito"),
            ("Mutuo", "Si compila", "Subito dopo Immobile",
             "Rata, imposta sostitutiva, oneri accessori, detrazione degli interessi anno per anno"),
            ("Locazione", "Si compila", "Solo se l'immobile si mette a reddito",
             "I quattro regimi fiscali a confronto sullo stesso canone, e il regime scelto che alimenta la proiezione"),
            ("Comproprieta", "Si compila", "Solo se si compra in piu' di uno",
             "Ripartizione per quote di esborso, rata e imposte, e l'avvertenza su cosa serve mettere per iscritto"),
        )),
        ("Come si comporta nel tempo", (
            ("Ammortamento", "Si legge", "Dopo aver compilato Mutuo",
             "Il piano rata per rata, fino a quarant'anni, con quota capitale e quota interessi separate"),
            ("Simulatore mutuo", "Si compila e si legge", "Prima di firmare un mutuo, soprattutto a tasso variabile",
             "Effetto dei rimborsi volontari e di un percorso del tasso a gradini, con la rata massima raggiunta e il segnale se il piano si chiude"),
            ("Cash flow", "Si legge", "Dopo Locazione",
             "La proiezione annuale sull'orizzonte scelto, con l'uscita finale"),
        )),
        ("L'esito, e quanto e' fragile", (
            ("Metriche", "Si legge", "Quando gli input sono completi",
             "Rendimento lordo e netto, cap rate, cash on cash, debt service coverage ratio, tasso interno di rendimento, valore attuale netto"),
            ("Confronto affitto", "Si legge", "Solo per l'abitazione propria",
             "Comprare con mutuo oppure restare in affitto investendo la differenza, a parita' di esborso"),
            ("Scenari", "Si compila e si legge", "Prima di decidere, mai dopo",
             "Tre ipotesi impostabili a mano, le tabelle di sensibilita', e il prezzo massimo che l'immobile giustifica al rendimento obiettivo"),
            ("Rischio", "Si legge", "Insieme a Scenari",
             "Mille scenari con estrazioni fisse, le probabilita' di cash flow negativo e di perdita, e il tornado che dice quale variabile pesa"),
        )),
        ("Prima di firmare", (
            ("Checklist", "Si compila", "Dal momento in cui si passa dalla valutazione alla proposta",
             "Le verifiche legali, urbanistiche, catastali e condominiali da chiudere, con la fase in cui vanno chiuse"),
            ("Dossier tecnico", "Si compila", "Prima della proposta, non dopo",
             "I documenti da farsi consegnare in trattativa, con chi li rilascia, la norma che li rende dovuti e il costo"),
        )),
        ("Casi speciali", (
            ("Asta", "Si compila e si legge", "Solo se l'immobile viene da una vendita giudiziaria",
             "Costo reale dell'aggiudicazione, sconto effettivo sul mercato, prezzo massimo a cui fermarsi in gara"),
        )),
        ("La ricerca, cioe' il passo che viene prima di tutti", (
            ("Annunci", "Si compila", "Da subito, appena si inizia a guardare",
             "Il registro degli immobili in valutazione, con prezzo al metro quadro, rendimento lordo e scarto sulla zona OMI calcolati"),
            ("Confronto immobili", "Si legge", "Quando a registro c'e' piu' di un immobile",
             "Tutti gli annunci in fila con lo stesso modello, imposte comprese, per scegliere quale approfondire"),
        )),
        ("Riferimento, si apre quando serve", (
            ("Parametri", "Si consulta", "All'aggiornamento fiscale annuale, o per capire da dove viene un numero",
             "Ogni aliquota e ogni assunzione del modello, con la fonte accanto e la data di revisione"),
            ("Fonti", "Si consulta", "Prima di fidarsi di un numero",
             "Da dove viene ciascun dato, con il collegamento alla fonte istituzionale"),
        )),
    )

    def foglio_guida(self) -> None:
        """Indice navigabile del workbook.

        Il foglio esisteva gia' come pagina di presentazione con un elenco
        descrittivo di undici voci su venti fogli, il che lo rendeva una guida
        parziale e non un indice. La differenza conta piu' di quanto sembri: in un
        file di venti fogli la navigazione via linguette in basso funziona solo per
        chi conosce a memoria dove sta cosa, e chi lo apre la prima volta, o lo
        riapre dopo un mese, non sa in che ordine leggere ne' quali fogli lo
        riguardano. La tabella qui sotto risponde a tre domande per ogni foglio,
        cioe' se si compila o si legge, quando lo si apre nel percorso, e cosa ne
        esce, e ogni nome e' un collegamento.
        """
        ws = self.wb.create_sheet(S.FOGLIO_INDICE)
        ws.sheet_view.showGridLines = False
        S.larghezze_colonne(ws, {"A": 5, "B": 24, "C": 20, "D": 44, "E": 76})
        r = S.titolo(
            ws,
            1,
            "Valutazione di un investimento immobiliare",
            f"Modello aggiornato al {P.REVISIONE.strftime('%d/%m/%Y')} con i parametri fiscali {P.ANNO_IMPOSTA}. "
            "Se non lo hai mai aperto, le cinque righe qui sotto e la legenda dei colori sono tutto quello che serve per cominciare.",
            5,
        )

        r = S.sezione(ws, r, "Se apri questo file per la prima volta, leggi queste cinque righe", 5)
        for testo in [
            "Uno. In questo file si scrive soltanto nelle celle colorate di giallo e di azzurro. Tutto il resto e' calcolato: se ci scrivi sopra, il calcolo si rompe e nessuno te lo dice, perche' il foglio continua a mostrare un numero.",
            "Due. Ogni foglio, in alto, ti dice in una riga se lui e' un foglio dove si scrive oppure uno dove si legge, quando conviene aprirlo e che cosa ne esce. Se la riga in alto e' gialla si compila, se e' grigia si legge.",
            "Tre. Non serve compilare tutto. Se non metti l'immobile a reddito, il foglio Locazione non ti riguarda; se non compri in piu' persone, Comproprieta' non ti riguarda; se non e' un'asta, Asta non ti riguarda. La tabella qui sotto lo dice foglio per foglio nella colonna Quando si apre.",
            "Quattro. I numeri che decidono stanno tutti nel Cruscotto, che e' il secondo foglio. Se hai poco tempo, compila i tre fogli gialli del percorso minimo e poi leggi solo quello.",
            "Cinque. Ogni nome di foglio in questa pagina e' un collegamento: clicci e ci vai. Da ogni foglio torni qui col collegamento in alto a sinistra, quello che dice Indice.",
        ]:
            r = S.nota_riga(ws, r, testo, 5)
        r += 1

        r = S.sezione(ws, r, "I colori delle celle, che sono la cosa piu' importante", 5)
        # I colori si mostrano, non si descrivono: la cella della legenda porta il
        # riempimento che spiega. Descriverli a parole avrebbe richiesto a chi
        # legge di ricordare un'associazione fra un nome e un colore, e la
        # segnalazione d'uso da cui nasce questa sezione era esattamente che quella
        # associazione non era chiara nemmeno a chi ha seguito il progetto.
        legenda = [
            ("Gialla", S.FILL_INPUT,
             "Ci scrivi tu, un numero o un testo. Sono le uniche celle da compilare, e sono poche per foglio."),
            ("Azzurra", S.FILL_SCELTA,
             "Ci scegli da un elenco: clicca la cella e a destra compare una freccia. Un valore scritto a mano fuori dall'elenco viene rifiutato, ed e' voluto."),
            ("Grigia", S.FILL_CALCOLO,
             "La calcola il foglio. Non ci si scrive: sovrascriverla rompe la catena di calcolo in silenzio, cioe' senza alcun messaggio di errore."),
            ("Verde", S.FILL_RISULTATO,
             "Risultato di sintesi. E' quello che sei venuto a leggere, e viene sempre da celle gialle e azzurre compilate altrove."),
            ("Rossa", S.FILL_ATTENZIONE,
             "Attenzione. Un valore ha superato una soglia oltre la quale e' un problema, oppure un controllo di plausibilita' non e' superato."),
        ]
        for nome, riempimento, spiegazione in legenda:
            c = ws.cell(row=r, column=2, value=nome)
            c.fill = riempimento
            c.border = S.BORDO
            c.font = S.ETICHETTA_BOLD
            c.alignment = S.CENTRO
            d = ws.cell(row=r, column=3, value=spiegazione)
            d.font = S.ETICHETTA
            d.alignment = S.SINISTRA
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
            ws.row_dimensions[r].height = 30
            r += 1
        r += 1

        r = S.sezione(ws, r, "Indice dei fogli, in ordine di lettura", 5)
        r = S.intestazioni(ws, r, ["N", "Foglio", "Che cosa si fa qui", "Quando si apre", "Che cosa ne esce"], [5, 24, 20, 44, 76])
        numero = 0
        for fase, fogli in self.PERCORSO:
            r = S.sezione(ws, r, fase, 5, secondaria=True)
            for nome, azione, quando, esito in fogli:
                numero += 1
                c = ws.cell(row=r, column=1, value=numero)
                c.number_format = S.NUMERO
                c.alignment = S.CENTRO
                S.collegamento(ws, r, 2, nome, nome)
                a = ws.cell(row=r, column=3, value=azione)
                a.font = S.ETICHETTA_BOLD if azione.startswith("Si compila") else S.ETICHETTA
                a.alignment = S.SINISTRA
                for colonna, testo in ((4, quando), (5, esito)):
                    d = ws.cell(row=r, column=colonna, value=testo)
                    d.font = S.ETICHETTA
                    d.alignment = S.SINISTRA
                for colonna in range(1, 6):
                    ws.cell(row=r, column=colonna).border = S.BORDO
                ws.row_dimensions[r].height = 30
                r += 1
        ws.freeze_panes = ws.cell(row=r - numero, column=1)
        r += 1

        r = S.sezione(ws, r, "Il percorso minimo, se si ha mezz'ora", 5)
        for testo in [
            "Primo, foglio Annunci: si mettono gli immobili che si stanno guardando, anche solo con link, Comune, metri quadri e prezzo. Il foglio Confronto immobili li mette in fila da solo e dice quale merita tempo.",
            "Secondo, foglio Immobile, sul candidato scelto: prezzo, rendita catastale, chi vende, prima casa. Ne escono imposte e costo reale. Poi foglio Mutuo, e foglio Locazione se l'immobile si affitta.",
            "Terzo, si torna al Cruscotto e si leggono i cinque numeri. Se il cash flow e' negativo, si va nel foglio Scenari e si guarda quanto lo diventa nell'ipotesi pessimistica: e' quella la cifra che si deve poter sostenere ogni mese.",
            "Quarto, prima di fare una proposta si aprono Checklist e Dossier tecnico. Una proposta accettata e' gia' un contratto, quindi le verifiche si chiudono prima, o diventano condizioni scritte nella proposta.",
        ]:
            r = S.nota_riga(ws, r, testo, 5)
        r += 1

        r = S.sezione(ws, r, "Le tre domande a cui il modello risponde", 5)
        for testo in [
            "Quanta cassa serve davvero per chiudere l'operazione, contando imposte, notaio, provvigione e oneri del mutuo, e non solo il prezzo.",
            "Quanto rende l'immobile al netto di tutto, e come si confronta con l'alternativa di non comprarlo.",
            "Quali verifiche legali e tecniche vanno chiuse prima di firmare, perche' una proposta accettata e' gia' un contratto.",
        ]:
            r = S.nota_riga(ws, r, testo, 5)
        r += 1

        r = S.sezione(ws, r, "Avvertenza", 5)
        for testo in [
            "Questo file e' uno strumento di analisi personale, non una consulenza fiscale, legale o finanziaria. Le aliquote sono quelle vigenti alla data di revisione indicata sopra e cambiano con ogni legge di bilancio: prima di firmare qualunque cosa vanno riverificate sulle fonti elencate nel foglio Fonti, e le posizioni soggettive vanno confermate da un notaio e da un commercialista.",
            "Il modello ignora deliberatamente la ristrutturazione come voce di progetto, secondo il perimetro con cui e' stato costruito. La ristrutturazione periodica di fine ciclo, invece, resta come costo ricorrente ammortizzato, perche' un immobile che si tiene quarant'anni va rifatto almeno una volta e ignorarlo falsa il rendimento.",
            "L'aliquota IMU e le spese condominiali sono le due voci che cambiano di piu' da un immobile all'altro: l'aliquota va letta nella delibera del Comune dell'anno in corso, le spese nel consuntivo condominiale degli ultimi due esercizi.",
            "La guida operativa completa, con ogni comando e ogni voce spiegati uno per uno, sta nel file docs/manuale-operativo.md dentro il progetto che ha generato questo workbook.",
        ]:
            r = S.nota_riga(ws, r, testo, 5)

    # -------------------------------------------------------------- cruscotto
    def foglio_cruscotto(self) -> None:
        """Pagina di sintesi: gli otto numeri che decidono, e cosa manca.

        Un workbook di diciotto fogli e' potente e insieme scoraggiante: chi lo apre
        la prima volta non sa dove guardare, e chi lo riapre dopo un mese non ricorda
        dove aveva lasciato. Il cruscotto risolve entrambe le cose senza aggiungere
        modello: non calcola nulla di nuovo, raccoglie con nomi definiti cio' che gli
        altri fogli hanno gia' calcolato e lo mette in una schermata sola.

        La disciplina che lo rende utile e' la rinuncia: qui non entra tutto, entra
        solo cio' che cambia una decisione. Ogni riga porta accanto la soglia oltre la
        quale il numero e' un problema, perche' un indicatore senza soglia si guarda
        e non si usa.
        """
        ws = self.wb.create_sheet("Cruscotto")
        ws.sheet_view.showGridLines = False
        S.larghezze_colonne(ws, {"A": 42, "B": 20, "C": 26, "D": 52})
        r = S.titolo(
            ws,
            1,
            "Cruscotto",
            "Tutto quello che sta qui e' calcolato altrove: questa pagina non aggiunge ipotesi, le riassume. Si legge in un minuto e dice se vale la pena aprire il resto.",
            4,
        )

        def riga_kpi(etichetta, formula, formato, soglia="", commento=""):
            nonlocal r
            e = ws.cell(row=r, column=1, value=etichetta)
            e.font = S.ETICHETTA
            e.alignment = S.SINISTRA
            v = ws.cell(row=r, column=2, value=formula)
            if formato:
                v.number_format = formato
            v.font = S.KPI
            v.fill = S.FILL_RISULTATO
            v.border = S.BORDO
            v.alignment = S.DESTRA
            if soglia:
                sc = ws.cell(row=r, column=3, value=soglia)
                sc.font = S.NOTA
                sc.alignment = S.SINISTRA
            if commento:
                cc = ws.cell(row=r, column=4, value=commento)
                cc.font = S.NOTA
                cc.alignment = S.SINISTRA
            ws.row_dimensions[r].height = 20
            r += 1
            return r - 1

        r = S.sezione(ws, r, "L'operazione", 4)
        riga_kpi("Immobile", '=IF(riferimento_immobile="","da compilare nel foglio Immobile",riferimento_immobile&IF(comune_immobile="",""," - "&comune_immobile))', None,
                 "", "Riferimento interno e Comune, dal foglio Immobile.")
        riga_kpi("Prezzo trattato", "=prezzo", S.EURO, "", "")
        riga_kpi("Costo totale dell'operazione", "=costo_totale", S.EURO,
                 "prezzo piu' imposte e costi", "E' il numero da avere in testa quando si fa la proposta, non il prezzo.")
        riga_cassa = riga_kpi("Cassa necessaria al rogito", "=esborso", S.EURO,
                              "", "Costo totale meno la parte finanziata dalla banca.")
        riga_inc = riga_kpi("Incidenza dei costi sul prezzo", "=incidenza_costi", S.PERC,
                            "attenzione sopra il 10%", "Sopra la soglia conviene capire quale voce pesa: di solito e' la provvigione o l'imposta sostitutiva.")
        r += 1

        r = S.sezione(ws, r, "I numeri che decidono", 4, secondaria=True)
        riga_rn = riga_kpi("Rendimento netto", "=utile_locazione/costo_totale", S.PERC,
                           "confronta con l'obiettivo", "Utile dopo tutti i costi e le imposte, sul costo totale. Fra lordo e netto si perdono di norma due punti e mezzo.")
        riga_cfm = riga_kpi("Cash flow mensile", "=cash_flow_primo_anno/12", S.EURO_DEC,
                            "negativo significa che ci metti", "Se negativo, e' quanto esce dalla tua tasca ogni mese. La domanda non e' se e' bello, ma se e' sostenibile per anni.")
        riga_dscr = riga_kpi("Debt service coverage ratio", '=IF(rata_annua>0,noi_annuo/rata_annua,"nessun mutuo")', S.NUMERO_DEC,
                             "sotto 1 il reddito non copre la rata", "E' l'indicatore che smaschera prima le operazioni troppo tirate.")
        riga_tir = riga_kpi("Tasso interno di rendimento", '=IFERROR(IRR(flussi_tir),"non calcolabile")', S.PERC,
                            "confronta con il portafoglio", "Include l'uscita. E' l'unico numero commensurabile con un investimento finanziario.")
        riga_rr = riga_kpi("Rapporto rata reddito", "=IF(reddito_mensile>0,rata_mensile/reddito_mensile,0)", S.PERC,
                           "le banche si fermano a un terzo", "Sopra il trentacinque per cento la pratica difficilmente passa.")
        r += 1

        r = S.sezione(ws, r, "Il rischio, non solo il caso centrale", 4, secondaria=True)
        riga_kpi("Probabilita' di cash flow negativo", "=prob_cash_negativo", S.PERC_1,
                 "", "Su mille scenari simulati. Con la leva e' quasi sempre alta.")
        riga_kpi("Probabilita' di battere il portafoglio alternativo", "=prob_batte_alternativa", S.PERC_1,
                 "", "Confronto con l'esborso investito al rendimento atteso per lo stesso orizzonte.")
        riga_kpi("Probabilita' di perdere capitale proprio", "=prob_perdita_capitale", S.PERC_1,
                 "", "Patrimonio finale sotto quanto messo all'inizio.")
        riga_kpi("Cash flow annuo nello scenario peggiore su venti", "=PERCENTILE(sim_cash_flow,0.05)", S.EURO,
                 "", "Il numero da poter sostenere, diviso dodici, prima di firmare.")
        r += 1

        r = S.sezione(ws, r, "Il verdetto e cosa manca", 4, secondaria=True)
        riga_ver = riga_kpi("Sintesi automatica", "=verdetto", None, "", "Confronta il tasso interno con il costo opportunita' e con il portafoglio alternativo.")
        ws.cell(row=riga_ver, column=2).alignment = S.SINISTRA
        ws.cell(row=riga_ver, column=2).font = S.ETICHETTA_BOLD
        # Nomi definiti, non coordinate: la versione precedente citava Immobile!$B$21
        # e 'Confronto affitto'!$B$52, e una riga inserita in uno dei due fogli
        # avrebbe cambiato il verdetto del Cruscotto senza sollevare nulla.
        riga_conf = riga_kpi("Comprare oppure restare in affitto", "=IF(abitazione_principale=\"SI\",IF(conf_differenza>0,\"conviene comprare\",\"conviene restare in affitto e investire\"),\"non pertinente: non e' abitazione principale\")", None,
                             "", "Il confronto ha senso solo se l'immobile e' destinato ad abitazione propria. Legge il foglio Confronto affitto per nome definito, non per coordinata.")
        ws.cell(row=riga_conf, column=2).alignment = S.SINISTRA
        ws.cell(row=riga_conf, column=2).font = S.ETICHETTA_BOLD
        riga_ap = riga_kpi("Verifiche ancora aperte", "=verifiche_aperte", S.NUMERO,
                           "vanno a zero prima di firmare", "Nel foglio Checklist. Una proposta accettata e' gia' un contratto: le verifiche si chiudono prima, o diventano condizioni scritte.")
        riga_doc = riga_kpi("Documenti bloccanti ancora da avere", "=documenti_bloccanti_aperti", S.NUMERO,
                            "senza questi non si verifica nulla", "Nel foglio Dossier tecnico. Sono le carte la cui assenza rende nullo l'atto, blocca il mutuo o lascia ignoto il costo di regolarizzazione.")
        riga_kpi("Completamento del fascicolo tecnico", "=documenti_completamento", S.PERC_1,
                 "", "Documenti ricevuti sul totale di quelli applicabili a questo immobile.")
        riga_kpi("Controlli di plausibilita' non superati", "=controlli_falliti", S.NUMERO,
                 "il dettaglio e' qui sotto", "Input che il modello considera non ancora attendibili: il dettaglio, con la ragione di ciascuno, sta nella sezione in fondo a questa pagina.")
        r += 1

        r = S.sezione(ws, r, "Controlli di plausibilita' sugli input", 4, secondaria=True)
        r = S.nota_riga(ws, r, "Il modello non puo' sapere se un input e' giusto, ma puo' sapere se e' ancora quello di esempio, se e' a zero dove uno zero non e' plausibile, o se e' incoerente con un'altra scelta. I controlli qui sotto sono di questo tipo, e ciascuno dice che cosa comporta il valore trovato. Nessuno di essi blocca il calcolo: il foglio continua a produrre numeri, e il punto e' proprio che li produrrebbe anche sbagliati senza dirlo.", 4)
        intest_ctrl = ["Controllo", "Esito", "Che cosa comporta", "Come si chiude"]
        r = S.intestazioni(ws, r, intest_ctrl, [42, 30, 60, 56])
        prima_ctrl = r

        # Ogni controllo e' una formula che restituisce la stringa vuota quando
        # e' superato e un messaggio quando non lo e'. La stringa vuota, e non
        # la parola ok, perche' una colonna di ok e' rumore che si impara a
        # ignorare, mentre una colonna quasi vuota rende visibile cio' che resta.
        controlli = [
            ("Rendita catastale contro prezzo-valore",
             '=IF(AND(usa_prezzo_valore="SI",rendita<=0),"rendita a zero con prezzo-valore attivo","")',
             "L'opzione prezzo-valore non si applica senza rendita, quindi le imposte si calcolano sul prezzo intero: e' la leva fiscale piu' grossa dell'operazione e si sta perdendo in silenzio.",
             "Si chiede la visura catastale all'agenzia, o la si estrae da una visura propria."),
            ("Aliquota IMU deliberata dal Comune",
             '=IF(imu_aliquota=imu_base,"ancora al valore base di legge","")',
             "I Comuni possono azzerarla o portarla all'1,06 per cento: sul valore base l'IMU stimata puo' sbagliare di un quarto, ogni anno per tutta la durata del possesso.",
             "Si legge nella delibera comunale dell'anno in corso e si scrive nel foglio Locazione."),
            ("Spese condominiali",
             '=IF(condominio<=0,"a zero","")',
             "Un immobile in condominio ha spese, e a zero il reddito operativo netto e' sovrastimato di tutto il loro importo.",
             "Dal consuntivo condominiale degli ultimi due esercizi, non dalla stima dell'agenzia; insieme si leggono i verbali per i lavori deliberati."),
            ("Canone atteso",
             '=IF(AND(abitazione_principale<>"SI",canone_mese<=0),"a zero su un immobile da mettere a reddito","")',
             "Senza canone il conto economico della locazione e ogni rendimento sono vuoti, e il cash flow risulta pari alla sola rata.",
             "Dagli annunci di affitto comparabili nella stessa zona, o dalle quotazioni OMI di locazione."),
            ("Superficie",
             '=IF(mq<=0,"a zero","")',
             "Senza superficie non esiste il prezzo al metro quadro, quindi non e' possibile alcun confronto con le quotazioni di zona.",
             "Dall'annuncio, verificando se la superficie dichiarata e' commerciale o calpestabile."),
            ("Comune",
             '=IF(comune="","non compilato","")',
             "Serve a ritrovare la delibera IMU e la zona OMI: senza, entrambe le verifiche restano aperte.",
             "Dall'annuncio."),
            ("Assicurazione del fabbricato",
             '=IF(assicurazione<=0,"a zero","")',
             "La polizza incendio e' obbligatoria con un mutuo e il suo costo esiste comunque: a zero il conto economico e' ottimistico.",
             "Dal preventivo della polizza, o da quello che la banca propone; si puo' portarne una propria equivalente."),
            ("Patrimonio complessivo",
             '=IF(patrimonio_totale<=0,"non compilato, controllo di concentrazione spento","")',
             "Il rischio di concentrazione e' quello che nessun rendimento vede: chi ha due terzi del patrimonio in mattone non ha un portafoglio ma una scommessa su una zona.",
             "Si somma immobili gia' posseduti, liquidita' e investimenti, incluso questo acquisto, nel foglio Metriche."),
        ]

        for etichetta, formula, comporta, chiude in controlli:
            e = ws.cell(row=r, column=1, value=etichetta)
            e.font = S.ETICHETTA
            e.alignment = S.SINISTRA
            v = ws.cell(row=r, column=2, value=formula)
            v.font = S.ETICHETTA_BOLD
            v.alignment = S.SINISTRA
            for colonna, testo in ((3, comporta), (4, chiude)):
                c = ws.cell(row=r, column=colonna, value=testo)
                c.font = S.NOTA
                c.alignment = S.SINISTRA
            for colonna in range(1, 5):
                ws.cell(row=r, column=colonna).border = S.BORDO
            ws.row_dimensions[r].height = 34
            r += 1
        ultima_ctrl = r - 1

        # Il contatore che il Cruscotto mostra in testa. Conta le celle non vuote
        # della colonna dell'esito, cioe' i controlli non superati.
        self.nome_intervallo("controlli_esiti", ws, f"$B${prima_ctrl}:$B${ultima_ctrl}")
        riga_cont = r
        r = S.campo(ws, r, "Controlli non superati", '=COUNTIF(controlli_esiti,"?*")', S.NUMERO, risultato=True,
                    nota="Conta le celle dell'esito che portano un messaggio. Il criterio conta le stringhe non vuote e non le celle non vuote, perche' una formula che restituisce la stringa vuota produce una cella tecnicamente non vuota.")
        self.nome("controlli_falliti", ws, f"B{riga_cont}")
        ws.conditional_formatting.add(
            f"B{riga_cont}:B{riga_cont}",
            CellIsRule(operator="greaterThan", formula=["0"], fill=S.FILL_ATTENZIONE),
        )
        ws.conditional_formatting.add(
            f"B{prima_ctrl}:B{ultima_ctrl}",
            CellIsRule(operator="notEqual", formula=['""'], fill=S.FILL_ATTENZIONE),
        )
        r = S.nota_riga(ws, r, "Un controllo non superato non e' un errore del modello: e' un input che non e' ancora un dato. La distinzione conta perche' il foglio calcola comunque, e un numero calcolato su un input di esempio ha la stessa faccia di un numero calcolato su un dato vero.", 4)
        r += 1

        for cella, regola in (
            (f"B{riga_inc}", CellIsRule(operator="greaterThan", formula=["0.10"], fill=S.FILL_ATTENZIONE)),
            (f"B{riga_cfm}", CellIsRule(operator="lessThan", formula=["0"], fill=S.FILL_ATTENZIONE)),
            (f"B{riga_dscr}", CellIsRule(operator="lessThan", formula=["1"], fill=S.FILL_ATTENZIONE)),
            (f"B{riga_rr}", CellIsRule(operator="greaterThan", formula=["0.35"], fill=S.FILL_ATTENZIONE)),
            (f"B{riga_ap}", CellIsRule(operator="greaterThan", formula=["0"], fill=S.FILL_ATTENZIONE)),
            (f"B{riga_doc}", CellIsRule(operator="greaterThan", formula=["0"], fill=S.FILL_ATTENZIONE)),
            (f"B{riga_rn}", CellIsRule(operator="lessThan", formula=["rend_obiettivo"], fill=S.FILL_ATTENZIONE)),
        ):
            ws.conditional_formatting.add(cella, regola)

        r = S.sezione(ws, r, "Dove andare, in ordine", 4, secondaria=True)
        percorso = [
            ("Annunci", "Butta dentro gli immobili che stai guardando: link, Comune, metri quadri, prezzo."),
            ("Confronto immobili", "Si popola da solo e dice quale merita tempo. Guarda cash flow e DSCR, non il rendimento lordo."),
            ("Immobile", "Per il candidato scelto: prezzo, rendita catastale, chi vende, prima casa. Ne escono imposte e costo reale."),
            ("Mutuo e Simulatore mutuo", "Importo, tasso, durata. Il simulatore prova rimborsi volontari e rialzi di tasso."),
            ("Locazione", "Canone, spese condominiali dal consuntivo, aliquota IMU dalla delibera. Quattro regimi a confronto."),
            ("Metriche, Scenari, Rischio", "Si leggono. Scenari da' tre ipotesi impostabili, Rischio la distribuzione su mille."),
            ("Comproprieta'", "Solo se comprate in piu' di uno: ripartisce per quote e calcola l'imposta di ciascuno."),
            ("Checklist", "Quando si passa dalla valutazione alla proposta. Filtra per fase e non firmare con righe rosse."),
            ("Asta", "Solo se l'immobile viene da una vendita giudiziaria: le regole sono altre e il foglio le mette in conto."),
            ("Dossier tecnico", "Le carte da farsi dare in trattativa, con la norma che le rende dovute. Si chiedono prima della proposta, non dopo."),
        ]
        for foglio, cosa in percorso:
            a = ws.cell(row=r, column=1, value=foglio)
            a.font = S.ETICHETTA_BOLD
            a.alignment = S.SINISTRA
            b = ws.cell(row=r, column=2, value=cosa)
            b.font = S.ETICHETTA
            b.alignment = S.SINISTRA
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
            ws.row_dimensions[r].height = 26
            r += 1
        r += 1
        r = S.nota_riga(ws, r, "Le celle di questa pagina sono tutte calcolate: non si scrive qui. Gli input stanno nei fogli gialli, e ogni numero del cruscotto risale a uno di quelli attraverso un nome definito.", 4)

    # -------------------------------------------------------------- parametri
    def foglio_parametri(self) -> None:
        ws = self.wb.create_sheet("Parametri")
        ws.sheet_view.showGridLines = False
        S.larghezze_colonne(ws, {"A": 46, "B": 14, "C": 58, "D": 46})
        r = S.titolo(
            ws,
            1,
            "Parametri fiscali e di mercato",
            "Ogni valore e' modificabile: se cambia la legge si aggiorna qui e tutto il modello segue. "
            "La colonna delle note dice da dove viene il numero.",
            4,
        )

        def riga(chiave, etichetta, valore, formato, nota, fonte=""):
            nonlocal r
            e = ws.cell(row=r, column=1, value=etichetta)
            e.font = S.ETICHETTA
            e.alignment = S.SINISTRA
            v = ws.cell(row=r, column=2, value=valore)
            v.number_format = formato
            v.fill = S.FILL_INPUT
            v.border = S.BORDO
            v.alignment = S.DESTRA
            n = ws.cell(row=r, column=3, value=nota)
            n.font = S.NOTA
            n.alignment = S.SINISTRA
            if fonte:
                f = ws.cell(row=r, column=4, value=fonte)
                f.font = S.NOTA
                f.alignment = S.SINISTRA
            if chiave:
                self.nome(chiave, ws, f"B{r}")
            r += 1

        t = P.IMPOSTE_TRASFERIMENTO
        r = S.sezione(ws, r, "Imposte sul trasferimento, acquisto da privato o esente IVA", 4)
        riga("reg_prima", "Imposta di registro, prima casa", t.registro_prima_casa, S.PERC, "Sulla base imponibile, che con il prezzo-valore e' il valore catastale.", P.FONTI["imposte_acquisto"])
        riga("reg_ord", "Imposta di registro, aliquota ordinaria", t.registro_ordinario, S.PERC, "Seconda casa o mancanza dei requisiti prima casa.")
        riga("reg_min", "Imposta di registro, minimo di legge", t.registro_minimo, S.EURO, "Si paga comunque, anche se la percentuale darebbe meno.")
        riga("ipo_priv", "Imposta ipotecaria, da privato", t.ipotecaria_da_privato, S.EURO, "Misura fissa.")
        riga("cat_priv", "Imposta catastale, da privato", t.catastale_da_privato, S.EURO, "Misura fissa.")
        r += 1

        r = S.sezione(ws, r, "Imposte sul trasferimento, acquisto da impresa con IVA", 4)
        riga("iva_prima", "IVA, prima casa", t.iva_prima_casa, S.PERC, "Cessione entro cinque anni dall'ultimazione, o con opzione dell'impresa.")
        riga("iva_ord", "IVA, aliquota ordinaria abitativa", t.iva_ordinaria, S.PERC, "Abitazione non di lusso senza requisiti prima casa.")
        riga("iva_lusso", "IVA, categorie A/1, A/8, A/9", t.iva_lusso, S.PERC, "Immobili di pregio, ville e castelli: nessuna agevolazione.")
        riga("fisso_impresa", "Registro, ipotecaria e catastale in misura fissa", t.registro_fisso_da_impresa, S.EURO, "Duecento euro ciascuna, quindi seicento in totale.")
        r += 1

        r = S.sezione(ws, r, "Regola prezzo-valore", 4)
        riga("riv_rendita", "Rivalutazione della rendita catastale", t.rivalutazione_rendita, S.NUMERO_DEC, "Coefficiente fisso del cinque per cento.")
        riga("molt_prima", "Moltiplicatore, prima casa", t.moltiplicatore_prima_casa, S.NUMERO, "Valore catastale uguale a rendita per 1,05 per 110.")
        riga("molt_ord", "Moltiplicatore, altri fabbricati abitativi", t.moltiplicatore_ordinario, S.NUMERO, "Valore catastale uguale a rendita per 1,05 per 120.")
        r = S.nota_riga(ws, r, "Il prezzo-valore vale solo fuori campo IVA, per persone fisiche, su immobili abitativi e pertinenze, e va chiesto espressamente al notaio in atto. Porta con se' anche la riduzione del trenta per cento dell'onorario notarile e il blocco dell'accertamento di valore.", 4)
        r += 1

        m = P.MUTUO
        r = S.sezione(ws, r, "Mutuo", 4)
        riga("sost_prima", "Imposta sostitutiva, mutuo prima casa", m.imposta_sostitutiva_prima_casa, S.PERC, "Trattenuta dalla banca sull'erogato. Assorbe registro, bollo, ipotecarie e catastali sul finanziamento.")
        riga("sost_ord", "Imposta sostitutiva, altri mutui", m.imposta_sostitutiva_ordinaria, S.PERC, "Otto volte l'aliquota agevolata: pesa molto sulle seconde case.")
        riga("detr_aliq", "Detrazione IRPEF sugli interessi passivi", m.detrazione_interessi_aliquota, S.PERC, "Solo su abitazione principale, con residenza trasferita entro dodici mesi.")
        riga("detr_max", "Massimale annuo di interessi detraibili", m.detrazione_interessi_massimale, S.EURO, "Riferito all'immobile: va diviso fra i cointestatari del mutuo.")
        riga("ltv_max", "Loan to value oltre cui serve garanzia esterna", m.ltv_ordinario_max, S.PERC, "Sopra questa soglia serve il fondo Consap o una garanzia integrativa.")
        r += 1

        l = P.LOCAZIONE
        r = S.sezione(ws, r, "Tassazione dei canoni", 4)
        riga("ced_libero", "Cedolare secca, canone libero", l.cedolare_libero, S.PERC, "Contratti 4 piu' 4 e transitori a canone libero.", P.FONTI["locazioni_brevi"])
        riga("ced_conc", "Cedolare secca, canone concordato", l.cedolare_concordato, S.PERC, "Contratti 3 piu' 2 e studenti, nei Comuni ad alta tensione abitativa.")
        riga("ced_breve1", "Cedolare secca, locazione breve prima unita'", l.cedolare_breve_prima_unita, S.PERC, "Una sola unita' per periodo d'imposta, a scelta in dichiarazione.")
        riga("ced_breve2", "Cedolare secca, locazione breve dalla seconda", l.cedolare_breve_altre_unita, S.PERC, "Dal 2026 il regime copre al massimo due unita': dalla terza scatta la presunzione di impresa.")
        riga("abbatt_ord", "Abbattimento forfettario, IRPEF ordinaria", l.abbattimento_forfettario_ordinario, S.PERC, "Imponibile pari al novantacinque per cento del canone.")
        riga("abbatt_conc", "Abbattimento forfettario, canone concordato", l.abbattimento_forfettario_concordato, S.PERC, "Imponibile pari al settantacinque per cento del canone.")
        riga("reg_loc", "Imposta di registro annuale sul canone", l.registro_annuo, S.PERC, "Solo in regime ordinario: la cedolare secca la sostituisce. Meta' per parte.")
        riga("reg_loc_min", "Imposta di registro, minimo", l.registro_minimo, S.EURO, "Per la prima annualita'.")
        r += 1

        r = S.sezione(ws, r, "IRPEF e addizionali", 4)
        riga("irpef_s1", "Primo scaglione, soglia", 28000, S.EURO, "Aliquota del ventitre' per cento fino a questa soglia.", P.FONTI["irpef_2026"])
        riga("irpef_a1", "Primo scaglione, aliquota", 0.23, S.PERC, "")
        riga("irpef_s2", "Secondo scaglione, soglia", 50000, S.EURO, "Ridotta dal trentacinque al trentatre' per cento dalla legge di bilancio 2026.")
        riga("irpef_a2", "Secondo scaglione, aliquota", 0.33, S.PERC, "")
        riga("irpef_a3", "Terzo scaglione, aliquota", 0.43, S.PERC, "Oltre i cinquantamila euro.")
        riga("addizionali", "Addizionali regionale e comunale, stima", P.IRPEF.addizionale_regionale_tipica + P.IRPEF.addizionale_comunale_tipica, S.PERC, "Variano per Regione e Comune: qui una stima prudenziale complessiva.")
        r += 1

        i = P.IMU
        r = S.sezione(ws, r, "IMU", 4)
        riga("imu_molt", "Moltiplicatore catastale, gruppo A", i.moltiplicatore_gruppo_a, S.NUMERO, "Base imponibile uguale a rendita per 1,05 per 160.")
        riga("imu_base", "Aliquota base", i.aliquota_base, S.PERC, "I Comuni possono azzerarla o portarla all'1,06 per cento: leggere la delibera.", P.FONTI["imu_2026"])
        riga("imu_conc", "Quota dovuta con canone concordato", i.riduzione_canone_concordato, S.PERC, "Sconto del venticinque per cento sull'imposta.")
        r += 1

        c = P.COSTI
        r = S.sezione(ws, r, "Costi accessori e di gestione", 4)
        riga("iva_provv", "IVA sulla provvigione di agenzia", c.iva_su_provvigione, S.PERC, "La provvigione e' dovuta alla conclusione dell'affare, cioe' all'accettazione della proposta.")
        riga("ristrutt_pct", "Ristrutturazione di fine ciclo, quota del valore", c.ristrutturazione_su_valore, S.PERC, "Un rifacimento completo costa circa un terzo del valore dell'immobile.")
        riga("ristrutt_anni", "Anni fra due ristrutturazioni complete", c.anni_fra_ristrutturazioni, S.NUMERO, "Impostazione ripresa dal foglio rendita immobiliare di Paolo Coletti.", P.FONTI["coletti_rendita"])
        r += 1

        f = P.FINANZA
        r = S.sezione(ws, r, "Assunzioni finanziarie", 4)
        riga("infl", "Inflazione attesa", f.inflazione_attesa, S.PERC, "Usata per indicizzare canone e valore nominale dell'immobile.")
        riga("rend_port", "Rendimento lordo atteso del portafoglio alternativo", f.rendimento_portafoglio_lordo, S.PERC, "Assunzione, non previsione. Serve solo al confronto con il non comprare.")
        riga("tax_port", "Tassazione delle rendite finanziarie", f.tassazione_rendite_finanziarie, S.PERC, "Dodici e mezzo per cento sui titoli di Stato, ventisei sul resto.")
        riga("tasso_sconto", "Tasso di sconto reale per il valore attuale netto", f.tasso_sconto_reale, S.PERC, "Il costo opportunita' del capitale immobilizzato.")
        r += 1

        p = P.PLUSVALENZA
        r = S.sezione(ws, r, "Rivendita", 4)
        riga("plus_anni", "Anni entro cui la plusvalenza e' imponibile", p.anni_imponibilita_ordinaria, S.NUMERO, "Dieci anni se l'immobile ha avuto interventi con superbonus.")
        riga("plus_aliq", "Imposta sostitutiva sulla plusvalenza", p.imposta_sostitutiva, S.PERC, "Alternativa alla tassazione IRPEF, si chiede al notaio in atto.", P.FONTI["plusvalenza"])
        riga("costi_vendita", "Costi di vendita, quota del prezzo", 0.03, S.PERC, "Provvigione di agenzia in uscita, APE, pratiche.")

    # --------------------------------------------------------------- immobile
    def foglio_immobile(self) -> None:
        ws = self.wb.create_sheet("Immobile")
        ws.sheet_view.showGridLines = False
        S.larghezze_colonne(ws, {"A": 44, "B": 18, "C": 66, "D": 16, "E": 16, "F": 16, "G": 16, "H": 16})
        r = S.titolo(
            ws,
            1,
            "Immobile e costo dell'operazione",
            "Il costo totale, non il prezzo, e' il denominatore corretto di ogni rendimento: comprende imposte, notaio, provvigione e oneri del mutuo.",
        )

        si_no = DataValidation(type="list", formula1='"SI,NO"', allow_blank=False)
        ws.add_data_validation(si_no)
        categorie = DataValidation(type="list", formula1='"A/1,A/2,A/3,A/4,A/5,A/6,A/7,A/8,A/9,A/11"', allow_blank=True)
        ws.add_data_validation(categorie)

        r = S.sezione(ws, r, "Identificazione")
        riga_rif = r
        r = S.campo(ws, r, "Riferimento interno", "house_1", input_utente=True, nota="Lo stesso identificativo usato nel foglio Annunci.")
        self.nome("riferimento_immobile", ws, f"B{riga_rif}")
        riga_com = r
        riga_comune = r
        r = S.campo(ws, r, "Comune", "", input_utente=True, nota="Serve a ritrovare la delibera IMU e la zona OMI di riferimento.")
        # Il nome esiste perche' la precompilazione da registro scrive per nome
        # definito e non per coordinata: senza, questa cella non sarebbe
        # raggiungibile e il Comune resterebbe l'unico campo da ridigitare.
        self.nome("comune", ws, f"B{riga_comune}")
        self.nome("comune_immobile", ws, f"B{riga_com}")
        r = S.campo(ws, r, "Indirizzo", "", input_utente=True)
        r = S.campo(ws, r, "Link annuncio", "", input_utente=True)
        riga_mq = r
        r = S.campo(ws, r, "Superficie commerciale", 55, S.NUMERO, input_utente=True, nota="Metri quadri commerciali, non calpestabili: e' la base del prezzo al metro quadro di mercato.")
        self.nome("mq", ws, f"B{riga_mq}")
        riga_cat = r
        r = S.campo(ws, r, "Categoria catastale", "A/3", input_utente=True, nota="A/1, A/8 e A/9 sono escluse dall'agevolazione prima casa e scontano IVA al ventidue per cento.")
        S.scelta(categorie, ws.cell(row=riga_cat, column=2))
        self.nome("categoria", ws, f"B{riga_cat}")
        riga_rendita = r
        r = S.campo(ws, r, "Rendita catastale", 450, S.EURO_DEC, input_utente=True, nota="Si legge nella visura catastale. Serve al prezzo-valore e all'IMU.")
        self.nome("rendita", ws, f"B{riga_rendita}")
        r += 1

        r = S.sezione(ws, r, "Prezzo e natura dell'operazione")
        riga_richiesto = r
        r = S.campo(ws, r, "Prezzo richiesto", 130000, S.EURO, input_utente=True)
        self.nome("prezzo_richiesto", ws, f"B{riga_richiesto}")
        riga_prezzo = r
        r = S.campo(ws, r, "Prezzo trattato, da mettere in proposta", 120000, S.EURO, input_utente=True, nota="E' il prezzo su cui si costruisce tutta l'analisi.")
        self.nome("prezzo", ws, f"B{riga_prezzo}")
        riga_impresa = r
        r = S.campo(ws, r, "Venditore impresa con IVA", "NO", input_utente=True, nota="SI se si compra da impresa costruttrice entro cinque anni dall'ultimazione, o con opzione per l'imponibilita'.")
        S.scelta(si_no, ws.cell(row=riga_impresa, column=2))
        self.nome("da_impresa", ws, f"B{riga_impresa}")
        riga_nuova = r
        r = S.campo(ws, r, "Nuova costruzione", "NO", input_utente=True, nota="Attiva le verifiche del decreto legislativo 122/2005 nella checklist: fideiussione e polizza decennale postuma.")
        S.scelta(si_no, ws.cell(row=riga_nuova, column=2))
        self.nome("nuova_costruzione", ws, f"B{riga_nuova}")
        riga_prima = r
        r = S.campo(ws, r, "Agevolazione prima casa", "SI", input_utente=True, nota="Richiede residenza nel Comune entro diciotto mesi e assenza di altra prima casa agevolata, salvo rivendita entro due anni.")
        S.scelta(si_no, ws.cell(row=riga_prima, column=2))
        self.nome("prima_casa", ws, f"B{riga_prima}")
        riga_pv = r
        r = S.campo(ws, r, "Opzione prezzo-valore", "SI", input_utente=True, nota="Da chiedere al notaio in atto. Non si applica se si compra da impresa con IVA.")
        S.scelta(si_no, ws.cell(row=riga_pv, column=2))
        self.nome("usa_prezzo_valore", ws, f"B{riga_pv}")
        riga_quota = r
        r = S.campo(ws, r, "Quota di acquisto", 1.0, S.PERC, input_utente=True, nota="Cinquanta per cento se si compra in due. Incide sul massimale della detrazione degli interessi.")
        self.nome("quota", ws, f"B{riga_quota}")
        riga_abit = r
        r = S.campo(ws, r, "Destinato ad abitazione principale", "NO", input_utente=True, nota="SI se ci si va a vivere: abilita la detrazione degli interessi e l'esenzione IMU. NO se si compra per affittarlo.")
        S.scelta(si_no, ws.cell(row=riga_abit, column=2))
        self.nome("abitazione_principale", ws, f"B{riga_abit}")
        r += 1

        r = S.sezione(ws, r, "Base imponibile e imposte di trasferimento", secondaria=True)
        # L'ordine conta: il moltiplicatore catastale dipende dall'agevolazione
        # effettivamente applicabile, non da quella richiesta, quindi le due celle di
        # controllo vanno calcolate prima del valore catastale. Usare la prima casa
        # richiesta darebbe il moltiplicatore 110 anche su una categoria di lusso, che
        # dall'agevolazione e' esclusa, e sottostimerebbe l'imposta.
        riga_lusso = r
        r = S.campo(ws, r, "Categoria di lusso", '=IF(OR(categoria="A/1",categoria="A/8",categoria="A/9"),"SI","NO")', nota="Esclude in ogni caso l'agevolazione prima casa.")
        self.nome("di_lusso", ws, f"B{riga_lusso}")
        riga_agev = r
        r = S.campo(ws, r, "Agevolazione effettivamente applicabile", '=IF(AND(prima_casa="SI",di_lusso="NO"),"SI","NO")', nota="Prima casa richiesta e categoria ammessa.")
        self.nome("agevolata", ws, f"B{riga_agev}")
        riga_vc = r
        r = S.campo(ws, r, "Valore catastale", '=rendita*riv_rendita*IF(agevolata="SI",molt_prima,molt_ord)', S.EURO, nota="Rendita per 1,05 per il moltiplicatore di categoria: 110 se agevolata, 120 altrimenti.")
        self.nome("valore_catastale", ws, f"B{riga_vc}")
        riga_base = r
        r = S.campo(
            ws, r, "Base imponibile del registro",
            '=IF(da_impresa="SI",prezzo,IF(AND(usa_prezzo_valore="SI",rendita>0),valore_catastale,prezzo))',
            S.EURO, nota="Con il prezzo-valore si tassa il valore catastale, che di norma e' molto piu' basso del prezzo.",
        )
        self.nome("base_registro", ws, f"B{riga_base}")
        r += 1

        riga_iva = r
        r = S.campo(ws, r, "IVA", '=IF(da_impresa="SI",prezzo*IF(agevolata="SI",iva_prima,IF(di_lusso="SI",iva_lusso,iva_ord)),0)', S.EURO)
        self.nome("imp_iva", ws, f"B{riga_iva}")
        riga_reg = r
        r = S.campo(ws, r, "Imposta di registro", '=IF(da_impresa="SI",fisso_impresa,MAX(base_registro*IF(agevolata="SI",reg_prima,reg_ord),reg_min))', S.EURO)
        self.nome("imp_registro", ws, f"B{riga_reg}")
        riga_ipo = r
        r = S.campo(ws, r, "Imposta ipotecaria", '=IF(da_impresa="SI",fisso_impresa,ipo_priv)', S.EURO)
        self.nome("imp_ipo", ws, f"B{riga_ipo}")
        riga_cat2 = r
        r = S.campo(ws, r, "Imposta catastale", '=IF(da_impresa="SI",fisso_impresa,cat_priv)', S.EURO)
        self.nome("imp_cat", ws, f"B{riga_cat2}")
        riga_imposte = r
        r = S.campo(ws, r, "Totale imposte di trasferimento", "=imp_iva+imp_registro+imp_ipo+imp_cat", S.EURO, risultato=True)
        self.nome("imposte_totali", ws, f"B{riga_imposte}")
        r = S.nota_riga(ws, r, "Confronto utile: se si comprasse senza agevolazione prima casa, le imposte sarebbero quelle indicate qui sotto. La differenza e' il valore economico del bonus, da pesare contro il fatto che il bonus si consuma e non si puo' riusare su un acquisto futuro finche' non si rivende.")
        riga_senza = r
        r = S.campo(
            ws, r, "Imposte senza agevolazione prima casa",
            '=IF(da_impresa="SI",prezzo*IF(di_lusso="SI",iva_lusso,iva_ord)+3*fisso_impresa,MAX(IF(AND(usa_prezzo_valore="SI",rendita>0),rendita*riv_rendita*molt_ord,prezzo)*reg_ord,reg_min)+ipo_priv+cat_priv)',
            S.EURO, nota="Scenario alternativo, per quantificare il beneficio.",
        )
        self.nome("imposte_senza_agev", ws, f"B{riga_senza}")
        riga_benef = r
        r = S.campo(ws, r, "Valore del bonus prima casa", "=imposte_senza_agev-imposte_totali", S.EURO, risultato=True)
        self.nome("valore_bonus", ws, f"B{riga_benef}")
        r += 1

        r = S.sezione(ws, r, "Costi accessori dell'operazione", secondaria=True)
        riga_provv_pct = r
        r = S.campo(ws, r, "Provvigione di agenzia", 0.03, S.PERC, input_utente=True, nota="Percentuale sul prezzo, al netto di IVA. Mettere zero se si tratta direttamente col privato.")
        self.nome("provv_pct", ws, f"B{riga_provv_pct}")
        riga_provv = r
        r = S.campo(ws, r, "Provvigione, IVA inclusa", "=prezzo*provv_pct*(1+iva_provv)", S.EURO, nota="Dovuta alla conclusione dell'affare: pattuire in proposta che nulla e' dovuto se la condizione sospensiva del mutuo non si avvera.")
        self.nome("provvigione", ws, f"B{riga_provv}")
        riga_notaio = r
        r = S.campo(ws, r, "Notaio, atto di compravendita", 2000, S.EURO, input_utente=True, nota="Con il prezzo-valore l'onorario scende del trenta per cento. Chiedere sempre due o tre preventivi scritti.")
        self.nome("notaio_cv", ws, f"B{riga_notaio}")
        riga_altri = r
        r = S.campo(ws, r, "Altri costi", 2000, S.EURO, input_utente=True, nota="Visure, relazione notarile preliminare, tecnico di parte, allacci, accatastamento, arredo minimo.")
        self.nome("altri_costi", ws, f"B{riga_altri}")
        r += 1

        r = S.sezione(ws, r, "Sintesi del fabbisogno di cassa", secondaria=True)
        riga_acc = r
        r = S.campo(ws, r, "Totale costi accessori", "=imposte_totali+provvigione+notaio_cv+altri_costi+oneri_mutuo", S.EURO, nota="Gli oneri del mutuo arrivano dal foglio Mutuo.")
        self.nome("costi_accessori", ws, f"B{riga_acc}")
        riga_tot = r
        r = S.campo(ws, r, "Costo totale dell'operazione", "=prezzo+costi_accessori", S.EURO, risultato=True)
        self.nome("costo_totale", ws, f"B{riga_tot}")
        riga_inc = r
        r = S.campo(ws, r, "Incidenza dei costi sul prezzo", "=costi_accessori/prezzo", S.PERC, nota="Sotto il sei per cento e' un'operazione leggera, sopra il dieci va capito quale voce pesa.")
        self.nome("incidenza_costi", ws, f"B{riga_inc}")
        riga_esb = r
        r = S.campo(ws, r, "Esborso iniziale, cassa necessaria", "=costo_totale-mutuo_importo", S.EURO, risultato=True, nota="E' il capitale proprio davvero immobilizzato: il denominatore del cash on cash.")
        self.nome("esborso", ws, f"B{riga_esb}")
        riga_mq_prezzo = r
        r = S.campo(ws, r, "Prezzo al metro quadro", "=IF(mq>0,prezzo/mq,0)", S.EURO, nota="Da confrontare con le quotazioni OMI della zona, foglio Fonti.")
        self.nome("prezzo_mq", ws, f"B{riga_mq_prezzo}")
        riga_sconto = r
        r = S.campo(ws, r, "Sconto ottenuto sul richiesto", "=IF(prezzo_richiesto>0,1-prezzo/prezzo_richiesto,0)", S.PERC)

        ws.conditional_formatting.add(
            f"B{riga_inc}",
            CellIsRule(operator="greaterThan", formula=["0.10"], fill=S.FILL_ATTENZIONE),
        )

    # ------------------------------------------------------------------ mutuo
    def foglio_mutuo(self) -> None:
        ws = self.wb.create_sheet("Mutuo")
        ws.sheet_view.showGridLines = False
        S.larghezze_colonne(ws, {"A": 44, "B": 18, "C": 64, "D": 12, "E": 16, "F": 16, "G": 16, "H": 16})
        r = S.titolo(
            ws,
            1,
            "Mutuo ipotecario",
            "Ammortamento alla francese a rata costante. La detrazione degli interessi spetta solo se l'immobile e' abitazione principale con residenza trasferita entro dodici mesi.",
        )

        r = S.sezione(ws, r, "Condizioni del finanziamento")
        riga_imp = r
        r = S.campo(ws, r, "Importo del mutuo", 90000, S.EURO, input_utente=True)
        self.nome("mutuo_importo", ws, f"B{riga_imp}")
        riga_ltv = r
        r = S.campo(ws, r, "Loan to value sul prezzo", "=IF(prezzo>0,mutuo_importo/prezzo,0)", S.PERC, nota="Oltre l'ottanta per cento serve il fondo Consap o una garanzia integrativa, e il tasso peggiora.")
        self.nome("ltv", ws, f"B{riga_ltv}")
        riga_alert = r
        r = S.campo(ws, r, "Serve garanzia esterna", '=IF(ltv>ltv_max,"SI, oltre la soglia","no")', nota="Il fondo di garanzia prima casa Consap copre fino all'ottanta per cento per under 36 con ISEE entro quarantamila euro, misura prorogata al 31 dicembre 2027.")
        riga_tasso = r
        r = S.campo(ws, r, "Tasso annuo nominale", 0.032, S.PERC, input_utente=True, nota="TAN del preventivo. Per il variabile mettere Euribor piu' spread e usare il foglio Scenari. Per sapere se il preventivo e' in linea col mercato: python tools/valuta.py tassi --tasso 0,032")
        self.nome("tasso", ws, f"B{riga_tasso}")
        riga_durata = r
        r = S.campo(ws, r, "Durata in anni", 25, S.NUMERO, input_utente=True)
        self.nome("durata", ws, f"B{riga_durata}")
        r += 1

        r = S.sezione(ws, r, "Oneri del mutuo", secondaria=True)
        riga_sost = r
        r = S.campo(ws, r, "Imposta sostitutiva", '=mutuo_importo*IF(agevolata="SI",sost_prima,sost_ord)', S.EURO, nota="Zero virgola venticinque per cento sulla prima casa, due per cento negli altri casi.")
        self.nome("sostitutiva", ws, f"B{riga_sost}")
        riga_istr = r
        r = S.campo(ws, r, "Spese di istruttoria", 500, S.EURO, input_utente=True, nota="Spesso azzerate nelle offerte commerciali: verificare sul foglio informativo.")
        self.nome("istruttoria", ws, f"B{riga_istr}")
        riga_per = r
        r = S.campo(ws, r, "Perizia", 300, S.EURO, input_utente=True)
        self.nome("perizia", ws, f"B{riga_per}")
        riga_not = r
        r = S.campo(ws, r, "Notaio, atto di mutuo", 1000, S.EURO, input_utente=True, nota="Fattura distinta da quella della compravendita, e detraibile fra gli oneri accessori se prima casa.")
        self.nome("notaio_mutuo", ws, f"B{riga_not}")
        riga_pol_modo = r
        r = S.campo(ws, r, "Polizza incendio, forma del premio", "annuo", input_utente=True, nota="annuo per il premio ricorrente, unico per il premio anticipato in un'unica soluzione, che le banche propongono spesso finanziandolo dentro il mutuo.")
        modo_pol = DataValidation(type="list", formula1='"annuo,unico"', allow_blank=False)
        ws.add_data_validation(modo_pol)
        S.scelta(modo_pol, ws.cell(row=riga_pol_modo, column=2))
        self.nome("polizza_forma", ws, f"B{riga_pol_modo}")
        riga_pol_imp = r
        r = S.campo(ws, r, "Polizza incendio e scoppio, importo", 180, S.EURO, input_utente=True, nota="Se la forma e' annua e' il premio di ogni anno; se e' unica e' il totale per l'intera durata, che sul mercato si osserva attorno a cinque euro ogni mille di capitale.")
        self.nome("polizza_importo", ws, f"B{riga_pol_imp}")
        riga_pol = r
        r = S.campo(ws, r, "Polizza, costo annuo equivalente", '=IF(polizza_forma="unico",IF(durata>0,polizza_importo/durata,0),polizza_importo)', S.EURO_DEC, nota="Il premio unico si ripartisce sulla durata per poterlo confrontare con quello annuo.")
        self.nome("polizza", ws, f"B{riga_pol}")
        r = S.nota_riga(ws, r, "La polizza incendio e scoppio e' obbligatoria per legge ma non e' obbligatorio comprarla dalla banca: il cliente puo' presentarne una reperita altrove, purche' di protezione equivalente, e la banca deve accettarla. Se invece si accetta quella proposta, il cliente ha diritto di sapere quanta provvigione la compagnia paga alla banca. Le polizze vita o sulla perdita dell'impiego non sono mai obbligatorie e la concessione del credito non puo' esservi subordinata.")
        riga_pol_vita = r
        r = S.campo(ws, r, "Polizza vita o impiego, premio annuo", 0, S.EURO, input_utente=True, nota="Facoltativa per legge: la banca non puo' subordinare la concessione del credito alla sua sottoscrizione.")
        self.nome("polizza_facoltativa", ws, f"B{riga_pol_vita}")
        riga_oneri = r
        r = S.campo(ws, r, "Oneri iniziali del mutuo", '=IF(mutuo_importo>0,sostitutiva+istruttoria+perizia+notaio_mutuo+IF(polizza_forma="unico",polizza_importo,0),0)', S.EURO, risultato=True, nota="Comprende il premio unico della polizza, se scelto: e' cassa che esce al rogito e va nel costo dell'operazione.")
        self.nome("oneri_mutuo", ws, f"B{riga_oneri}")
        r += 1

        r = S.sezione(ws, r, "Rata e costo del debito", secondaria=True)
        riga_rata = r
        r = S.campo(ws, r, "Rata mensile", "=IF(mutuo_importo>0,PMT(tasso/12,durata*12,-mutuo_importo),0)", S.EURO_DEC, risultato=True)
        self.nome("rata_mensile", ws, f"B{riga_rata}")
        riga_rata_a = r
        r = S.campo(ws, r, "Rata annua", "=rata_mensile*12", S.EURO)
        self.nome("rata_annua", ws, f"B{riga_rata_a}")
        riga_tot_int = r
        r = S.campo(ws, r, "Interessi totali sull'intera durata", "=rata_mensile*durata*12-mutuo_importo", S.EURO, risultato=True, nota="E' il vero prezzo del debito: confrontarlo con il valore del bonus prima casa mette le cose in prospettiva.")
        self.nome("interessi_totali", ws, f"B{riga_tot_int}")
        riga_costo_deb = r
        r = S.campo(ws, r, "Costo totale del debito", "=interessi_totali+oneri_mutuo+polizza*durata+polizza_facoltativa*durata", S.EURO)
        riga_taeg = r
        r = S.campo(
            ws, r, "Tasso effettivo stimato",
            "=IF(mutuo_importo>0,RATE(durata*12,-(rata_mensile+(polizza+polizza_facoltativa)/12),mutuo_importo-oneri_mutuo)*12,0)",
            S.PERC, nota="Stima interna, non il TAEG di legge: include oneri iniziali e premi assicurativi ripartiti sulle rate.",
        )
        riga_reddito = r
        r = S.campo(ws, r, "Reddito netto mensile del richiedente", 1800, S.EURO, input_utente=True)
        self.nome("reddito_mensile", ws, f"B{riga_reddito}")
        riga_rr = r
        r = S.campo(ws, r, "Rapporto rata reddito", "=IF(reddito_mensile>0,rata_mensile/reddito_mensile,0)", S.PERC, nota="Le banche si fermano di norma a un terzo del reddito netto: sopra il trentacinque per cento la pratica difficilmente passa.")
        ws.conditional_formatting.add(
            f"B{riga_rr}", CellIsRule(operator="greaterThan", formula=["0.35"], fill=S.FILL_ATTENZIONE)
        )
        r += 1

        r = S.sezione(ws, r, "Interessi e detrazione, anno per anno", secondaria=True)
        r = S.nota_riga(ws, r, "La detrazione vale il diciannove per cento degli interessi entro un massimale di quattromila euro riferito all'immobile, quindi da dividere per la quota se il mutuo e' cointestato. Spetta solo sull'abitazione principale: sull'immobile comprato per affittare non spetta, e la colonna resta a zero.")
        intestazione = r
        r = S.intestazioni(ws, r, ["Anno", "Interessi", "Capitale rimborsato", "Debito residuo", "Detrazione IRPEF", "Rata annua", "Beneficio netto"], [10, 16, 18, 16, 18, 16, 16])
        prima_riga_tab = r
        for anno in range(1, MAX_ANNI + 1):
            ws.cell(row=r, column=1, value=anno).number_format = S.NUMERO
            ws.cell(row=r, column=2, value=f'=IF({anno}<=durata,SUMIFS(Ammortamento!$D:$D,Ammortamento!$H:$H,$A{r}),0)').number_format = S.EURO
            ws.cell(row=r, column=3, value=f'=IF({anno}<=durata,SUMIFS(Ammortamento!$E:$E,Ammortamento!$H:$H,$A{r}),0)').number_format = S.EURO
            ws.cell(row=r, column=4, value=f'=IF({anno}<=durata,MAX(0,mutuo_importo-SUM($C${prima_riga_tab}:$C{r})),0)').number_format = S.EURO
            ws.cell(row=r, column=5, value=f'=IF(abitazione_principale="SI",MIN($B{r},detr_max*quota)*detr_aliq,0)').number_format = S.EURO
            ws.cell(row=r, column=6, value=f'=IF({anno}<=durata,rata_annua,0)').number_format = S.EURO
            ws.cell(row=r, column=7, value=f"=$F{r}-$E{r}").number_format = S.EURO
            for col in range(1, 8):
                ws.cell(row=r, column=col).border = S.BORDO
                if col > 1:
                    ws.cell(row=r, column=col).fill = S.FILL_CALCOLO
            r += 1
        ultima_riga_tab = r - 1
        self.nome_intervallo("interessi_anno", ws, f"$B${prima_riga_tab}:$B${ultima_riga_tab}")
        self.nome_intervallo("detrazione_anno", ws, f"$E${prima_riga_tab}:$E${ultima_riga_tab}")
        self.nome_intervallo("rata_anno", ws, f"$F${prima_riga_tab}:$F${ultima_riga_tab}")
        self.nome_intervallo("debito_residuo_anno", ws, f"$D${prima_riga_tab}:$D${ultima_riga_tab}")
        ws.freeze_panes = ws.cell(row=prima_riga_tab, column=1)
        r += 1
        c = ws.cell(row=r, column=1, value="Totale detrazione sull'intera durata")
        c.font = S.ETICHETTA_BOLD
        v = ws.cell(row=r, column=5, value=f"=SUM($E${prima_riga_tab}:$E${ultima_riga_tab})")
        v.number_format = S.EURO
        v.fill = S.FILL_RISULTATO
        v.font = S.ETICHETTA_BOLD

    # ----------------------------------------------------------- ammortamento
    def foglio_ammortamento(self) -> None:
        ws = self.wb.create_sheet("Ammortamento")
        ws.sheet_view.showGridLines = False
        r = S.titolo(ws, 1, "Piano di ammortamento", "Una riga per rata. Le righe oltre la durata restano vuote.", 8)
        r = S.intestazioni(
            ws, r,
            ["N. rata", "Data", "Debito iniziale", "Quota interessi", "Quota capitale", "Rata", "Debito residuo", "Anno"],
            [10, 14, 18, 18, 18, 16, 18, 10],
        )
        prima = r
        for k in range(1, MAX_RATE + 1):
            ws.cell(row=r, column=1, value=k).number_format = S.NUMERO
            ws.cell(row=r, column=2, value=f'=IF($A{r}<=durata*12,EDATE(data_erogazione,$A{r}),"")').number_format = S.DATA
            if k == 1:
                ws.cell(row=r, column=3, value='=IF($A{0}<=durata*12,mutuo_importo,"")'.format(r)).number_format = S.EURO_DEC
            else:
                ws.cell(row=r, column=3, value=f'=IF($A{r}<=durata*12,$G{r-1},"")').number_format = S.EURO_DEC
            ws.cell(row=r, column=4, value=f'=IF($A{r}<=durata*12,$C{r}*tasso/12,"")').number_format = S.EURO_DEC
            ws.cell(row=r, column=5, value=f'=IF($A{r}<=durata*12,$F{r}-$D{r},"")').number_format = S.EURO_DEC
            ws.cell(row=r, column=6, value=f'=IF($A{r}<=durata*12,rata_mensile,"")').number_format = S.EURO_DEC
            ws.cell(row=r, column=7, value=f'=IF($A{r}<=durata*12,MAX(0,$C{r}-$E{r}),"")').number_format = S.EURO_DEC
            ws.cell(row=r, column=8, value=f'=IF($A{r}<=durata*12,INT(($A{r}-1)/12)+1,"")').number_format = S.NUMERO
            for col in range(1, 9):
                ws.cell(row=r, column=col).border = S.BORDO
            r += 1
        ws.freeze_panes = ws.cell(row=prima, column=1)

    # ------------------------------------------------------ simulatore mutuo
    def foglio_simulatore(self) -> None:
        """Simulatore del mutuo con rimborsi volontari e percorso del tasso.

        Il foglio Ammortamento risponde alla domanda semplice, cioe' come si spegne
        il debito se non succede nulla. Questo risponde alle due domande che si
        fanno davvero durante la vita di un mutuo: che cosa succede se verso denaro
        in anticipo, e che cosa succede se il tasso si muove.

        L'impianto riprende il foglio "calcolatore mutuo" di Paolo Coletti, che
        ricalcola la rata mese per mese sul debito residuo effettivo: e' l'unico
        modo di rappresentare correttamente un rimborso volontario, perche' dopo un
        versamento straordinario o la rata scende, a parita' di durata, o la durata
        si accorcia, a parita' di rata, e sono due scelte diverse che si dichiarano
        alla banca.

        Sulla conversione del tasso annuo in mensile convivono due convenzioni. Le
        banche italiane dividono per dodici, ed e' quella usata dal resto del
        workbook; la conversione finanziariamente esatta e' il tasso equivalente
        composto. La differenza e' piccola ma non nulla, e qui si sceglie con una
        cella invece di nasconderla.
        """
        ws = self.wb.create_sheet("Simulatore mutuo")
        ws.sheet_view.showGridLines = False
        S.larghezze_colonne(ws, {"A": 44, "B": 18, "C": 62})
        r = S.titolo(
            ws,
            1,
            "Simulatore: rimborsi volontari e percorso del tasso",
            "Indipendente dal foglio Mutuo, cosi' si puo' provare uno scenario senza toccare l'analisi principale. I valori di partenza sono ripresi da li'.",
        )

        r = S.sezione(ws, r, "Condizioni di partenza")
        riga = r
        r = S.campo(ws, r, "Capitale erogato", "=mutuo_importo", S.EURO, input_utente=True, nota="Ripreso dal foglio Mutuo: si puo' sovrascrivere con un altro importo.")
        self.nome("sim_capitale", ws, f"B{riga}")
        riga = r
        r = S.campo(ws, r, "Durata in mesi", "=durata*12", S.NUMERO, input_utente=True)
        self.nome("sim_mesi", ws, f"B{riga}")
        riga = r
        r = S.campo(ws, r, "Tasso annuo di partenza", "=tasso", S.PERC, input_utente=True)
        self.nome("sim_tasso", ws, f"B{riga}")
        riga = r
        r = S.campo(ws, r, "Conversione del tasso mensile", "banca", input_utente=True, nota="banca per la divisione per dodici, che e' la convenzione dei contratti italiani; composta per il tasso equivalente finanziariamente esatto.")
        conv = DataValidation(type="list", formula1='"banca,composta"', allow_blank=False)
        ws.add_data_validation(conv)
        S.scelta(conv, ws.cell(row=riga, column=2))
        self.nome("sim_convenzione", ws, f"B{riga}")
        r += 1

        r = S.sezione(ws, r, "Percorso del tasso, per il variabile", secondaria=True)
        r = S.nota_riga(ws, r, "Il percorso si descrive a gradini: ogni riga dice da quale mese vale quale variazione rispetto al tasso di partenza, e la variazione e' cumulata, non incrementale. Vale l'ultimo gradino il cui mese e' stato raggiunto, quindi le righe vanno riempite dall'alto verso il basso in ordine di mese crescente e le righe lasciate vuote non hanno alcun effetto. Con il solo primo gradino compilato il comportamento e' quello di un gradino singolo, cioe' quello che il foglio faceva prima.")
        intest_perc = ["Gradino", "Dal mese", "Variazione cumulata"]
        r = S.intestazioni(ws, r, intest_perc, [12, 14, 20])
        prima_perc = r
        # Il primo gradino conserva i due nomi definiti che c'erano prima, cosi' che
        # nulla di quanto li citava debba cambiare: il gradino singolo non e' stato
        # sostituito, e' diventato la prima riga di un percorso.
        valori_iniziali = [(25, 0.0), (None, None), (None, None), (None, None), (None, None), (None, None)]
        for indice, (mese, delta) in enumerate(valori_iniziali, start=1):
            ws.cell(row=r, column=1, value=indice).number_format = S.NUMERO
            cella_mese = ws.cell(row=r, column=2, value=mese)
            cella_mese.number_format = S.NUMERO
            cella_delta = ws.cell(row=r, column=3, value=delta)
            cella_delta.number_format = S.PERC
            for colonna in (1, 2, 3):
                c = ws.cell(row=r, column=colonna)
                c.border = S.BORDO
                c.fill = S.FILL_CALCOLO if colonna == 1 else S.FILL_INPUT
            if indice == 1:
                self.nome("sim_shock_mese", ws, f"B{r}")
                self.nome("sim_shock", ws, f"C{r}")
            r += 1
        ultima_perc = r - 1
        ws.cell(row=prima_perc, column=5, value="Il primo gradino porta i valori di partenza: mese 25 e variazione zero, cioe' nessun effetto.").font = S.NOTA
        ws.cell(row=prima_perc, column=5).alignment = S.SINISTRA
        r += 1
        r = S.nota_riga(ws, r, f"Quanto puo' salire un tasso variabile non e' una domanda di opinione: la serie mensile dell'{P.RISALITE_EURIBOR.indice} pubblicata dalla Banca centrale europea copre {P.RISALITE_EURIBOR.copertura}, e la peggiore finestra di dodici mesi che contiene e' un rialzo di {P.RISALITE_EURIBOR.risalita_12_mesi:.2f} punti, avvenuto fra {P.RISALITE_EURIBOR.finestra_12_mesi.split(',')[0]}. Su ventiquattro mesi la peggiore vale {P.RISALITE_EURIBOR.risalita_24_mesi:.2f} punti, su trentasei {P.RISALITE_EURIBOR.risalita_36_mesi:.2f}. Chi simula un punto percentuale sta simulando un quinto di cio' che e' appena successo. I valori si rileggono con: python tools/valuta.py tassi --risalita")
        r = S.nota_riga(ws, r, f"Come si riempie il percorso per riprodurre quel rialzo, che e' l'esercizio utile prima di firmare un variabile. Primo gradino a un terzo della risalita dal mese in cui si vuole far partire lo scenario, secondo a due terzi quattro mesi dopo, terzo alla risalita intera altri quattro mesi dopo: sui {P.RISALITE_EURIBOR.risalita_12_mesi:.2f} punti della finestra 2022-2023 significa piu' 1,26, piu' 2,52 e piu' 3,78 per cento. Poi si legge la rata massima raggiunta e si decide se quella cifra si puo' sostenere, che e' la sola domanda che conta.")
        r = S.nota_riga(ws, r, f"Una verifica di plausibilita' che il solo scarto non da'. La stessa risalita produce tassi molto diversi a seconda del livello di partenza: quella del 2022-2023 partiva da un Euribor negativo e arrivava al {P.RISALITE_EURIBOR.finestra_12_mesi.split('a ')[-1]} per cento, mentre applicata al livello di oggi, {P.RISALITE_EURIBOR.livello_corrente:.2f} per cento, porterebbe a un tasso che nella serie storica dal 1994 compare soltanto negli anni Novanta, con il massimo al {P.RISALITE_EURIBOR.massimo_storico:.2f} per cento del {P.RISALITE_EURIBOR.periodo_massimo}. Non e' una ragione per escluderlo: e' una ragione per sapere che si sta guardando la coda della distribuzione e non il centro.")
        r = S.nota_riga(ws, r, "Per un tasso fisso si lascia il percorso com'e', con la variazione a zero. Per un variabile si compilano i gradini, e la riga della rata massima dice se lo scenario resta sostenibile: e' la sola domanda che conta davvero prima di firmare un variabile.")
        r = S.nota_riga(ws, r, "Una precisazione che cambia il risultato. Il mutuo a tasso variabile italiano tiene ferma la scadenza e sposta l'aumento sulla rata, quindi per simulare un rialzo dei tassi va scelto sotto l'effetto \"riduci rata\". Con \"riduci durata\" la rata resta quella di partenza e a crescere e' il numero di mesi: e' il funzionamento del mutuo a rata costante e durata variabile, che esiste ma e' meno diffuso, e in caso di forte rialzo puo' allungare il piano oltre la scadenza contrattuale.")
        r += 1

        r = S.sezione(ws, r, "Rimborsi volontari", secondaria=True)
        riga = r
        r = S.campo(ws, r, "Versamento mensile aggiuntivo", 0.0, S.EURO, input_utente=True, nota="Somma che si aggiunge a ogni rata, oltre a quella dovuta.")
        self.nome("sim_extra_mese", ws, f"B{riga}")
        riga = r
        r = S.campo(ws, r, "Versamento una tantum", 0.0, S.EURO, input_utente=True, nota="Una somma sola, per esempio una liquidazione o un premio.")
        self.nome("sim_extra_unico", ws, f"B{riga}")
        riga = r
        r = S.campo(ws, r, "Mese del versamento una tantum", 37, S.NUMERO, input_utente=True)
        self.nome("sim_extra_mese_unico", ws, f"B{riga}")
        riga = r
        r = S.campo(ws, r, "Effetto del rimborso", "riduci durata", input_utente=True, nota="riduci durata tiene ferma la rata e accorcia il piano; riduci rata tiene ferma la scadenza e abbassa la rata. Va dichiarato alla banca: non lo sceglie il foglio.")
        eff = DataValidation(type="list", formula1='"riduci durata,riduci rata"', allow_blank=False)
        ws.add_data_validation(eff)
        S.scelta(eff, ws.cell(row=riga, column=2))
        self.nome("sim_effetto", ws, f"B{riga}")
        r = S.nota_riga(ws, r, "Sul se convenga rimborsare in anticipo la regola e' una sola, e non e' quella che si sente ripetere. Non conta che all'inizio si paghino soprattutto interessi: il denaro e' fungibile e ogni mese la scelta e' la stessa, cioe' estinguere adesso oppure pagare gli interessi per rimandare la decisione. Conviene rimborsare se non si trova un impiego che renda, al netto delle imposte, almeno quanto il tasso del mutuo. Restano due argomenti non finanziari: non avere debiti fa dormire meglio, e un mutuo estinto oggi non si riottiene domani.")
        r += 1

        r = S.sezione(ws, r, "Esito della simulazione", secondaria=True)
        riga_rata0 = r
        r = S.campo(ws, r, "Rata iniziale", "=IF(sim_capitale>0,PMT(sim_tasso_mensile,sim_mesi,-sim_capitale),0)", S.EURO_DEC, risultato=True)
        self.nome("sim_rata_iniziale", ws, f"B{riga_rata0}")
        riga = r
        r = S.campo(ws, r, "Tasso mensile applicato", '=IF(sim_convenzione="composta",(1+sim_tasso)^(1/12)-1,sim_tasso/12)', "0.0000%")
        self.nome("sim_tasso_mensile", ws, f"B{riga}")
        riga_int = r
        r = S.campo(ws, r, "Interessi totali pagati", "=SUM(sim_interessi)", S.EURO, risultato=True)
        riga_base = r
        r = S.campo(ws, r, "Interessi senza rimborsi volontari", "=sim_rata_iniziale*sim_mesi-sim_capitale", S.EURO, nota="Il piano di partenza, a tasso invariato e senza versamenti aggiuntivi.")
        riga_risp = r
        r = S.campo(ws, r, "Interessi risparmiati", f"=B{riga_base}-B{riga_int}", S.EURO, risultato=True, nota="Positivo se i rimborsi volontari, o un tasso in discesa, hanno ridotto il costo del debito.")
        riga_dur = r
        r = S.campo(ws, r, "Durata effettiva in mesi", "=COUNTIF(sim_pagato,\">0\")", S.NUMERO, risultato=True, nota="Piu' bassa della durata contrattuale se i rimborsi hanno accorciato il piano.")
        r = S.campo(ws, r, "Anni risparmiati", f"=(sim_mesi-B{riga_dur})/12", S.NUMERO_DEC)
        r = S.campo(ws, r, "Rata massima raggiunta", "=MAX(sim_rate)", S.EURO_DEC, risultato=True, nota="Solo la rata dovuta, senza i versamenti volontari. Su un tasso variabile e' il numero che dice se lo scenario e' sostenibile, e va confrontato con il reddito.")
        r = S.campo(ws, r, "Massimo esborso in un mese", "=MAX(sim_pagato)", S.EURO_DEC, nota="Comprende anche l'eventuale versamento una tantum.")
        r = S.campo(ws, r, "Totale versato", "=SUM(sim_pagato)", S.EURO)
        r = S.campo(ws, r, "Costo effettivo annuo", "=IFERROR(IRR(sim_flussi)*12,\"non calcolabile\")", S.PERC, nota="Tasso interno dei flussi del solo mutuo: coincide col nominale se non ci sono rimborsi ne' variazioni di tasso.")

        # Il piano modellato si ferma a quarant'anni di rate. Sotto l'effetto che
        # riduce la durata, un rialzo forte del tasso allunga il piano invece di
        # alzare la rata, e con una risalita delle dimensioni di quella del
        # 2022-2023 il piano puo' arrivare al fondo della tabella con il debito non
        # ancora estinto. In quel caso la durata effettiva mostra 480 mesi e gli
        # interessi totali mostrano la somma di cio' che sta in tabella: due numeri
        # veri e due numeri che rispondono a una domanda diversa da quella posta,
        # perche' il piano non si e' chiuso e il costo totale del debito e'
        # maggiore di quanto scritto. Le due righe seguenti esistono per dirlo,
        # invece di lasciarlo dedurre da una durata sospettosamente rotonda.
        riga_res = r
        r = S.campo(ws, r, "Debito residuo alla fine del piano", "=MIN(sim_debito)", S.EURO,
                    nota=f"Deve essere zero. La tabella modella {MAX_RATE} mesi, cioe' quarant'anni di rate.")
        r = S.campo(
            ws, r, "Il piano si chiude",
            f'=IF(B{riga_res}>0.005,"NO: il debito non si estingue entro i {MAX_RATE} mesi modellati, quindi durata effettiva e interessi totali sono troncati e non risolti","SI")',
            risultato=True,
            nota="Con l'effetto che riduce la durata un rialzo forte allunga il piano invece di alzare la rata: se il piano non si chiude, lo scenario va riletto sotto l'effetto che riduce la rata, che e' anche il funzionamento del variabile italiano.",
        )
        ws.cell(row=r - 1, column=2).alignment = S.SINISTRA
        ws.conditional_formatting.add(
            f"B{r-1}:B{r-1}",
            CellIsRule(operator="notEqual", formula=['"SI"'], fill=S.FILL_ATTENZIONE),
        )
        r += 1

        intest = ["Mese", "Data", "Tasso annuo", "Tasso mensile", "Debito iniziale",
                  "Interessi maturati", "Rata dovuta", "Versamento extra", "Totale pagato",
                  "Quota interessi", "Quota capitale", "Debito residuo", "Flusso"]
        r = S.intestazioni(ws, r, intest, [8, 13, 13, 13, 16, 16, 14, 14, 14, 14, 14, 16, 14])
        prima = r

        # Riga zero: erogazione. Serve al tasso interno, che altrimenti non avrebbe
        # l'incasso iniziale contro cui misurare i pagamenti.
        ws.cell(row=r, column=1, value=0).number_format = S.NUMERO
        ws.cell(row=r, column=2, value="=data_erogazione").number_format = S.DATA
        ws.cell(row=r, column=12, value="=sim_capitale").number_format = S.EURO
        ws.cell(row=r, column=13, value="=sim_capitale").number_format = S.EURO
        for col in (6, 9, 10, 11):
            ws.cell(row=r, column=col, value=0).number_format = S.EURO
        for col in range(1, 14):
            ws.cell(row=r, column=col).border = S.BORDO
            ws.cell(row=r, column=col).fill = S.FILL_CALCOLO
        r += 1

        for mese in range(1, MAX_RATE + 1):
            p = r - 1                       # riga precedente
            attivo = f"$L{p}>0.005"         # finche' resta debito, oltre l'arrotondamento
            ws.cell(row=r, column=1, value=mese).number_format = S.NUMERO
            ws.cell(row=r, column=2, value=f"=EDATE(data_erogazione,$A{r})").number_format = S.DATA
            # Il tasso del mese: tasso di partenza piu' l'ultimo gradino raggiunto.
            # La catena di IF viene generata dal basso verso l'alto, cosi' che il
            # primo confronto vero sia quello del gradino piu' avanzato: e' il modo
            # di ottenere "vale l'ultimo gradino raggiunto" senza CERCA, che
            # pretenderebbe una colonna ordinata e si comporterebbe in modo opaco
            # sulle righe lasciate vuote. Il test di ogni gradino include la
            # presenza del mese, percio' una riga vuota non partecipa.
            gradino = "0"
            for riga_g in range(prima_perc, ultima_perc + 1):
                gradino = (
                    f'IF(AND($B${riga_g}<>"",$A{r}>=$B${riga_g}),$C${riga_g},{gradino})'
                )
            ws.cell(row=r, column=3, value=f"=sim_tasso+{gradino}").number_format = S.PERC
            ws.cell(row=r, column=4, value=f'=IF(sim_convenzione="composta",(1+$C{r})^(1/12)-1,$C{r}/12)').number_format = "0.0000%"
            ws.cell(row=r, column=5, value=f"=IF({attivo},$L{p},0)").number_format = S.EURO_DEC
            ws.cell(row=r, column=6, value=f"=$E{r}*$D{r}").number_format = S.EURO_DEC
            # La rata dovuta: costante nella modalita' che accorcia il piano, ricalcolata
            # sui mesi che restano in quella che abbassa la rata. In entrambi i casi non
            # puo' superare quanto serve a chiudere il debito.
            ws.cell(
                row=r, column=7,
                value=(
                    f'=IF(NOT({attivo}),0,MIN($E{r}+$F{r},'
                    f'IF(sim_effetto="riduci rata",'
                    f'PMT($D{r},MAX(sim_mesi-$A{r}+1,1),-$E{r}),sim_rata_iniziale)))'
                ),
            ).number_format = S.EURO_DEC
            ws.cell(
                row=r, column=8,
                value=(
                    f"=IF(NOT({attivo}),0,MIN($E{r}+$F{r}-$G{r},"
                    f"sim_extra_mese+IF($A{r}=sim_extra_mese_unico,sim_extra_unico,0)))"
                ),
            ).number_format = S.EURO_DEC
            ws.cell(row=r, column=9, value=f"=$G{r}+$H{r}").number_format = S.EURO_DEC
            ws.cell(row=r, column=10, value=f"=MIN($I{r},$F{r})").number_format = S.EURO_DEC
            ws.cell(row=r, column=11, value=f"=$I{r}-$J{r}").number_format = S.EURO_DEC
            ws.cell(row=r, column=12, value=f"=MAX(0,$E{r}+$F{r}-$I{r})").number_format = S.EURO_DEC
            ws.cell(row=r, column=13, value=f"=-$I{r}").number_format = S.EURO_DEC
            for col in range(1, 14):
                ws.cell(row=r, column=col).border = S.BORDO
                ws.cell(row=r, column=col).fill = S.FILL_CALCOLO
            r += 1

        ultima = r - 1
        self.nome_intervallo("sim_interessi", ws, f"$J${prima+1}:$J${ultima}")
        self.nome_intervallo("sim_pagato", ws, f"$I${prima+1}:$I${ultima}")
        self.nome_intervallo("sim_rate", ws, f"$G${prima+1}:$G${ultima}")
        self.nome_intervallo("sim_flussi", ws, f"$M${prima}:$M${ultima}")
        # Serve al controllo di chiusura del piano, che sta nella sezione dell'esito
        # e quindi viene scritto prima di questa tabella: il nome definito e' cio'
        # che permette di citarla in avanti senza conoscerne le coordinate.
        self.nome_intervallo("sim_debito", ws, f"$L${prima+1}:$L${ultima}")
        ws.freeze_panes = ws.cell(row=prima, column=3)

    # -------------------------------------------------------------- locazione
    def foglio_locazione(self) -> None:
        ws = self.wb.create_sheet("Locazione")
        ws.sheet_view.showGridLines = False
        S.larghezze_colonne(ws, {"A": 44, "B": 18, "C": 60, "D": 18, "E": 18, "F": 18, "G": 18, "H": 18})
        r = S.titolo(
            ws,
            1,
            "Messa a reddito",
            "Quattro regimi a confronto sullo stesso immobile. Il canone effettivo non e' quello di contratto: sfitto e morosita' vanno tolti prima, non dopo.",
        )

        r = S.sezione(ws, r, "Ipotesi di ricavo")
        riga_canone = r
        r = S.campo(ws, r, "Canone mensile atteso, canone libero", 500, S.EURO, input_utente=True, nota="Da verificare sulle quotazioni OMI di locazione della zona e sugli annunci reali comparabili.")
        self.nome("canone_mese", ws, f"B{riga_canone}")
        riga_canone_conc = r
        r = S.campo(ws, r, "Canone mensile a canone concordato", 420, S.EURO, input_utente=True, nota="Deriva dall'accordo territoriale del Comune, con attestazione di un'associazione firmataria. Di norma inferiore del dieci-venti per cento al libero.")
        self.nome("canone_conc_mese", ws, f"B{riga_canone_conc}")
        riga_brevi = r
        r = S.campo(ws, r, "Ricavi lordi annui da locazione breve", 9000, S.EURO, input_utente=True, nota="Tariffa media per notte moltiplicata per le notti occupate attese. Va stimata sui dati reali della zona, non sul picco di agosto.")
        self.nome("ricavi_brevi", ws, f"B{riga_brevi}")
        riga_sfitto = r
        r = S.campo(ws, r, "Mesi di sfitto attesi all'anno", 1, S.NUMERO_DEC, input_utente=True, nota="Un mese l'anno e' l'assunzione prudenziale standard. Fra un contratto e l'altro il vuoto e' di norma piu' lungo.")
        self.nome("mesi_sfitto", ws, f"B{riga_sfitto}")
        riga_moros = r
        r = S.campo(ws, r, "Accantonamento per morosita'", 0.03, S.PERC, input_utente=True, nota="Uno sfratto per morosita' richiede un anno e mezzo e nel frattempo il canone non entra ma le imposte si pagano lo stesso.")
        self.nome("morosita_pct", ws, f"B{riga_moros}")
        r += 1

        r = S.sezione(ws, r, "Costi ricorrenti a carico del proprietario", secondaria=True)
        riga_cond = r
        r = S.campo(ws, r, "Spese condominiali annue totali", 1200, S.EURO, input_utente=True, nota="Dal consuntivo condominiale, non dalla stima dell'agenzia. Chiedere anche il verbale dell'ultima assemblea per i lavori deliberati.")
        self.nome("condominio", ws, f"B{riga_cond}")
        riga_quota_cond = r
        r = S.campo(ws, r, "Quota a carico del proprietario", 0.4, S.PERC, input_utente=True, nota="La ripartizione fra proprietario e inquilino segue la tabella degli oneri accessori: straordinaria al proprietario, ordinaria all'inquilino.")
        self.nome("quota_condominio", ws, f"B{riga_quota_cond}")
        riga_manut = r
        r = S.campo(ws, r, "Manutenzione ordinaria, quota del valore", 0.01, S.PERC, input_utente=True, nota="Un per cento del valore l'anno e' la regola empirica: caldaia, infissi, elettrodomestici, tinteggiature fra un inquilino e l'altro.")
        self.nome("manut_pct", ws, f"B{riga_manut}")
        riga_ass = r
        r = S.campo(ws, r, "Assicurazione fabbricato", 200, S.EURO, input_utente=True)
        self.nome("assicurazione", ws, f"B{riga_ass}")
        riga_imu_a = r
        r = S.campo(ws, r, "Aliquota IMU deliberata dal Comune", P.IMU.aliquota_base, S.PERC, input_utente=True, nota="Da leggere nella delibera comunale dell'anno, non dal valore base di legge.")
        self.nome("imu_aliquota", ws, f"B{riga_imu_a}")
        riga_gest = r
        r = S.campo(ws, r, "Gestione affidata a terzi, quota del canone", 0.0, S.PERC, input_utente=True, nota="Un property manager costa il dieci per cento sul lungo periodo, il venti sul breve.")
        self.nome("gestione_pct", ws, f"B{riga_gest}")
        riga_costi_brevi = r
        r = S.campo(ws, r, "Costi variabili della locazione breve", 0.15, S.PERC, input_utente=True, nota="Pulizie, biancheria, utenze, commissioni dei portali, consumabili. Sono sui ricavi, non sull'utile.")
        self.nome("costi_brevi_pct", ws, f"B{riga_costi_brevi}")
        riga_marg = r
        r = S.campo(ws, r, "Aliquota marginale IRPEF del contribuente", 0.33, S.PERC, input_utente=True, nota="Lo scaglione in cui cade l'ultimo euro di reddito: 23 per cento fino a ventottomila euro, 33 fino a cinquantamila, 43 oltre. Serve solo al confronto con il regime ordinario.")
        self.nome("irpef_marginale", ws, f"B{riga_marg}")
        r += 1

        r = S.sezione(ws, r, "Costo figurativo del proprio tempo", secondaria=True)
        r = S.nota_riga(ws, r, "Gestire un immobile locato costa tempo: registrare e rinnovare i contratti, seguire le assemblee, cercare l'artigiano il 23 dicembre, selezionare gli inquilini, rincorrere un pagamento in ritardo. E' un costo reale che quasi nessuna analisi mette a bilancio, e non metterlo significa confrontare il rendimento di un immobile con quello di un investimento finanziario che di tempo non ne chiede. Va dato un valore alle proprie ore e va calcolato.")
        riga_ore = r
        r = S.campo(ws, r, "Ore all'anno dedicate alla gestione", 30, S.NUMERO, input_utente=True, nota="Trenta ore l'anno e' l'ordine di grandezza di una locazione lunga che va liscia. Una locazione breve gestita in proprio sta su un altro ordine di grandezza, dalle duecento ore in su.")
        self.nome("ore_gestione", ws, f"B{riga_ore}")
        riga_valore_ora = r
        r = S.campo(ws, r, "Valore di un'ora del proprio tempo", 0, S.EURO, input_utente=True, nota="Zero per escludere questa voce dal conto. Un riferimento onesto e' il proprio costo orario netto, oppure quanto si pagherebbe qualcuno per farlo al posto proprio.")
        self.nome("valore_ora", ws, f"B{riga_valore_ora}")
        riga_costo_tempo = r
        r = S.campo(ws, r, "Costo figurativo annuo", "=ore_gestione*valore_ora", S.EURO, nota="Entra nel conto economico come una qualsiasi altra voce di costo. Se resta a zero il modello si comporta come prima.")
        self.nome("costo_tempo", ws, f"B{riga_costo_tempo}")
        riga_coef = r
        r = S.campo(ws, r, "Moltiplicatore del tempo per la locazione breve", 6, S.NUMERO_DEC, input_utente=True, nota="Quante volte il tempo della locazione lunga. La locazione breve non e' un investimento passivo: e' piu' vicina a un mestiere, e va confrontata con gli altri regimi tenendone conto.")
        self.nome("coefficiente_tempo_breve", ws, f"B{riga_coef}")
        r += 1

        r = S.sezione(ws, r, "Confronto fra regimi", secondaria=True)
        r = S.intestazioni(
            ws, r,
            ["Voce", "Cedolare libero", "Cedolare concordato", "IRPEF ordinaria", "Locazione breve"],
            [42, 20, 20, 20, 20],
        )

        def riga_conf(etichetta, f_lib, f_conc, f_irp, f_brev, formato=S.EURO, risultato=False):
            """Scrive una voce del conto economico e restituisce la riga occupata.

            Il valore di ritorno e' il presidio contro il difetto piu' insidioso di
            questo generatore. La versione precedente calcolava gli indici delle
            righe come base piu' una costante scritta a mano accanto a ogni
            chiamata: inserire una voce in mezzo al conto economico spostava di uno
            tutte le righe successive e lasciava le costanti dov'erano, quindi il
            reddito operativo netto sommava un intervallo traslato e l'utile netto
            leggeva la riga sbagliata. Il file si apriva, nessuna cella andava in
            errore e i valori di sintesi restavano plausibili, che e' esattamente il
            difetto che una revisione a video non vede. Chiedendo la riga a chi la
            scrive, l'allineamento smette di essere una cosa da ricordare.
            """
            nonlocal r
            scritta = r
            e = ws.cell(row=r, column=1, value=etichetta)
            e.font = S.ETICHETTA_BOLD if risultato else S.ETICHETTA
            e.alignment = S.SINISTRA
            for i, f in enumerate([f_lib, f_conc, f_irp, f_brev], start=2):
                c = ws.cell(row=r, column=i, value=f)
                c.number_format = formato
                c.border = S.BORDO
                c.fill = S.FILL_RISULTATO if risultato else S.FILL_CALCOLO
                if risultato:
                    c.font = S.ETICHETTA_BOLD
            r += 1
            return scritta

        riga_pot = riga_conf("Canone o ricavo lordo annuo", "=canone_mese*12", "=canone_conc_mese*12", "=canone_mese*12", "=ricavi_brevi")
        riga_sf = riga_conf("Perdita per sfitto", "=-canone_mese*mesi_sfitto", "=-canone_conc_mese*mesi_sfitto", "=-canone_mese*mesi_sfitto", "=0",)
        riga_mo = riga_conf(
            "Accantonamento morosita'",
            f"=-(B{riga_pot}+B{riga_sf})*morosita_pct",
            f"=-(C{riga_pot}+C{riga_sf})*morosita_pct",
            f"=-(D{riga_pot}+D{riga_sf})*morosita_pct",
            "=0",
        )
        riga_eff = riga_conf(
            "Ricavo effettivo",
            f"=SUM(B{riga_pot}:B{riga_mo})",
            f"=SUM(C{riga_pot}:C{riga_mo})",
            f"=SUM(D{riga_pot}:D{riga_mo})",
            f"=SUM(E{riga_pot}:E{riga_mo})",
        )
        # Il blocco dei costi. Si catturano la prima e l'ultima riga, e il reddito
        # operativo netto somma fra le due invece che fra due costanti: una voce di
        # costo aggiunta qui in mezzo entra da sola nella somma, che e' il solo modo
        # di rendere sicura un'operazione che prima richiedeva di ricordarsi di
        # aggiornare quattro numeri sparsi.
        riga_primo_costo = riga_conf("Spese condominiali a carico", "=-condominio*quota_condominio", "=-condominio*quota_condominio", "=-condominio*quota_condominio", "=-condominio")
        riga_conf("Manutenzione ordinaria", "=-prezzo*manut_pct", "=-prezzo*manut_pct", "=-prezzo*manut_pct", "=-prezzo*manut_pct")
        riga_conf("Assicurazione", "=-assicurazione", "=-assicurazione", "=-assicurazione", "=-assicurazione")
        riga_conf(
            "Costo figurativo del proprio tempo",
            "=-costo_tempo", "=-costo_tempo", "=-costo_tempo",
            "=-costo_tempo*coefficiente_tempo_breve",
        )
        riga_conf(
            "Accantonamento ristrutturazione di fine ciclo",
            "=-accantonamento_ristrutturazione", "=-accantonamento_ristrutturazione",
            "=-accantonamento_ristrutturazione", "=-accantonamento_ristrutturazione",
        )
        riga_conf(
            "IMU",
            "=-rendita*riv_rendita*imu_molt*imu_aliquota",
            "=-rendita*riv_rendita*imu_molt*imu_aliquota*imu_conc",
            "=-rendita*riv_rendita*imu_molt*imu_aliquota",
            "=-rendita*riv_rendita*imu_molt*imu_aliquota",
        )
        riga_ultimo_costo = riga_conf(
            "Gestione e costi variabili",
            f"=-B{riga_eff}*gestione_pct",
            f"=-C{riga_eff}*gestione_pct",
            f"=-D{riga_eff}*gestione_pct",
            f"=-E{riga_eff}*(gestione_pct+costi_brevi_pct)",
        )
        riga_noi = riga_conf(
            "Reddito operativo netto",
            f"=B{riga_eff}+SUM(B{riga_primo_costo}:B{riga_ultimo_costo})",
            f"=C{riga_eff}+SUM(C{riga_primo_costo}:C{riga_ultimo_costo})",
            f"=D{riga_eff}+SUM(D{riga_primo_costo}:D{riga_ultimo_costo})",
            f"=E{riga_eff}+SUM(E{riga_primo_costo}:E{riga_ultimo_costo})",
            risultato=True,
        )
        riga_imp = riga_conf(
            "Imposta sul reddito da locazione",
            f"=-B{riga_eff}*ced_libero",
            f"=-C{riga_eff}*ced_conc",
            f"=-(D{riga_eff}*(1-abbatt_ord)*(irpef_marginale+addizionali)+MAX(D{riga_eff}*reg_loc,reg_loc_min)/2)",
            f"=-E{riga_eff}*ced_breve1",
        )
        riga_utile = riga_conf(
            "Utile netto annuo",
            f"=B{riga_noi}+B{riga_imp}",
            f"=C{riga_noi}+C{riga_imp}",
            f"=D{riga_noi}+D{riga_imp}",
            f"=E{riga_noi}+E{riga_imp}",
            risultato=True,
        )
        riga_conf(
            "Rendimento lordo sul prezzo",
            f"=B{riga_pot}/prezzo", f"=C{riga_pot}/prezzo", f"=D{riga_pot}/prezzo", f"=E{riga_pot}/prezzo",
            formato=S.PERC,
        )
        riga_conf(
            "Rendimento netto sul costo totale",
            f"=B{riga_utile}/costo_totale", f"=C{riga_utile}/costo_totale",
            f"=D{riga_utile}/costo_totale", f"=E{riga_utile}/costo_totale",
            formato=S.PERC, risultato=True,
        )
        r += 1
        riga_scelta = r
        r = S.campo(
            ws, r, "Regime scelto per la proiezione",
            "cedolare_libero", input_utente=True,
            nota="Uno fra cedolare_libero, cedolare_concordato, irpef_ordinario, breve. Alimenta il foglio Cash flow.",
        )
        dv = DataValidation(type="list", formula1='"cedolare_libero,cedolare_concordato,irpef_ordinario,breve"', allow_blank=False)
        ws.add_data_validation(dv)
        S.scelta(dv, ws.cell(row=riga_scelta, column=2))
        self.nome("regime_scelto", ws, f"B{riga_scelta}")

        riga_sel_noi = r
        r = S.campo(
            ws, r, "Reddito operativo netto del regime scelto",
            f'=IF(regime_scelto="cedolare_libero",B{riga_noi},IF(regime_scelto="cedolare_concordato",C{riga_noi},IF(regime_scelto="irpef_ordinario",D{riga_noi},E{riga_noi})))',
            S.EURO, risultato=True,
        )
        self.nome("noi_annuo", ws, f"B{riga_sel_noi}")
        riga_sel_utile = r
        r = S.campo(
            ws, r, "Utile netto annuo del regime scelto",
            f'=IF(regime_scelto="cedolare_libero",B{riga_utile},IF(regime_scelto="cedolare_concordato",C{riga_utile},IF(regime_scelto="irpef_ordinario",D{riga_utile},E{riga_utile})))',
            S.EURO, risultato=True,
        )
        self.nome("utile_locazione", ws, f"B{riga_sel_utile}")
        riga_sel_canone = r
        r = S.campo(
            ws, r, "Ricavo lordo annuo del regime scelto",
            f'=IF(regime_scelto="cedolare_libero",B{riga_pot},IF(regime_scelto="cedolare_concordato",C{riga_pot},IF(regime_scelto="irpef_ordinario",D{riga_pot},E{riga_pot})))',
            S.EURO,
        )
        self.nome("ricavo_lordo", ws, f"B{riga_sel_canone}")
        riga_sel_eff = r
        r = S.campo(
            ws, r, "Ricavo effettivo del regime scelto",
            f'=IF(regime_scelto="cedolare_libero",B{riga_eff},IF(regime_scelto="cedolare_concordato",C{riga_eff},IF(regime_scelto="irpef_ordinario",D{riga_eff},E{riga_eff})))',
            S.EURO,
        )
        self.nome("ricavo_effettivo", ws, f"B{riga_sel_eff}")
        r += 1

        r = S.sezione(ws, r, "Vincoli e adempimenti del regime scelto", secondaria=True)
        for testo in [
            "Cedolare secca: si opta in sede di registrazione o per le annualita' successive, sostituisce IRPEF, addizionali, registro e bollo, ma comporta la rinuncia all'aggiornamento ISTAT del canone per tutta la durata dell'opzione.",
            "Canone concordato: il canone e' vincolato dall'accordo territoriale del Comune e serve l'attestazione di conformita' rilasciata da un'associazione firmataria. In cambio si ha la cedolare al dieci per cento e, nella maggior parte dei Comuni, lo sconto del venticinque per cento sull'IMU.",
            "IRPEF ordinaria: conviene solo con redditi bassi o con molti oneri da far valere. Il canone concorre al reddito complessivo e puo' spostare l'intero reddito in uno scaglione superiore.",
            "Locazione breve: dal 2026 il regime copre al massimo due unita' per periodo d'imposta, dalla terza scatta la presunzione di impresa con obbligo di partita IVA. Servono il codice identificativo nazionale in ogni annuncio, la comunicazione alla questura degli alloggiati, i dispositivi di sicurezza obbligatori dal 2025 e il rispetto dei regolamenti comunali e condominiali.",
        ]:
            r = S.nota_riga(ws, r, testo)

    # ---------------------------------------------------------------- cashflow
    def foglio_cashflow(self) -> None:
        ws = self.wb.create_sheet("Cash flow")
        ws.sheet_view.showGridLines = False
        S.larghezze_colonne(ws, {"A": 40, "B": 18, "C": 58})
        r = S.titolo(
            ws,
            1,
            "Proiezione del flusso di cassa",
            "Un anno per riga. L'ultimo anno incorpora la vendita: valore rivalutato, meno debito residuo, meno costi di uscita e imposta sulla plusvalenza.",
        )

        r = S.sezione(ws, r, "Assunzioni della proiezione")
        riga_or = r
        r = S.campo(ws, r, "Orizzonte di analisi in anni", 25, S.NUMERO, input_utente=True, nota="Quanti anni si pensa di tenere l'immobile. Oltre i costi di ingresso pesano meno.")
        self.nome("orizzonte", ws, f"B{riga_or}")
        riga_riv = r
        r = S.campo(ws, r, "Rivalutazione nominale annua dell'immobile", 0.02, S.PERC, input_utente=True, nota="In termini reali il mattone italiano e' rimasto sostanzialmente fermo per vent'anni: mettere qui l'inflazione e' gia' un'ipotesi ottimistica.")
        self.nome("riv_immobile", ws, f"B{riga_riv}")
        riga_ind = r
        r = S.campo(ws, r, "Indicizzazione annua del canone", 0.0, S.PERC, input_utente=True, nota="Con la cedolare secca l'aggiornamento ISTAT non si puo' applicare: lasciare zero.")
        self.nome("indicizzazione", ws, f"B{riga_ind}")
        riga_data = r
        r = S.campo(ws, r, "Data di erogazione del mutuo", date(2026, 9, 1), S.DATA, input_utente=True)
        self.nome("data_erogazione", ws, f"B{riga_data}")
        riga_ristr = r
        r = S.campo(ws, r, "Accantonamento annuo per ristrutturazione di fine ciclo", "=prezzo*ristrutt_pct/ristrutt_anni", S.EURO, nota="Un rifacimento completo ogni quarant'anni, spalmato. Ignorarlo e' l'errore piu' comune nelle valutazioni ottimistiche.")
        self.nome("accantonamento_ristrutturazione", ws, f"B{riga_ristr}")
        r += 2

        # I costi operativi della colonna D comprendono gia' l'accantonamento per la
        # ristrutturazione, perche' quest'ultimo e' una riga del conto economico nel
        # foglio Locazione ed entra quindi nel reddito operativo netto. Una colonna
        # separata qui lo conterebbe due volte.
        intest = [
            "Anno", "Ricavo lordo", "Ricavo effettivo", "Costi operativi",
            "Imposta sul reddito", "Rata mutuo", "Detrazione interessi", "Cash flow netto",
            "Cash flow cumulato", "Valore immobile", "Debito residuo", "Patrimonio netto", "Flusso per TIR",
        ]
        r = S.intestazioni(ws, r, intest, [8, 16, 16, 18, 16, 16, 16, 18, 18, 18, 16, 18, 18])
        prima = r

        # Anno zero: esce solo l'esborso iniziale.
        ws.cell(row=r, column=1, value=0).number_format = S.NUMERO
        for col in range(2, 10):
            ws.cell(row=r, column=col, value=0).number_format = S.EURO
        ws.cell(row=r, column=9, value="=-esborso").number_format = S.EURO
        ws.cell(row=r, column=10, value="=prezzo").number_format = S.EURO
        ws.cell(row=r, column=11, value="=mutuo_importo").number_format = S.EURO
        ws.cell(row=r, column=12, value="=prezzo-mutuo_importo").number_format = S.EURO
        ws.cell(row=r, column=13, value="=-esborso").number_format = S.EURO
        for col in range(1, 14):
            ws.cell(row=r, column=col).border = S.BORDO
            ws.cell(row=r, column=col).fill = S.FILL_CALCOLO
        r += 1

        for anno in range(1, ORIZZONTE_MAX + 1):
            attivo = f"$A{r}<=orizzonte"
            ws.cell(row=r, column=1, value=anno).number_format = S.NUMERO
            ws.cell(row=r, column=2, value=f'=IF({attivo},ricavo_lordo*(1+indicizzazione)^($A{r}-1),0)').number_format = S.EURO
            ws.cell(row=r, column=3, value=f'=IF({attivo},ricavo_effettivo*(1+indicizzazione)^($A{r}-1),0)').number_format = S.EURO
            ws.cell(row=r, column=4, value=f'=IF({attivo},-(ricavo_effettivo-noi_annuo)*(1+infl)^($A{r}-1),0)').number_format = S.EURO
            ws.cell(row=r, column=5, value=f'=IF({attivo},-(noi_annuo-utile_locazione)*(1+indicizzazione)^($A{r}-1),0)').number_format = S.EURO
            ws.cell(row=r, column=6, value=f'=IF(AND({attivo},$A{r}<=durata),-rata_annua,0)').number_format = S.EURO
            ws.cell(row=r, column=7, value=f'=IF({attivo},INDEX(detrazione_anno,MIN($A{r},{MAX_ANNI})),0)').number_format = S.EURO
            ws.cell(row=r, column=8, value=f"=SUM($C{r}:$G{r})").number_format = S.EURO
            ws.cell(row=r, column=9, value=f"=$I{r-1}+$H{r}").number_format = S.EURO
            ws.cell(row=r, column=10, value=f'=IF({attivo},prezzo*(1+riv_immobile)^$A{r},0)').number_format = S.EURO
            ws.cell(row=r, column=11, value=f'=IF({attivo},INDEX(debito_residuo_anno,MIN($A{r},{MAX_ANNI})),0)').number_format = S.EURO
            ws.cell(row=r, column=12, value=f'=IF({attivo},$J{r}-$K{r},0)').number_format = S.EURO
            ws.cell(
                row=r, column=13,
                value=(
                    f'=IF({attivo},$H{r}+IF($A{r}=orizzonte,'
                    f'$J{r}*(1-costi_vendita)-$K{r}-IF($A{r}<plus_anni,MAX(0,$J{r}-costo_totale)*plus_aliq,0),0),0)'
                ),
            ).number_format = S.EURO
            for col in range(1, 14):
                ws.cell(row=r, column=col).border = S.BORDO
                ws.cell(row=r, column=col).fill = S.FILL_CALCOLO
            r += 1

        ultima = r - 1
        self.nome_intervallo("flussi_tir", ws, f"$M${prima}:$M${ultima}")
        self.nome_intervallo("flussi_tir_dal_primo", ws, f"$M${prima+1}:$M${ultima}")
        self.nome_intervallo("cash_flow_annuo_serie", ws, f"$H${prima+1}:$H${ultima}")
        self.nome("cash_flow_primo_anno", ws, f"H{prima+1}")
        ws.freeze_panes = ws.cell(row=prima, column=2)

        ws.conditional_formatting.add(
            f"H{prima+1}:H{ultima}",
            CellIsRule(operator="lessThan", formula=["0"], fill=S.FILL_ATTENZIONE),
        )

    # ---------------------------------------------------------------- metriche
    def foglio_metriche(self) -> None:
        ws = self.wb.create_sheet("Metriche")
        ws.sheet_view.showGridLines = False
        S.larghezze_colonne(ws, {"A": 46, "B": 20, "C": 74})
        r = S.titolo(
            ws,
            1,
            "Indicatori di sintesi",
            "Tutti i rendimenti sono calcolati sul costo totale dell'operazione, non sul prezzo: usare il prezzo gonfia il risultato di circa un decimo.",
        )

        r = S.sezione(ws, r, "Capitale impiegato")
        r = S.campo(ws, r, "Prezzo", "=prezzo", S.EURO)
        r = S.campo(ws, r, "Costi accessori", "=costi_accessori", S.EURO)
        r = S.campo(ws, r, "Costo totale", "=costo_totale", S.EURO, risultato=True)
        r = S.campo(ws, r, "Capitale proprio immobilizzato", "=esborso", S.EURO, risultato=True)
        r = S.campo(ws, r, "Leva finanziaria", "=IF(esborso>0,costo_totale/esborso,0)", S.NUMERO_DEC, nota="Quante volte il capitale proprio: la leva amplifica sia il rendimento sia la perdita.")
        r += 1

        r = S.sezione(ws, r, "Redditivita' corrente", secondaria=True)
        r = S.campo(ws, r, "Rendimento lordo", "=ricavo_lordo/prezzo", S.PERC, nota="Canone annuo diviso prezzo. E' il numero che si legge negli annunci, ed e' il meno informativo.")
        r = S.campo(ws, r, "Rendimento netto", "=utile_locazione/costo_totale", S.PERC, risultato=True, nota="Utile dopo costi e imposte, sul costo totale. Fra il lordo e il netto si perdono di norma due punti e mezzo.")
        r = S.campo(ws, r, "Cap rate", "=noi_annuo/costo_totale", S.PERC, nota="Reddito operativo netto sul costo totale, prima delle imposte sul reddito e del mutuo. Serve a confrontare immobili fra loro, a prescindere da come sono finanziati.")
        riga_coc = r
        r = S.campo(ws, r, "Cash on cash", "=IF(esborso>0,cash_flow_primo_anno/esborso,0)", S.PERC, risultato=True, nota="Cassa netta del primo anno sul capitale proprio. Con la leva puo' essere negativo anche se l'operazione e' sana.")
        riga_dscr = r
        r = S.campo(ws, r, "Debt service coverage ratio", "=IF(rata_annua>0,noi_annuo/rata_annua,\"n.d.\")", S.NUMERO_DEC, nota="Sotto 1 il reddito non copre la rata e la differenza esce dalla tasca del proprietario ogni mese.")
        r = S.campo(ws, r, "Cash flow del primo anno", "=cash_flow_primo_anno", S.EURO, risultato=True)
        r = S.campo(ws, r, "Cash flow mensile", "=cash_flow_primo_anno/12", S.EURO_DEC)
        r += 1

        r = S.sezione(ws, r, "Redditivita' sull'intero orizzonte", secondaria=True)
        r = S.campo(ws, r, "Tasso interno di rendimento", "=IFERROR(IRR(flussi_tir),\"non calcolabile\")", S.PERC, risultato=True, nota="Include l'uscita finale. E' il numero da confrontare con il rendimento atteso di un portafoglio alternativo.")
        r = S.campo(ws, r, "Valore attuale netto", "=IFERROR(NPV(tasso_sconto,flussi_tir_dal_primo)-esborso,\"non calcolabile\")", S.EURO, nota="Positivo se l'operazione batte il tasso di sconto scelto nei parametri.")
        r = S.campo(ws, r, "Cassa cumulata a fine orizzonte", "=SUM(cash_flow_annuo_serie)", S.EURO)
        r = S.campo(ws, r, "Multiplo sul capitale proprio", "=IF(esborso>0,SUM(flussi_tir)/esborso+1,0)", S.NUMERO_DEC)
        r = S.campo(ws, r, "Anni per rientrare del capitale proprio", "=IF(cash_flow_primo_anno>0,esborso/cash_flow_primo_anno,\"mai, con questo cash flow\")", S.NUMERO_DEC)
        r += 1

        r = S.sezione(ws, r, "Effetto dell'inflazione: dal nominale al reale", secondaria=True)
        r = S.nota_riga(ws, r, "Tutti i rendimenti sopra sono nominali, cioe' espressi in euro del futuro. Il rendimento che conta e' quello reale, perche' un investimento non serve a possedere piu' euro ma a comprare piu' cose. La conversione non e' una sottrazione: e' il rapporto fra il potere d'acquisto finale e quello iniziale, e le due forme divergono in modo misurabile.", 3)
        riga_infl_r = r
        r = S.campo(ws, r, "Inflazione attesa", "=infl", S.PERC, nota="Dal foglio Parametri. Si tara sui dati correnti con: python tools/valuta.py indicatori")
        riga_rn = r
        r = S.campo(ws, r, "Rendimento netto nominale", "=utile_locazione/costo_totale", S.PERC)
        riga_rr = r
        r = S.campo(ws, r, "Rendimento netto reale", f"=(1+B{riga_rn})/(1+infl)-1", S.PERC, risultato=True,
                    nota="Equazione di Fisher in forma esatta: (1+r)/(1+i)-1. E' il rendimento in potere d'acquisto, cioe' quello che dice se l'operazione arricchisce o impoverisce.")
        riga_err = r
        r = S.campo(ws, r, "Errore della sottrazione r meno i", f"=(B{riga_rn}-infl)-B{riga_rr}", "0.0000%",
                    nota="Quanto si sbaglia usando la forma approssimata invece di quella esatta. Piccolo su un anno, si compone sull'orizzonte.")
        riga_tir_n = r
        r = S.campo(ws, r, "Tasso interno di rendimento nominale", "=IFERROR(IRR(flussi_tir),\"non calcolabile\")", S.PERC)
        r = S.campo(ws, r, "Tasso interno di rendimento reale", f"=IFERROR((1+B{riga_tir_n})/(1+infl)-1,\"non calcolabile\")", S.PERC, risultato=True,
                    nota="E' il numero da confrontare con il rendimento reale di un portafoglio alternativo, non con quello nominale.")
        r += 1

        r = S.nota_riga(ws, r, "L'inflazione non agisce nello stesso verso su tutte le componenti dell'operazione, e in un acquisto a leva i versi si compensano solo in parte. Le quattro righe seguenti la scompongono voce per voce.", 3)
        riga_can_r = r
        r = S.campo(ws, r, "Variazione reale annua del canone", "=(1+indicizzazione)/(1+infl)-1", S.PERC,
                    nota="Negativa quando l'indicizzazione non copre l'inflazione. Con la cedolare secca l'aggiornamento ISTAT non si puo' applicare, quindi l'indicizzazione e' zero e il canone perde in termini reali quanto l'inflazione intera.")
        riga_riv_r = r
        r = S.campo(ws, r, "Rivalutazione reale annua dell'immobile", "=(1+riv_immobile)/(1+infl)-1", S.PERC,
                    nota="Zero se la rivalutazione nominale assunta e' pari all'inflazione, che e' esattamente cio' che il mercato residenziale italiano ha fatto negli ultimi vent'anni: rivalutazione nominale, nessuna reale.")
        riga_deb_n = r
        r = S.campo(ws, r, "Debito residuo a fine orizzonte, nominale", f"=IF(mutuo_importo>0,INDEX(debito_residuo_anno,MIN(orizzonte,{MAX_ANNI})),0)", S.EURO)
        riga_deb_r = r
        r = S.campo(ws, r, "Lo stesso debito in euro di oggi", f"=B{riga_deb_n}/(1+infl)^orizzonte", S.EURO)
        riga_regalo = r
        r = S.campo(ws, r, "Sconto che l'inflazione fa sul debito", f"=B{riga_deb_n}-B{riga_deb_r}", S.EURO, risultato=True,
                    nota="Il debito e' un importo nominale: l'inflazione lo eroda a favore di chi lo ha contratto. E' un trasferimento reale di ricchezza dalla banca al mutuatario, e vale solo a tasso fisso.")
        riga_rata_r = r
        r = S.campo(ws, r, "Rata annua dell'ultimo anno, in euro di oggi", "=IF(mutuo_importo>0,rata_annua/(1+infl)^MIN(orizzonte,durata),0)", S.EURO,
                    nota="La rata di un fisso e' un numero nominale fermo, quindi si alleggerisce ogni anno in termini reali. Confrontarla con la rata annua di oggi rende visibile l'effetto.")
        r += 1

        r = S.sezione(ws, r, "Potere d'acquisto a fine orizzonte", secondaria=True)
        riga_val_n = r
        r = S.campo(ws, r, "Valore dell'immobile a fine orizzonte, nominale", "=prezzo*(1+riv_immobile)^orizzonte", S.EURO)
        r = S.campo(ws, r, "Lo stesso valore in euro di oggi", f"=B{riga_val_n}/(1+infl)^orizzonte", S.EURO, risultato=True,
                    nota="Se coincide col prezzo pagato, la rivalutazione nominale ha solo tenuto il passo dell'inflazione e in termini reali l'immobile non ha guadagnato nulla.")
        r = S.campo(ws, r, "Cassa cumulata a fine orizzonte, nominale", "=SUM(cash_flow_annuo_serie)", S.EURO)
        r = S.campo(ws, r, "Patrimonio netto reale a fine orizzonte", f"=B{riga_val_n}/(1+infl)^orizzonte-B{riga_deb_r}", S.EURO, risultato=True,
                    nota="Valore dell'immobile meno debito residuo, entrambi in potere d'acquisto di oggi. E' la grandezza da confrontare con il capitale proprio immobilizzato all'inizio.")
        riga_conf_reale = r
        r = S.campo(ws, r, "Confronto col capitale proprio immobilizzato", f"=B{riga_conf_reale-1}-esborso", S.EURO,
                    nota="Positivo se, in potere d'acquisto, a fine orizzonte si ha piu' di quanto si e' messo. Non tiene conto della cassa versata o incassata nel frattempo, che sta nella riga sopra.")
        r += 1

        r = S.sezione(ws, r, "Il costo dell'indicizzazione rinunciata", secondaria=True)
        r = S.nota_riga(ws, r, "La cedolare secca sostituisce l'IRPEF sul canone con un'aliquota fissa, ed e' quasi sempre conveniente guardando la sola aliquota. In cambio, per l'articolo 3 comma 11 del d.lgs. 23 del 2011, chi la opta rinuncia all'aggiornamento ISTAT del canone per la durata dell'opzione. Le righe seguenti quantificano i due lati, perche' la scelta si fa di solito guardando solo il primo.", 3)
        # I due fattori di attualizzazione di una rendita crescente. Stanno in celle
        # visibili e non dentro la formula finale, perche' un risultato di cui non si
        # possono riconoscere i pezzi non si puo' contestare. La forma chiusa e'
        # q(1-q^n)/((1+x)(1-q)) con q=(1+x)/(1+s), e il caso q=1, cioe' crescita pari
        # al tasso di sconto, va trattato a parte perche' annullerebbe il denominatore.
        riga_f_infl = r
        r = S.campo(
            ws, r, "Fattore di un canone indicizzato all'inflazione piena",
            "=IF(ABS((1+infl)/(1+tasso_sconto)-1)<0.0000001,orizzonte/(1+infl),"
            "((1+infl)/(1+tasso_sconto))*(1-((1+infl)/(1+tasso_sconto))^orizzonte)/((1+infl)*(1-(1+infl)/(1+tasso_sconto))))",
            S.NUMERO_DEC,
            nota="Valore attuale di una rendita unitaria che cresce all'inflazione, sull'orizzonte scelto.",
        )
        riga_f_ind = r
        r = S.campo(
            ws, r, "Fattore del canone indicizzato come dichiarato",
            "=IF(ABS((1+indicizzazione)/(1+tasso_sconto)-1)<0.0000001,orizzonte/(1+indicizzazione),"
            "((1+indicizzazione)/(1+tasso_sconto))*(1-((1+indicizzazione)/(1+tasso_sconto))^orizzonte)/((1+indicizzazione)*(1-(1+indicizzazione)/(1+tasso_sconto))))",
            S.NUMERO_DEC,
        )
        riga_perso = r
        r = S.campo(ws, r, "Canone rinunciato, valore attuale lordo", f"=ricavo_lordo*(B{riga_f_infl}-B{riga_f_ind})", S.EURO)
        riga_perso_netto = r
        r = S.campo(ws, r, "Lo stesso al netto dell'imposta sul canone", f"=B{riga_perso}*(1-ced_libero)", S.EURO, risultato=True,
                    nota="Il canone in piu' sarebbe stato tassato, quindi il confronto va fatto al netto: e' questa la cifra comparabile col risparmio d'imposta.")
        riga_risp = r
        r = S.campo(
            ws, r, "Risparmio d'imposta della cedolare, valore attuale",
            "=(ricavo_effettivo*(1-abbatt_ord)*(irpef_marginale+addizionali)+MAX(ricavo_effettivo*reg_loc,reg_loc_min)/2-ricavo_effettivo*ced_libero)"
            "*IF(tasso_sconto=0,orizzonte,(1-(1+tasso_sconto)^-orizzonte)/tasso_sconto)",
            S.EURO, risultato=True,
            nota="Differenza annua fra imposta in regime ordinario e cedolare, attualizzata sull'orizzonte. Positiva quando la cedolare conviene.",
        )
        riga_saldo = r
        r = S.campo(ws, r, "Saldo della scelta cedolare", f"=B{riga_risp}-B{riga_perso_netto}", S.EURO, risultato=True,
                    nota="Positivo se il risparmio d'imposta supera l'indicizzazione rinunciata, negativo se e' il contrario.")
        ws.conditional_formatting.add(
            f"B{riga_saldo}:B{riga_saldo}",
            CellIsRule(operator="lessThan", formula=["0"], fill=S.FILL_ATTENZIONE),
        )
        for testo in [
            "Tre avvertenze sul saldo, perche' e' un confronto fra due grandezze che non sono simmetriche e leggerlo come un verdetto sarebbe sbagliato. L'orizzonte usato e' quello del modello, tipicamente decenni, mentre l'opzione per la cedolare si esercita per contratto e si puo' riconsiderare a ogni rinnovo: la cifra e' quindi il caso peggiore, cioe' quello di chi la rinnova sempre senza ripensarci. Il risparmio d'imposta dipende dall'aliquota marginale personale, che qui e' un input e non un dato. E il canone concordato ha regole di aggiornamento proprie, fissate dall'accordo territoriale, che questa riga non modella.",
            "Cio' che il saldo dice con certezza e' che la scelta ha due lati e che il secondo non e' trascurabile. Su un orizzonte lungo l'indicizzazione rinunciata puo' valere piu' del risparmio d'imposta che l'ha motivata, ed e' un confronto che quasi nessuno fa perche' il risparmio si vede subito nella dichiarazione e la mancata indicizzazione non si vede mai, essendo un canone che non e' stato chiesto.",
            "Sul rendimento reale in generale: se e' negativo, l'operazione perde potere d'acquisto pur mostrando un utile in euro. Non e' di per se' una ragione per non comprare, perche' l'alternativa va confrontata anch'essa in termini reali e perche' l'abitazione propria produce un servizio abitativo che nessun rendimento cattura; e' una ragione per non chiamare investimento quello che e' consumo o copertura.",
            "L'effetto dell'inflazione sul debito, quello favorevole, esiste solo a tasso fisso. Su un variabile l'inflazione fa salire il tasso di riferimento, la rata cresce, e lo sconto sul debito si paga in interessi: il foglio Simulatore mutuo mostra di quanto, e il percorso del tasso a gradini serve esattamente a questo.",
        ]:
            r = S.nota_riga(ws, r, testo, 3)
        r += 1

        r = S.sezione(ws, r, "Concentrazione del patrimonio", secondaria=True)
        riga_patr = r
        r = S.campo(ws, r, "Patrimonio complessivo, immobili inclusi", 0, S.EURO, input_utente=True, nota="Somma di immobili gia' posseduti, liquidita' e investimenti finanziari, incluso questo acquisto. Zero per saltare il controllo.")
        self.nome("patrimonio_totale", ws, f"B{riga_patr}")
        riga_imm = r
        r = S.campo(ws, r, "Valore immobiliare complessivo dopo questo acquisto", 0, S.EURO, input_utente=True, nota="Prima casa, seconde case, quote ereditate e questo immobile.")
        self.nome("patrimonio_immobiliare", ws, f"B{riga_imm}")
        riga_quota_imm = r
        r = S.campo(ws, r, "Quota immobiliare del patrimonio", "=IF(patrimonio_totale>0,patrimonio_immobiliare/patrimonio_totale,\"n.d.\")", S.PERC, risultato=True)
        r = S.campo(
            ws, r, "Lettura della concentrazione",
            f'=IF(patrimonio_totale=0,"controllo non attivo",IF(B{riga_quota_imm}>0.66,"molto concentrato: oltre due terzi in immobili",'
            f'IF(B{riga_quota_imm}>0.33,"concentrato: oltre un terzo in immobili","entro un terzo del patrimonio")))',
            risultato=True,
        )
        for testo in [
            "Il controllo esiste perche' e' il rischio che il rendimento non vede. Un immobile e' un singolo bene, in una singola via, di un singolo Comune, comprato in un singolo momento del ciclo: porta insieme rischio di timing, di ciclo economico, di tasso e di localizzazione, e non si vende in tre giorni. Chi ha due terzi del patrimonio in mattone non ha un portafoglio diversificato, ha una scommessa sul mercato immobiliare della sua zona.",
            "Va poi sfatata un'aspettativa ricorrente: l'immobiliare non decorrela dall'azionario quando servirebbe. Nelle recessioni i due si muovono insieme, perche' e' la stessa contrazione del credito e della domanda a colpirli. Cio' che ha decorrelato nei momenti di crisi, in modo diverso a seconda dello scenario, e' stato semmai il reddito fisso o il bene rifugio.",
            "L'abitazione principale merita un discorso a parte. Il suo trattamento fiscale, dall'esenzione IMU alla detrazione degli interessi, e' un vantaggio che nessun altro investimento replica, ma resta un capitale che non si puo' diversificare, ne' rendere liquido, ne' bilanciare con il resto. E' la ragione per cui va contata nella concentrazione anche se non e' un investimento.",
        ]:
            r = S.nota_riga(ws, r, testo, 3)
        r += 1

        r = S.sezione(ws, r, "Lettura", secondaria=True)
        riga_verdetto = r
        self.nome("verdetto", ws, f"B{riga_verdetto}")
        r = S.campo(
            ws, r, "Sintesi automatica",
            '=IF(IFERROR(IRR(flussi_tir),-1)<0,"Operazione in perdita sull\'orizzonte scelto",'
            'IF(IFERROR(IRR(flussi_tir),0)<tasso_sconto,"Rende meno del costo opportunita\' del capitale",'
            'IF(IFERROR(IRR(flussi_tir),0)<rend_port,"Rende, ma meno del portafoglio alternativo",'
            '"Batte il portafoglio alternativo sulle assunzioni date")))',
            risultato=True,
        )
        ws.cell(row=riga_verdetto, column=2).alignment = S.SINISTRA
        for testo in [
            "Il verdetto automatico confronta il tasso interno di rendimento con il tasso di sconto e con il rendimento atteso del portafoglio alternativo. Sono tutte assunzioni impostate nel foglio Parametri: cambiandole cambia il verdetto, ed e' esattamente il punto.",
            "Il tasso interno di rendimento non pesa il rischio. Un immobile ha rischio di sfitto, di morosita', di deterioramento, di illiquidita' e di concentrazione su un singolo bene in una singola via di un singolo Comune. Un portafoglio diversificato con lo stesso rendimento atteso non e' la stessa cosa, e la differenza va messa a mano nel giudizio.",
            "Il modello non prezza il lavoro. Gestire un immobile in affitto significa registrare i contratti, seguire le assemblee, rincorrere le manutenzioni, gestire l'inquilino che non paga. Quel tempo ha un costo che nessuna cella cattura.",
        ]:
            r = S.nota_riga(ws, r, testo, 3)

    # --------------------------------------------------------------- confronto
    def foglio_confronto(self) -> None:
        ws = self.wb.create_sheet("Confronto affitto")
        ws.sheet_view.showGridLines = False
        S.larghezze_colonne(ws, {"A": 44, "B": 20, "C": 66})
        r = S.titolo(
            ws,
            1,
            "Comprare oppure restare in affitto investendo la differenza",
            "Confronto a parita' di esborso, nell'impostazione dei fogli di Paolo Coletti: chi non compra investe l'anticipo e ogni anno la differenza fra la rata piu' i costi da proprietario e il canone che paga.",
        )

        r = S.sezione(ws, r, "Assunzioni del confronto")
        riga_can_alt = r
        r = S.campo(ws, r, "Canone mensile che si pagherebbe restando in affitto", 550, S.EURO, input_utente=True, nota="Per un immobile equivalente a quello che si comprerebbe, non per quello in cui si vive oggi.")
        self.nome("canone_alternativo", ws, f"B{riga_can_alt}")
        riga_costi_prop = r
        r = S.campo(ws, r, "Costi annui del proprietario", "=condominio*quota_condominio+prezzo*manut_pct+assicurazione+accantonamento_ristrutturazione+IF(abitazione_principale=\"SI\",0,rendita*riv_rendita*imu_molt*imu_aliquota)", S.EURO, nota="Condominio a carico, manutenzione, assicurazione, accantonamento per la ristrutturazione, IMU se non e' abitazione principale.")
        self.nome("costi_proprietario", ws, f"B{riga_costi_prop}")
        r += 1

        r = S.intestazioni(ws, r, ["Anno", "Uscite comprando", "Uscite affittando", "Differenza investita", "Portafoglio", "Versato", "Valore immobile", "Debito residuo"], [8, 20, 20, 20, 20, 18, 20, 18])
        prima = r
        ws.cell(row=r, column=1, value=0).number_format = S.NUMERO
        ws.cell(row=r, column=2, value="=esborso").number_format = S.EURO
        ws.cell(row=r, column=3, value=0).number_format = S.EURO
        ws.cell(row=r, column=4, value="=esborso").number_format = S.EURO
        ws.cell(row=r, column=5, value="=esborso").number_format = S.EURO
        ws.cell(row=r, column=6, value="=esborso").number_format = S.EURO
        ws.cell(row=r, column=7, value="=prezzo").number_format = S.EURO
        ws.cell(row=r, column=8, value="=mutuo_importo").number_format = S.EURO
        for col in range(1, 9):
            ws.cell(row=r, column=col).border = S.BORDO
            ws.cell(row=r, column=col).fill = S.FILL_CALCOLO
        r += 1

        for anno in range(1, ORIZZONTE_MAX + 1):
            attivo = f"$A{r}<=orizzonte"
            ws.cell(row=r, column=1, value=anno).number_format = S.NUMERO
            ws.cell(row=r, column=2, value=f'=IF({attivo},IF($A{r}<=durata,rata_annua,0)+costi_proprietario*(1+infl)^($A{r}-1)-IF(abitazione_principale="SI",INDEX(detrazione_anno,MIN($A{r},{MAX_ANNI})),0),0)').number_format = S.EURO
            ws.cell(row=r, column=3, value=f'=IF({attivo},canone_alternativo*12*(1+infl)^($A{r}-1),0)').number_format = S.EURO
            ws.cell(row=r, column=4, value=f"=$B{r}-$C{r}").number_format = S.EURO
            ws.cell(row=r, column=5, value=f'=IF({attivo},$E{r-1}*(1+rend_port)+$D{r},0)').number_format = S.EURO
            ws.cell(row=r, column=6, value=f'=IF({attivo},$F{r-1}+$D{r},0)').number_format = S.EURO
            ws.cell(row=r, column=7, value=f'=IF({attivo},prezzo*(1+riv_immobile)^$A{r},0)').number_format = S.EURO
            ws.cell(row=r, column=8, value=f'=IF({attivo},INDEX(debito_residuo_anno,MIN($A{r},{MAX_ANNI})),0)').number_format = S.EURO
            for col in range(1, 9):
                ws.cell(row=r, column=col).border = S.BORDO
                ws.cell(row=r, column=col).fill = S.FILL_CALCOLO
            r += 1
        ultima = r - 1
        self.nome_intervallo("conf_portafoglio", ws, f"$E${prima}:$E${ultima}")
        self.nome_intervallo("conf_versato", ws, f"$F${prima}:$F${ultima}")
        self.nome_intervallo("conf_valore", ws, f"$G${prima}:$G${ultima}")
        self.nome_intervallo("conf_debito", ws, f"$H${prima}:$H${ultima}")
        r += 1

        r = S.sezione(ws, r, "Esito a fine orizzonte", secondaria=True)
        riga_pc = r
        r = S.campo(ws, r, "Patrimonio comprando", "=INDEX(conf_valore,orizzonte+1)*(1-costi_vendita)-INDEX(conf_debito,orizzonte+1)", S.EURO, risultato=True, nota="Valore dell'immobile al netto dei costi di vendita e del debito residuo.")
        r = S.campo(ws, r, "Portafoglio lordo affittando", "=INDEX(conf_portafoglio,orizzonte+1)", S.EURO)
        r = S.campo(ws, r, "Capitale versato nel portafoglio", "=INDEX(conf_versato,orizzonte+1)", S.EURO)
        riga_pa = r
        r = S.campo(ws, r, "Patrimonio affittando, al netto dell'imposta", "=INDEX(conf_portafoglio,orizzonte+1)-MAX(0,INDEX(conf_portafoglio,orizzonte+1)-INDEX(conf_versato,orizzonte+1))*tax_port", S.EURO, risultato=True)
        riga_diff = r
        r = S.campo(ws, r, "Differenza a favore dell'acquisto", f"=B{riga_pc}-B{riga_pa}", S.EURO, risultato=True)
        # Il nome definito esiste per il Cruscotto, che leggeva questa cella per
        # coordinata. Una coordinata fissa dentro una formula che punta a un altro
        # foglio e' il difetto piu' insidioso del generatore: inserire una riga qui
        # sopra non produce alcun errore, produce un verdetto sbagliato sul primo
        # foglio del workbook, cioe' quello che si legge per decidere.
        self.nome("conf_differenza", ws, f"B{riga_diff}")
        riga_verd = r
        r = S.campo(ws, r, "Esito", f'=IF(B{riga_diff}>0,"Conviene comprare","Conviene restare in affitto e investire la differenza")', risultato=True)
        ws.cell(row=riga_verd, column=2).alignment = S.SINISTRA
        riga_avv = r
        r = S.campo(
            ws, r, "Avvertenza",
            '=IF(mutuo_importo>0,"","Acquisto senza mutuo: il confronto sopra non e\' significativo, vedi la nota in fondo")',
            risultato=True,
            nota="Compare da sola quando l'importo del mutuo e' zero, e resta bianca altrimenti.",
        )
        ws.cell(row=riga_avv, column=2).alignment = S.SINISTRA
        r += 1
        for testo in [
            "Il confronto e' sensibile a tre soli numeri: il rendimento atteso del portafoglio, la rivalutazione dell'immobile e il canone alternativo. Cambiando il primo di un punto l'esito spesso si ribalta, il che dice quanto poco vada preso come verdetto e quanto vada preso come mappa di sensibilita'.",
            "Il modello assume disciplina perfetta di chi affitta: investe davvero ogni euro di differenza, ogni anno, senza toccarlo. Nella realta' quasi nessuno lo fa, e il mutuo funziona come piano di accumulo forzato. E' un vantaggio comportamentale reale che il foglio non sa misurare.",
            "Restano fuori dal conto la sicurezza abitativa, la liberta' di ristrutturare, il rischio di sfratto e il vincolo di mobilita' lavorativa. Sono decisivi nella scelta di dove vivere e irrilevanti nella scelta di dove investire: e' la ragione per cui le due domande vanno tenute separate.",
            "Se l'importo del mutuo e' zero il foglio continua a calcolare ma il confronto perde significato, e conviene sapere perche'. Il confronto e' costruito a parita' di esborso: chi non compra investe l'anticipo e poi, ogni anno, la differenza fra quanto esce a chi ha comprato e il canone che paga. Senza mutuo l'anticipo diventa il prezzo intero, quindi il portafoglio alternativo parte con tutto il capitale investito, mentre le uscite di chi ha comprato si riducono ai soli costi da proprietario. Il conto che ne esce non e' sbagliato, e' un'altra domanda: non piu' comprare a debito contro affittare e investire, ma immobilizzare il capitale in un immobile contro investirlo sui mercati. Per quella domanda il foglio da leggere e' Metriche, con il tasso interno di rendimento contro il rendimento del portafoglio, e non questo.",
        ]:
            r = S.nota_riga(ws, r, testo, 3)

    # ----------------------------------------------------------------- scenari
    def foglio_scenari(self) -> None:
        ws = self.wb.create_sheet("Scenari")
        ws.sheet_view.showGridLines = False
        S.larghezze_colonne(ws, {"A": 30, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16, "H": 16})
        r = S.titolo(
            ws,
            1,
            "Sensibilita'",
            "Il valore centrale di ogni tabella e' lo scenario base. Le variazioni sono quelle che si osservano davvero: mezzo punto di tasso, un decimo di canone, un decimo di prezzo.",
        )

        r = S.sezione(ws, r, "Cash flow annuo al variare di tasso e canone")
        r = S.nota_riga(ws, r, "Righe: tasso del mutuo. Colonne: canone mensile. Il cash flow e' il reddito operativo netto meno l'imposta e meno la rata, ricalcolati per ciascuna combinazione. Le celle rosse sono quelle in cui l'immobile costa piu' di quanto rende.")
        variazioni_tasso = [-0.010, -0.005, 0.0, 0.005, 0.010, 0.015]
        variazioni_canone = [-0.20, -0.10, 0.0, 0.10, 0.20]

        r = S.intestazioni(ws, r, ["Tasso \\ canone"] + [f"{v:+.0%}" for v in variazioni_canone], [24, 16, 16, 16, 16, 16])
        prima = r
        for dt in variazioni_tasso:
            c = ws.cell(row=r, column=1, value=f"=TEXT(tasso{dt:+f},\"0.00%\")")
            c.font = S.ETICHETTA_BOLD
            c.fill = S.FILL_CALCOLO
            c.border = S.BORDO
            for j, dc in enumerate(variazioni_canone, start=2):
                canone = f"canone_mese*12*(1+{dc})"
                effettivo = f"({canone}-canone_mese*mesi_sfitto*(1+{dc}))*(1-morosita_pct)"
                costi = "(condominio*quota_condominio+prezzo*manut_pct+assicurazione+rendita*riv_rendita*imu_molt*imu_aliquota+accantonamento_ristrutturazione)"
                imposta = f"{effettivo}*ced_libero"
                rata = f"IF(mutuo_importo>0,PMT((tasso{dt:+f})/12,durata*12,-mutuo_importo)*12,0)"
                cella = ws.cell(row=r, column=j, value=f"={effettivo}-{costi}-{imposta}-{rata}")
                cella.number_format = S.EURO
                cella.border = S.BORDO
                cella.fill = S.FILL_CALCOLO
            r += 1
        ultima = r - 1
        ws.conditional_formatting.add(
            f"B{prima}:F{ultima}",
            CellIsRule(operator="lessThan", formula=["0"], fill=S.FILL_ATTENZIONE),
        )
        ws.conditional_formatting.add(
            f"B{prima}:F{ultima}",
            ColorScaleRule(start_type="min", start_color="F8CBAD", mid_type="num", mid_value=0, mid_color="FFF2CC", end_type="max", end_color="C6E0B4"),
        )
        r += 2

        r = S.sezione(ws, r, "Rendimento netto al variare del prezzo", secondaria=True)
        r = S.nota_riga(ws, r, "Il prezzo entra due volte: al numeratore attraverso la manutenzione, al denominatore attraverso il costo totale e attraverso le imposte, che con il prezzo-valore restano pero' ancorate alla rendita catastale e non scendono con il prezzo.")
        variazioni_prezzo = [-0.20, -0.15, -0.10, -0.05, 0.0, 0.05, 0.10]
        r = S.intestazioni(ws, r, ["Variazione prezzo", "Prezzo", "Imposte", "Costo totale", "Utile netto", "Rendimento netto", "Cash flow"], [20, 16, 16, 18, 16, 18, 16])
        prima2 = r
        for dp in variazioni_prezzo:
            prezzo_v = f"prezzo*(1+{dp})"
            imposte_v = (
                f'IF(da_impresa="SI",{prezzo_v}*IF(agevolata="SI",iva_prima,IF(di_lusso="SI",iva_lusso,iva_ord))+3*fisso_impresa,'
                f'MAX(IF(AND(usa_prezzo_valore="SI",rendita>0),valore_catastale,{prezzo_v})*IF(agevolata="SI",reg_prima,reg_ord),reg_min)+ipo_priv+cat_priv)'
            )
            costo_v = f"{prezzo_v}+{imposte_v}+{prezzo_v}*provv_pct*(1+iva_provv)+notaio_cv+altri_costi+oneri_mutuo"
            utile_v = f"utile_locazione-({prezzo_v}-prezzo)*manut_pct"
            ws.cell(row=r, column=1, value=f"{dp:+.0%}").font = S.ETICHETTA_BOLD
            ws.cell(row=r, column=2, value=f"={prezzo_v}").number_format = S.EURO
            ws.cell(row=r, column=3, value=f"={imposte_v}").number_format = S.EURO
            ws.cell(row=r, column=4, value=f"={costo_v}").number_format = S.EURO
            ws.cell(row=r, column=5, value=f"={utile_v}").number_format = S.EURO
            ws.cell(row=r, column=6, value=f"=({utile_v})/({costo_v})").number_format = S.PERC
            ws.cell(row=r, column=7, value=f"={utile_v}-rata_annua").number_format = S.EURO
            for col in range(1, 8):
                ws.cell(row=r, column=col).border = S.BORDO
                ws.cell(row=r, column=col).fill = S.FILL_CALCOLO
            r += 1
        ultima2 = r - 1
        ws.conditional_formatting.add(
            f"F{prima2}:F{ultima2}",
            ColorScaleRule(start_type="min", start_color="F8CBAD", end_type="max", end_color="C6E0B4"),
        )
        r += 2

        r = S.sezione(ws, r, "Tre scenari a confronto", secondaria=True)
        r = S.nota_riga(ws, r, "Le celle gialle di questa sezione sono indipendenti dal resto del workbook: si impostano qui le tre ipotesi e si legge come cambia l'esito. Serve a rispondere alla domanda che conta davvero prima di comprare, cioe' non quanto rende se tutto va bene, ma quanto si perde se va male.")
        intest_sc = ["Voce", "Pessimistico", "Base", "Ottimistico"]
        r = S.intestazioni(ws, r, intest_sc, [46, 18, 18, 18])

        def riga_scenario(etichetta, valori, formato=S.EURO, input_utente=False, risultato=False, nota=""):
            """Scrive una riga della tabella a tre scenari e restituisce la sua riga."""
            nonlocal r
            scritta = r
            e = ws.cell(row=r, column=1, value=etichetta)
            e.font = S.ETICHETTA_BOLD if risultato else S.ETICHETTA
            e.alignment = S.SINISTRA
            for i, v in enumerate(valori, start=2):
                c = ws.cell(row=r, column=i, value=v)
                c.number_format = formato
                c.border = S.BORDO
                if input_utente:
                    c.fill = S.FILL_INPUT
                elif risultato:
                    c.fill = S.FILL_RISULTATO
                    c.font = S.ETICHETTA_BOLD
                else:
                    c.fill = S.FILL_CALCOLO
            if nota:
                n_ = ws.cell(row=r, column=5, value=nota)
                n_.font = S.NOTA
                n_.alignment = S.SINISTRA
            r += 1
            return scritta

        # Ogni riga si registra in `posizioni` sotto una chiave breve, e le formule
        # citano le altre righe solo attraverso quelle chiavi. La differenza con gli
        # offset numerici usati prima non e' di stile: una chiave assente solleva un
        # KeyError alla generazione, mentre un offset sbagliato produce un
        # riferimento valido a una riga diversa, cioe' un numero plausibile e falso.
        # Il vincolo che ne discende e' esplicito: una formula puo' citare solo righe
        # gia' scritte, e il tentativo di citarne una successiva fallisce subito.
        posizioni = {}
        posizioni["ca"] = riga_scenario("Canone mensile", ["=canone_mese*0.85", "=canone_mese", "=canone_mese*1.1"], S.EURO, input_utente=True,
                                        nota="Il pessimistico sconta un mercato in cui si affitta solo abbassando.")
        posizioni["sf"] = riga_scenario("Mesi di sfitto all'anno", [3, "=mesi_sfitto", 0.5], S.NUMERO_DEC, input_utente=True,
                                        nota="Tre mesi e' un anno con un cambio di inquilino andato male.")
        posizioni["mo"] = riga_scenario("Morosita'", [0.08, "=morosita_pct", 0], S.PERC, input_utente=True)
        posizioni["ta"] = riga_scenario("Tasso del mutuo", ["=tasso+0.015", "=tasso", "=tasso-0.005"], S.PERC, input_utente=True,
                                        nota="Sul fisso i tre valori coincidono; sul variabile il pessimistico e' lo scenario da sostenere.")
        posizioni["riv"] = riga_scenario("Rivalutazione annua dell'immobile", [-0.01, "=riv_immobile", 0.03], S.PERC, input_utente=True,
                                         nota="In termini reali il mattone italiano e' rimasto fermo per vent'anni: il pessimistico non e' catastrofismo.")

        colonne = ("B", "C", "D")
        for etichetta, chiave, formula, formato, risultato in [
            ("Ricavo effettivo", "ric", "({c}{ca}*12-{c}{ca}*{c}{sf})*(1-{c}{mo})", S.EURO, False),
            ("Costi operativi", "cos", "condominio*quota_condominio+prezzo*manut_pct+assicurazione+costo_tempo+rendita*riv_rendita*imu_molt*imu_aliquota+accantonamento_ristrutturazione", S.EURO, False),
            ("Reddito operativo netto", "noi", "{c}{ric}-{c}{cos}", S.EURO, False),
            ("Imposta sul canone", "imp", "{c}{ric}*ced_libero", S.EURO, False),
            ("Utile netto", "uti", "{c}{noi}-{c}{imp}", S.EURO, True),
            ("Rata annua", "rat", "IF(mutuo_importo>0,PMT({c}{ta}/12,durata*12,-mutuo_importo)*12,0)", S.EURO, False),
            ("Cash flow annuo", "cf", "{c}{uti}-{c}{rat}", S.EURO, True),
            ("Rendimento netto", "rnet", "{c}{uti}/costo_totale", S.PERC, True),
            ("Debt service coverage ratio", "dscr", "IF({c}{rat}>0,{c}{noi}/{c}{rat},\"n.d.\")", S.NUMERO_DEC, False),
            ("Valore dell'immobile a fine orizzonte", "val", "prezzo*(1+{c}{riv})^orizzonte", S.EURO, False),
            ("Debito residuo a fine orizzonte", "deb", "IF(mutuo_importo>0,mutuo_importo*((1+{c}{ta}/12)^(durata*12)-(1+{c}{ta}/12)^(MIN(orizzonte,durata)*12))/((1+{c}{ta}/12)^(durata*12)-1),0)", S.EURO, False),
            ("Patrimonio netto a fine orizzonte", "pat", "{c}{val}*(1-costi_vendita)-{c}{deb}", S.EURO, True),
        ]:
            valori = ["=" + formula.format(c=c, **posizioni) for c in colonne]
            posizioni[chiave] = riga_scenario(etichetta, valori, formato, risultato=risultato)

        ws.conditional_formatting.add(
            f"B{posizioni['cf']}:D{posizioni['cf']}",
            CellIsRule(operator="lessThan", formula=["0"], fill=S.FILL_ATTENZIONE),
        )
        r += 1
        r = S.nota_riga(ws, r, "Il patrimonio netto a fine orizzonte usa la formula chiusa del debito residuo nell'ammortamento alla francese, quindi resta esatto anche cambiando il tasso di scenario. Il cash flow annuo va letto come impegno mensile: diviso dodici e' quanto si mette di tasca propria ogni mese in quello scenario, ed e' il numero che dice se lo scenario e' sostenibile o solo sfavorevole.")
        r += 1

        r = S.sezione(ws, r, "Prezzo massimo sostenibile", secondaria=True)
        riga_obiettivo = r
        r = S.campo(ws, r, "Rendimento netto obiettivo", 0.04, S.PERC, input_utente=True, nota="Il rendimento sotto il quale l'operazione non ha senso rispetto alle alternative. Lo usa anche il foglio Confronto immobili per dare l'esito di ciascun annuncio.")
        self.nome("rend_obiettivo", ws, f"B{riga_obiettivo}")
        r = S.campo(ws, r, "Costo totale sostenibile a quel rendimento", f"=utile_locazione/B{riga_obiettivo}", S.EURO, nota="Serve come ordine di grandezza: il costo totale non e' proporzionale al prezzo, perche' una parte delle sue voci non dipende dal prezzo.")
        riga_costo_sost = r - 1

        # Il prezzo massimo si ricava in forma chiusa, non per proporzione. La
        # versione precedente divideva il costo sostenibile per uno piu' l'incidenza
        # percentuale dei costi accessori dello scenario base, cioe' assumeva che
        # quell'incidenza restasse costante al variare del prezzo. Non resta
        # costante, e sbaglia sempre nella stessa direzione: notaio, oneri del mutuo,
        # imposte ipotecaria e catastale e, con il prezzo-valore, l'intera imposta di
        # registro sono importi fissi, quindi la loro incidenza percentuale cresce se
        # il prezzo scende e cala se sale. Su un immobile piccolo con prezzo-valore la
        # distorsione e' di alcune migliaia di euro, cioe' proprio nell'ordine di
        # grandezza della trattativa che questo numero dovrebbe guidare.
        #
        # L'algebra e' elementare e vale la pena scriverla, perche' e' cio' che le
        # tre celle seguenti calcolano. Il costo totale in funzione del prezzo P e'
        # lineare a tratti, cioe' P*(1+k)+c dove k raccoglie tutto cio' che scala col
        # prezzo e c tutto cio' che non scala. L'utile netto annuo, a sua volta,
        # scende quando il prezzo sale, perche' manutenzione e accantonamento per la
        # ristrutturazione sono quote del valore: utile(P) = utile_base - (P-prezzo)*m.
        # Imporre utile(P)/costo(P) = obiettivo da' un'equazione di primo grado in P,
        # la cui soluzione e' (utile_base + prezzo*m - obiettivo*c)/(obiettivo*(1+k)+m).
        riga_k = r
        r = S.campo(
            ws, r, "Quota del prezzo che diventa costo aggiuntivo",
            '=IF(da_impresa="SI",IF(agevolata="SI",iva_prima,IF(di_lusso="SI",iva_lusso,iva_ord)),'
            'IF(AND(usa_prezzo_valore="SI",rendita>0),0,IF(agevolata="SI",reg_prima,reg_ord)))'
            "+provv_pct*(1+iva_provv)",
            S.PERC_1,
            nota="Per ogni euro di prezzo in piu', quanti centesimi di costi accessori si aggiungono. Con il prezzo-valore l'imposta di registro non entra, perche' resta ancorata alla rendita catastale e non cresce col prezzo.",
        )
        riga_c = r
        r = S.campo(
            ws, r, "Costi che non dipendono dal prezzo",
            '=IF(da_impresa="SI",3*fisso_impresa,'
            'IF(AND(usa_prezzo_valore="SI",rendita>0),MAX(valore_catastale*IF(agevolata="SI",reg_prima,reg_ord),reg_min),0)'
            "+ipo_priv+cat_priv)+notaio_cv+altri_costi+oneri_mutuo",
            S.EURO,
            nota="Notaio, altri costi, oneri del mutuo, imposte fisse, e con il prezzo-valore l'intera imposta di registro. Sono la ragione per cui l'incidenza percentuale dei costi non e' costante.",
        )
        riga_m = r
        r = S.campo(
            ws, r, "Costi annui che scalano col prezzo",
            "=manut_pct+ristrutt_pct/ristrutt_anni",
            S.PERC_1,
            nota="Manutenzione ordinaria e accantonamento per la ristrutturazione di fine ciclo sono quote del valore, quindi un prezzo piu' alto abbassa l'utile netto oltre ad alzare il costo totale.",
        )
        riga_pmax = r
        r = S.campo(
            ws, r, "Prezzo massimo corrispondente",
            f"=IFERROR((utile_locazione+prezzo*B{riga_m}-B{riga_obiettivo}*B{riga_c})/(B{riga_obiettivo}*(1+B{riga_k})+B{riga_m}),\"non calcolabile\")",
            S.EURO, risultato=True,
            nota="Soluzione esatta, non piu' una proporzione sull'incidenza dei costi. Se il rendimento obiettivo e' tanto alto da non essere raggiungibile a nessun prezzo positivo, il valore risulta negativo: e' la risposta corretta, e va letta come non esiste un prezzo che giustifichi l'operazione a quella soglia.",
        )
        r = S.campo(ws, r, "Scarto rispetto al prezzo trattato", f"=B{riga_pmax}-prezzo", S.EURO, nota="Se negativo, il prezzo trattato e' sopra quello che l'immobile puo' giustificare a quel rendimento, e la differenza e' lo sconto da ottenere.")

        # Controllo di chiusura. Ricalcola il rendimento netto al prezzo appena
        # trovato usando le formule esatte delle imposte, floor di legge compreso, e
        # lo confronta con l'obiettivo. Deve dare zero: se non lo da', un'assunzione
        # della linearizzazione non tiene, tipicamente il minimo di legge
        # dell'imposta di registro che diventa vincolante sui prezzi molto bassi. E'
        # scritto qui, e non lasciato a un test, perche' chi cambia gli input a video
        # deve poterlo vedere nel momento in cui succede.
        imposte_pmax = (
            f'IF(da_impresa="SI",B{riga_pmax}*IF(agevolata="SI",iva_prima,IF(di_lusso="SI",iva_lusso,iva_ord))+3*fisso_impresa,'
            f'MAX(IF(AND(usa_prezzo_valore="SI",rendita>0),valore_catastale,B{riga_pmax})*IF(agevolata="SI",reg_prima,reg_ord),reg_min)+ipo_priv+cat_priv)'
        )
        costo_pmax = f"B{riga_pmax}+{imposte_pmax}+B{riga_pmax}*provv_pct*(1+iva_provv)+notaio_cv+altri_costi+oneri_mutuo"
        utile_pmax = f"utile_locazione-(B{riga_pmax}-prezzo)*B{riga_m}"
        riga_ver = r
        r = S.campo(ws, r, "Verifica: rendimento netto a quel prezzo", f"=IFERROR(({utile_pmax})/({costo_pmax}),\"non calcolabile\")", S.PERC_1,
                    nota="Ricalcolato con le formule esatte delle imposte, minimo di legge compreso.")
        r = S.campo(ws, r, "Scarto dalla soglia, deve essere zero", f"=IFERROR(B{riga_ver}-B{riga_obiettivo},\"non calcolabile\")", "0.0000%",
                    nota="Diverso da zero solo se un'assunzione della soluzione chiusa non tiene: il caso noto e' il minimo di legge dell'imposta di registro, che su prezzi molto bassi diventa vincolante e rende il costo totale non piu' lineare nel prezzo.")
        r = S.campo(ws, r, "Canone minimo per un cash flow non negativo", "=(rata_annua+condominio*quota_condominio+prezzo*manut_pct+assicurazione+rendita*riv_rendita*imu_molt*imu_aliquota+accantonamento_ristrutturazione)/((12-mesi_sfitto)*(1-morosita_pct)*(1-ced_libero))", S.EURO_DEC, risultato=True, nota="Canone mensile sotto il quale l'immobile assorbe cassa invece di generarla.")

    # ------------------------------------------------------------- estrazioni
    def foglio_estrazioni(self) -> None:
        """Foglio tecnico nascosto: le estrazioni casuali della simulazione.

        La simulazione probabilistica ha un problema pratico in Excel. Usare la
        funzione casuale nativa la renderebbe volatile, cioe' ogni ricalcolo
        cambierebbe tutti i numeri e due letture consecutive dello stesso file
        darebbero risultati diversi: inutilizzabile per decidere e impossibile da
        verificare.

        La soluzione adottata separa le due cose. Le estrazioni sono numeri fissi,
        generati una volta sola alla creazione del file con un seme dichiarato,
        quindi identiche a ogni riapertura e riproducibili. Il calcolo che sta sopra,
        invece, e' formula viva: cambiando un input tutti i mille scenari si
        ricalcolano davanti agli occhi, ma sulla stessa estrazione. Si ottiene cosi'
        una simulazione stabile e insieme interattiva.

        Il foglio e' nascosto perche' non c'e' nulla da leggerci: si scopre solo se
        si vuole ispezionare il motore.
        """
        import random

        ws = self.wb.create_sheet("_Estrazioni")
        ws.sheet_state = "hidden"
        generatore = random.Random(SEME_SIMULAZIONE)

        intest = ["n", "z canone", "z sfitto", "z tasso", "z rivalutazione", "u evento",
                  "Canone annuo", "Mesi sfitto", "Tasso", "Rivalutazione", "Ricavo effettivo",
                  "NOI", "Utile netto", "Rata annua", "Cash flow", "Valore finale",
                  "Debito residuo", "Patrimonio finale", "Montante", "Rendimento netto"]
        for i, t in enumerate(intest, start=1):
            ws.cell(row=1, column=i, value=t).font = S.ETICHETTA_BOLD
        prima = 2

        for k in range(ESTRAZIONI):
            r = prima + k
            ws.cell(row=r, column=1, value=k + 1)
            # Quattro normali standard e una uniforme, fisse.
            for col in range(2, 6):
                ws.cell(row=r, column=col, value=round(generatore.gauss(0, 1), 6))
            ws.cell(row=r, column=6, value=round(generatore.random(), 6))

            ws.cell(row=r, column=7, value=f"=MAX(0,canone_mese*(1+$B{r}*vol_canone))*12")
            ws.cell(row=r, column=8, value=f"=MEDIAN(0,mesi_sfitto+$C{r}*vol_sfitto,12)")
            ws.cell(row=r, column=9, value=f"=MAX(0,tasso+$D{r}*vol_tasso)")
            # La rivalutazione si compone su tutto l'orizzonte, quindi l'estrazione
            # non e' la variazione di un anno ma la media dell'intero periodo, e la
            # sua dispersione scende con la radice del numero di anni. Senza questa
            # correzione un'estrazione verrebbe trattata come un regime permanente e
            # la coda alta produrrebbe patrimoni finali fuori scala.
            ws.cell(row=r, column=10, value=f"=riv_immobile+$E{r}*vol_rivalutazione/SQRT(MAX(orizzonte,1))")
            # L'evento di morosita' grave toglie i mesi di canone impostati.
            ws.cell(
                row=r, column=11,
                value=(
                    f"=MAX(0,($G{r}-$G{r}/12*$H{r})*(1-morosita_pct)"
                    f"-IF($F{r}<prob_morosita_grave,$G{r}/12*mesi_persi_morosita,0))"
                ),
            )
            ws.cell(row=r, column=12, value=f"=$K{r}-(ricavo_effettivo-noi_annuo)")
            ws.cell(row=r, column=13, value=f"=$L{r}-$K{r}*ced_libero")
            ws.cell(row=r, column=14, value=f"=IF(mutuo_importo>0,PMT($I{r}/12,durata*12,-mutuo_importo)*12,0)")
            ws.cell(row=r, column=15, value=f"=$M{r}-$N{r}")
            ws.cell(row=r, column=16, value=f"=prezzo*(1+$J{r})^orizzonte")
            ws.cell(
                row=r, column=17,
                value=(
                    f"=IF(mutuo_importo>0,mutuo_importo*((1+$I{r}/12)^(durata*12)"
                    f"-(1+$I{r}/12)^(MIN(orizzonte,durata)*12))/((1+$I{r}/12)^(durata*12)-1),0)"
                ),
            )
            ws.cell(row=r, column=18, value=f"=$P{r}*(1-costi_vendita)-$Q{r}")
            # Il montante confronta due strade che partono dallo stesso esborso. Chi
            # compra, se il flusso di cassa e' negativo, deve versare quella somma ogni
            # anno prendendola da altrove, e quel denaro ha un costo opportunita': i
            # flussi vanno quindi capitalizzati al rendimento del portafoglio
            # alternativo, non sommati a valore nominale. Con flusso positivo vale il
            # simmetrico, cioe' la cassa incassata si reinveste.
            ws.cell(
                row=r, column=19,
                value=f"=$R{r}+IF(rend_port=0,$O{r}*orizzonte,$O{r}*((1+rend_port)^orizzonte-1)/rend_port)",
            )
            ws.cell(row=r, column=20, value=f"=IF(costo_totale>0,$M{r}/costo_totale,0)")

        ultima = prima + ESTRAZIONI - 1
        for nome, colonna in (("sim_cash_flow", "O"), ("sim_patrimonio", "R"),
                              ("sim_montante", "S"), ("sim_rendimento", "T"),
                              ("sim_utile", "M")):
            self.nome_intervallo(nome, ws, f"${colonna}${prima}:${colonna}${ultima}")

    # ---------------------------------------------------------------- rischio
    def foglio_rischio(self) -> None:
        """Distribuzione degli esiti e classifica delle variabili che pesano.

        Il foglio Scenari risponde alla domanda su cosa succede in tre casi scelti a
        mano. Questo risponde a due domande diverse e piu' difficili: quanto e'
        probabile ciascun esito, e quale delle ipotesi conta davvero.

        La prima e' una simulazione su mille estrazioni, che restituisce percentili e
        probabilita' invece di un numero solo. La seconda e' un'analisi a tornado, che
        muove una variabile per volta di una percentuale uguale per tutte e ordina le
        variabili per quanto spostano il risultato: serve a sapere su cosa vale la
        pena raccogliere informazione migliore, e su cosa invece non cambia nulla.
        """
        ws = self.wb.create_sheet("Rischio")
        ws.sheet_view.showGridLines = False
        S.larghezze_colonne(ws, {"A": 46, "B": 18, "C": 18, "D": 18, "E": 44})
        r = S.titolo(
            ws,
            1,
            "Rischio: distribuzione degli esiti e peso delle ipotesi",
            f"Simulazione su {ESTRAZIONI} estrazioni con seme fisso, quindi riproducibile: gli stessi input danno sempre gli stessi risultati, ma cambiando un input tutto si ricalcola.",
            5,
        )

        r = S.sezione(ws, r, "Quanto sono incerte le ipotesi", 5)
        r = S.nota_riga(ws, r, "Non si dichiara quanto valgono le variabili, che sta negli altri fogli, ma quanto ci si sbaglia nel prevederle. Sono le uniche cose da impostare qui.", 5)
        for chiave, etichetta, valore, formato, nota in [
            ("vol_canone", "Incertezza sul canone", 0.10, S.PERC,
             "Scarto tipico rispetto al canone atteso. Il dieci per cento significa che due volte su tre il canone vero sta fra il novanta e il centodieci per cento di quello ipotizzato."),
            ("vol_sfitto", "Incertezza sui mesi di sfitto", 1.5, S.NUMERO_DEC,
             "In mesi. Uno e mezzo copre l'anno con un cambio di inquilino andato lungo."),
            ("vol_tasso", "Incertezza sul tasso", 0.0, S.PERC,
             "Zero per un tasso fisso, che e' certo per definizione. Su un variabile un punto percentuale e' un'ipotesi ordinaria."),
            ("vol_rivalutazione", "Incertezza sulla rivalutazione annua", 0.04, S.PERC,
             "Riferita al singolo anno. Poiche' la rivalutazione si compone, la simulazione la scala per la radice dell'orizzonte: su venticinque anni l'incertezza sulla media annua e' un quinto di quella su un anno solo."),
            ("prob_morosita_grave", "Probabilita' annua di morosita' grave", 0.05, S.PERC,
             "Non il piccolo ritardo, che sta gia' nell'accantonamento ordinario, ma l'inquilino che smette di pagare e va sfrattato."),
            ("mesi_persi_morosita", "Mesi di canone persi in caso di morosita' grave", 12, S.NUMERO_DEC,
             "Fra convalida di sfratto ed esecuzione passa in genere piu' di un anno, e nel frattempo le imposte e le spese si pagano lo stesso."),
        ]:
            riga = r
            r = S.campo(ws, r, etichetta, valore, formato, input_utente=True, nota=nota)
            self.nome(chiave, ws, f"B{riga}")
        r += 1

        r = S.sezione(ws, r, "Come vanno a finire i mille scenari", 5, secondaria=True)
        r = S.intestazioni(ws, r, ["Grandezza", "Peggiore 5%", "Mediana", "Migliore 5%"], [46, 18, 18, 18])
        for etichetta, intervallo, formato, nota in [
            ("Cash flow annuo", "sim_cash_flow", S.EURO, "Diviso dodici e' l'impegno mensile. La colonna di sinistra e' lo scenario che va messo in conto, non quello da escludere."),
            ("Utile netto annuo", "sim_utile", S.EURO, ""),
            ("Rendimento netto", "sim_rendimento", S.PERC, ""),
            ("Patrimonio a fine orizzonte", "sim_patrimonio", S.EURO, "Valore dell'immobile al netto dei costi di vendita e del debito residuo."),
            ("Montante complessivo", "sim_montante", S.EURO, "Patrimonio piu' i flussi di cassa capitalizzati al rendimento del portafoglio alternativo. E' la grandezza confrontabile con il non comprare, perche' tratta allo stesso modo il denaro impiegato nelle due strade."),
        ]:
            e = ws.cell(row=r, column=1, value=etichetta)
            e.font = S.ETICHETTA
            e.alignment = S.SINISTRA
            for col, formula in ((2, f"=PERCENTILE({intervallo},0.05)"),
                                 (3, f"=MEDIAN({intervallo})"),
                                 (4, f"=PERCENTILE({intervallo},0.95)")):
                c = ws.cell(row=r, column=col, value=formula)
                c.number_format = formato
                c.border = S.BORDO
                c.fill = S.FILL_CALCOLO
            if nota:
                n_ = ws.cell(row=r, column=5, value=nota)
                n_.font = S.NOTA
                n_.alignment = S.SINISTRA
            r += 1
        r += 1

        r = S.sezione(ws, r, "Le probabilita' che contano", 5, secondaria=True)
        nomi_prob = ["prob_cash_negativo", "prob_sotto_obiettivo", "prob_perdita_capitale", "prob_batte_alternativa"]
        for indice, (etichetta, formula, nota) in enumerate([
            ("Probabilita' di cash flow negativo",
             '=COUNTIF(sim_cash_flow,"<0")/COUNT(sim_cash_flow)',
             "Quanto spesso l'immobile assorbe cassa invece di darne. Con la leva e' quasi sempre alta: la domanda non e' se accade ma se e' sostenibile."),
            ("Probabilita' di rendimento sotto l'obiettivo",
             '=COUNTIF(sim_rendimento,"<"&rend_obiettivo)/COUNT(sim_rendimento)',
             "Rispetto alla soglia impostata nel foglio Scenari."),
            ("Probabilita' di perdere capitale proprio",
             '=COUNTIF(sim_patrimonio,"<"&esborso)/COUNT(sim_patrimonio)',
             "Patrimonio finale inferiore a quanto messo all'inizio, senza contare l'inflazione."),
            ("Probabilita' di battere il portafoglio alternativo",
             '=COUNTIF(sim_montante,">"&(esborso*(1+rend_port)^orizzonte))/COUNT(sim_montante)',
             "Entrambe le strade partono dallo stesso esborso; chi compra versa anche i flussi negativi, che nel montante sono capitalizzati allo stesso rendimento. Per l'abitazione principale il confronto corretto resta quello del foglio Confronto affitto, che tiene conto del canone risparmiato."),
        ]):
            riga = r
            r = S.campo(ws, r, etichetta, formula, S.PERC_1, risultato=True, nota=nota)
            self.nome(nomi_prob[indice], ws, f"B{riga}")
        r += 1

        r = S.sezione(ws, r, "Quale ipotesi pesa di piu'", 5, secondaria=True)
        r = S.nota_riga(ws, r, "Ogni variabile viene mossa da sola del dieci per cento in meno e in piu', tenendo ferme le altre, e si misura di quanto si sposta il cash flow annuo. L'ordine dice dove conviene spendere tempo a raccogliere un dato migliore: sulla variabile in cima si',  su quella in fondo no.", 5)
        r = S.intestazioni(ws, r, ["Variabile", "Meno 10%", "Piu' 10%", "Ampiezza"], [46, 18, 18, 18])
        prima_t = r
        costi_fissi = "(ricavo_effettivo-noi_annuo)"
        base_cf = f"(ricavo_effettivo-{costi_fissi}-ricavo_effettivo*ced_libero-rata_annua)"
        variabili = [
            ("Canone", "((ricavo_effettivo*{f})-{cf}-(ricavo_effettivo*{f})*ced_libero-rata_annua)"),
            ("Costi operativi", "(ricavo_effettivo-{cf}*{f}-ricavo_effettivo*ced_libero-rata_annua)"),
            ("Tasso del mutuo", "(ricavo_effettivo-{cf}-ricavo_effettivo*ced_libero-IF(mutuo_importo>0,PMT(tasso*{f}/12,durata*12,-mutuo_importo)*12,0))"),
            ("Importo del mutuo", "(ricavo_effettivo-{cf}-ricavo_effettivo*ced_libero-IF(mutuo_importo>0,PMT(tasso/12,durata*12,-mutuo_importo*{f})*12,0))"),
            ("Durata del mutuo", "(ricavo_effettivo-{cf}-ricavo_effettivo*ced_libero-IF(mutuo_importo>0,PMT(tasso/12,ROUND(durata*{f},0)*12,-mutuo_importo)*12,0))"),
            ("Prezzo", "(ricavo_effettivo-({cf}+prezzo*manut_pct*({f}-1)+prezzo*ristrutt_pct/ristrutt_anni*({f}-1))-ricavo_effettivo*ced_libero-rata_annua)"),
            ("Aliquota sul canone", "(ricavo_effettivo-{cf}-ricavo_effettivo*ced_libero*{f}-rata_annua)"),
            ("Spese condominiali", "(ricavo_effettivo-({cf}+condominio*quota_condominio*({f}-1))-ricavo_effettivo*ced_libero-rata_annua)"),
        ]
        for nome, schema in variabili:
            e = ws.cell(row=r, column=1, value=nome)
            e.font = S.ETICHETTA
            e.alignment = S.SINISTRA
            giu = "=" + schema.format(f="0.9", cf=costi_fissi)
            su = "=" + schema.format(f="1.1", cf=costi_fissi)
            for col, formula in ((2, giu), (3, su)):
                c = ws.cell(row=r, column=col, value=formula)
                c.number_format = S.EURO
                c.border = S.BORDO
                c.fill = S.FILL_CALCOLO
            amp = ws.cell(row=r, column=4, value=f"=ABS($C{r}-$B{r})")
            amp.number_format = S.EURO
            amp.border = S.BORDO
            amp.fill = S.FILL_RISULTATO
            amp.font = S.ETICHETTA_BOLD
            r += 1
        ultima_t = r - 1
        ws.conditional_formatting.add(
            f"D{prima_t}:D{ultima_t}",
            ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="F8CBAD"),
        )
        r += 1
        r = S.campo(ws, r, "Cash flow di riferimento", "=" + base_cf, S.EURO, nota="Il valore centrale rispetto a cui si misurano gli scostamenti.")
        r += 1
        for testo in [
            "Una precisazione sulla simulazione, per non farle dire piu' di quello che sa. Le estrazioni assumono che le variabili siano indipendenti fra loro, e nella realta' non lo sono: quando i tassi salgono i prezzi tendono a scendere, e quando il mercato del lavoro peggiora aumentano insieme sfitto e morosita'. La distribuzione va quindi letta come una misura della dispersione degli esiti, non come una probabilita' oggettiva.",
            "L'incertezza dichiarata in cima al foglio e' l'unica cosa che si sceglie, ed e' anche l'unica che non si puo' verificare: nessuno conosce la volatilita' vera del proprio canone. Il modo onesto di usarla e' provare valori diversi e guardare se la decisione cambia. Se cambia, la decisione non era solida.",
        ]:
            r = S.nota_riga(ws, r, testo, 5)

    # ------------------------------------------------------------ comproprieta
    def foglio_comproprieta(self) -> None:
        """Ripartizione dell'operazione fra piu' acquirenti.

        Comprare in due, in tre o in N non richiede di costituire una societa': la
        comunione costituita o mantenuta al solo scopo del godimento di una o piu'
        cose non e' un contratto di societa' ed e' regolata dalle norme sulla
        comunione, art. 2248 del codice civile. La societa' nasce quando si conferisce
        per esercitare in comune un'attivita' economica, art. 2247, che e' cosa diversa
        dal possedere insieme un immobile e incassarne il canone.

        Il foglio serve a due cose. La prima e' ripartire correttamente: ciascuno
        sopporta pesi e vantaggi in proporzione alla propria quota, art. 1101, e sul
        piano fiscale ciascuno fa storia a se', perche' l'opzione per la cedolare
        secca si esercita disgiuntamente e produce effetti solo per chi l'ha
        esercitata, e perche' l'aliquota marginale IRPEF e' personale. La seconda e'
        rendere visibili le regole di governo della comunione, che sono la parte che
        si scopre tardi e che conviene scrivere in un patto prima di firmare.
        """
        ws = self.wb.create_sheet("Comproprieta")
        ws.sheet_view.showGridLines = False
        S.larghezze_colonne(ws, {"A": 26, "B": 12, "C": 16, "D": 14, "E": 20, "F": 12})
        r = S.titolo(
            ws,
            1,
            "Acquisto in piu' persone",
            "Una riga per acquirente. Le quote devono sommare a cento. Ciascuno sceglie il proprio regime fiscale e ha la propria aliquota marginale: in comproprieta' la cedolare secca si opta individualmente.",
            20,
        )

        r = S.sezione(ws, r, "Chi compra e con quale quota", 20)
        intest = [
            "Acquirente", "Quota", "Aliquota IRPEF", "Regime fiscale", "Prima casa",
            "Quota prezzo", "Quota imposte", "Quota costo totale", "Quota mutuo",
            "Esborso proprio", "Quota ricavo", "Quota costi", "Quota NOI",
            "Imposta personale", "Detrazione interessi", "Quota rata", "Utile netto",
            "Cash flow", "Rendimento sul proprio capitale", "Peso decisionale",
        ]
        larghezze = [22, 9, 13, 20, 11, 14, 13, 16, 13, 15, 13, 13, 13, 15, 15, 13, 13, 13, 18, 14]
        r = S.intestazioni(ws, r, intest, larghezze)
        prima = r

        regimi = DataValidation(
            type="list",
            formula1='"cedolare_libero,cedolare_concordato,irpef_ordinario,breve"',
            allow_blank=True,
        )
        ws.add_data_validation(regimi)
        si_no = DataValidation(type="list", formula1='"SI,NO"', allow_blank=True)
        ws.add_data_validation(si_no)

        esempi = [("Acquirente 1", 0.5, 0.33, "cedolare_libero", "SI"),
                  ("Acquirente 2", 0.5, 0.23, "cedolare_libero", "NO")]
        for indice in range(8):
            vuoto = f'$A{r}=""'
            if indice < len(esempi):
                nome, quota, aliq, reg, pc = esempi[indice]
                ws.cell(row=r, column=1, value=nome)
                ws.cell(row=r, column=2, value=quota).number_format = S.PERC_1
                ws.cell(row=r, column=3, value=aliq).number_format = S.PERC_1
                ws.cell(row=r, column=4, value=reg)
                ws.cell(row=r, column=5, value=pc)
            else:
                ws.cell(row=r, column=2).number_format = S.PERC_1
                ws.cell(row=r, column=3).number_format = S.PERC_1
            for col in range(1, 6):
                ws.cell(row=r, column=col).fill = S.FILL_INPUT
            S.scelta(regimi, ws.cell(row=r, column=4))
            S.scelta(si_no, ws.cell(row=r, column=5))

            ws.cell(row=r, column=6, value=f'=IF({vuoto},"",prezzo*$B{r})').number_format = S.EURO
            ws.cell(row=r, column=7, value=f'=IF({vuoto},"",imposte_totali*$B{r})').number_format = S.EURO
            ws.cell(row=r, column=8, value=f'=IF({vuoto},"",costo_totale*$B{r})').number_format = S.EURO
            ws.cell(row=r, column=9, value=f'=IF({vuoto},"",mutuo_importo*$B{r})').number_format = S.EURO
            ws.cell(row=r, column=10, value=f'=IF({vuoto},"",esborso*$B{r})').number_format = S.EURO
            ws.cell(row=r, column=11, value=f'=IF({vuoto},"",ricavo_effettivo*$B{r})').number_format = S.EURO
            ws.cell(row=r, column=12, value=f'=IF({vuoto},"",(ricavo_effettivo-noi_annuo)*$B{r})').number_format = S.EURO
            ws.cell(row=r, column=13, value=f'=IF({vuoto},"",noi_annuo*$B{r})').number_format = S.EURO
            # L'imposta e' personale: dipende dal regime scelto da ciascuno e, in
            # regime ordinario, dalla sua aliquota marginale.
            ws.cell(
                row=r, column=14,
                value=(
                    f'=IF({vuoto},"",IF($D{r}="cedolare_libero",$K{r}*ced_libero,'
                    f'IF($D{r}="cedolare_concordato",$K{r}*ced_conc,'
                    f'IF($D{r}="breve",$K{r}*ced_breve1,'
                    f'$K{r}*(1-abbatt_ord)*($C{r}+addizionali)+MAX($K{r}*reg_loc,reg_loc_min*$B{r})/2))))'
                ),
            ).number_format = S.EURO
            ws.cell(
                row=r, column=15,
                value=f'=IF({vuoto},"",IF(abitazione_principale="SI",MIN(INDEX(interessi_anno,1)*$B{r},detr_max*$B{r})*detr_aliq,0))',
            ).number_format = S.EURO
            ws.cell(row=r, column=16, value=f'=IF({vuoto},"",rata_annua*$B{r})').number_format = S.EURO
            ws.cell(row=r, column=17, value=f'=IF({vuoto},"",$M{r}-$N{r})').number_format = S.EURO
            ws.cell(row=r, column=18, value=f'=IF({vuoto},"",$Q{r}-$P{r}+$O{r})').number_format = S.EURO
            ws.cell(row=r, column=19, value=f'=IF(OR({vuoto},N($J{r})=0),"",$R{r}/$J{r})').number_format = S.PERC
            ws.cell(row=r, column=20, value=f'=IF({vuoto},"",$B{r})').number_format = S.PERC_1
            for col in range(6, 21):
                ws.cell(row=r, column=col).fill = S.FILL_CALCOLO
            for col in range(1, 21):
                ws.cell(row=r, column=col).border = S.BORDO
            r += 1

        ultima = r - 1
        # Riga di controllo: le quote devono sommare a uno, altrimenti tutto il foglio mente.
        c = ws.cell(row=r, column=1, value="Totale quote")
        c.font = S.ETICHETTA_BOLD
        somma = ws.cell(row=r, column=2, value=f"=SUM($B${prima}:$B${ultima})")
        somma.number_format = S.PERC_1
        somma.font = S.ETICHETTA_BOLD
        somma.border = S.BORDO
        controllo = ws.cell(
            row=r, column=4,
            value=f'=IF(ABS($B{r}-1)<0.0001,"quote coerenti","ATTENZIONE: le quote non sommano a cento")',
        )
        controllo.font = S.ETICHETTA_BOLD
        ws.conditional_formatting.add(
            f"B{r}", CellIsRule(operator="notEqual", formula=["1"], fill=S.FILL_ATTENZIONE)
        )
        for colonna, sorgente in ((10, "J"), (17, "Q"), (18, "R")):
            t = ws.cell(row=r, column=colonna, value=f"=SUM(${sorgente}${prima}:${sorgente}${ultima})")
            t.number_format = S.EURO
            t.font = S.ETICHETTA_BOLD
        r += 2

        r = S.sezione(ws, r, "Come si governa una comunione, e cosa conviene scrivere prima", 20, secondaria=True)
        for testo in [
            "Non serve costituire una societa'. La comunione costituita o mantenuta al solo scopo del godimento di una o piu' cose e' regolata dalle norme sulla comunione e non e' un contratto di societa', art. 2248 del codice civile. Il contratto di societa' presuppone che si conferisca per esercitare in comune un'attivita' economica, art. 2247: possedere insieme un immobile e incassarne il canone non lo e'. Se pero' l'attivita' diventa organizzata, tipicamente la locazione turistica gestita con servizi oppure il comprare per ristrutturare e rivendere, si scivola nell'impresa e, in mancanza di forma, in una societa' di fatto con responsabilita' illimitata e solidale di tutti.",
            "Le maggioranze si contano per valore delle quote e non per teste. L'ordinaria amministrazione si decide a maggioranza, che vincola la minoranza dissenziente, ma tutti vanno informati prima, art. 1105. Le innovazioni e gli atti eccedenti l'ordinaria amministrazione richiedono una maggioranza che rappresenti almeno due terzi del valore, art. 1108. Serve invece l'unanimita' per vendere, per costituire diritti reali e per le locazioni di durata superiore a nove anni.",
            "La regola che decide il destino dell'operazione e' l'articolo 1111: ciascun partecipante puo' sempre domandare lo scioglimento della comunione. Chiunque, in qualsiasi momento, puo' costringere gli altri a dividere e quindi, su un immobile indivisibile, a vendere. Il correttivo esiste ed e' il patto di rimanere in comunione, valido per un massimo di dieci anni e opponibile anche agli aventi causa: oltre i dieci si riduce automaticamente. Va messo per iscritto e rinnovato.",
            "Vale la pena adottare un regolamento della comunione a maggioranza, art. 1106, che disciplini l'ordinaria amministrazione e deleghi la gestione a uno dei partecipanti o a un terzo, definendone poteri e obblighi. Accanto al regolamento conviene un patto che copra cio' che il codice non risolve: patto di indivisione, diritto di prelazione reciproco sulle quote, criterio di valorizzazione della quota in caso di uscita, ripartizione delle spese straordinarie, e cosa succede se uno smette di contribuire.",
            "Ciascuno puo' disporre della propria quota e cederla a terzi, art. 1103: senza un patto di prelazione ci si puo' ritrovare in comunione con uno sconosciuto. Ciascuno deve contribuire alle spese in proporzione alla quota, ma puo' liberarsene rinunciando al proprio diritto, art. 1104, e il cessionario risponde in solido con il cedente dei contributi non versati.",
            "Sul fisco ciascuno fa storia a se'. Il reddito si dichiara pro quota, l'opzione per la cedolare secca si esercita disgiuntamente e vale solo per chi l'ha esercitata, quindi in un immobile in due uno puo' stare in cedolare e l'altro in IRPEF ordinaria. Il massimale della detrazione degli interessi e' riferito all'immobile e si ripartisce fra i cointestatari del mutuo. Sull'agevolazione prima casa ciascuno verifica i requisiti per la propria quota, con una differenza che conta: possedere una quota di altra abitazione nello stesso Comune insieme a un fratello, a un genitore o a un estraneo non preclude l'agevolazione, mentre possederla insieme al coniuge la preclude.",
            "Quando invece una struttura societaria ha senso. La societa' semplice immobiliare regge il mero godimento, evita la doppia imposizione e la disciplina delle societa' non operative, e non paga la plusvalenza oltre il quinquennio; in cambio non puo' esercitare attivita' commerciale, quindi niente compravendita speculativa e niente locazione turistica organizzata, non puo' optare per la cedolare secca perche' l'opzione e' riservata al locatore persona fisica, e i soci rispondono illimitatamente. La societa' a responsabilita' limitata separa il patrimonio ma porta reddito d'impresa, doppia imposizione sulla distribuzione, costi di contabilita' e il rischio di essere qualificata non operativa. La regola pratica: per tenere e affittare si resta in comunione, per fare impresa si costituisce una societa'.",
        ]:
            r = S.nota_riga(ws, r, testo, 20)

    # ---------------------------------------------------------------- checklist
    def foglio_checklist(self) -> None:
        ws = self.wb.create_sheet("Checklist")
        ws.sheet_view.showGridLines = False
        r = S.titolo(
            ws,
            1,
            "Verifiche prima di firmare",
            "Una proposta di acquisto accettata dal venditore e' gia' un contratto preliminare vincolante: le verifiche vanno chiuse prima, oppure vanno trasformate in condizioni scritte nella proposta stessa.",
            7,
        )
        r = S.intestazioni(
            ws, r,
            ["Fase", "Verifica", "Perche' conta", "Documento o fonte", "Chi la fa", "Stato", "Note"],
            [16, 40, 62, 34, 20, 14, 30],
        )
        prima = r

        voci = [
            ("Prima della proposta", "Visura catastale aggiornata e planimetria depositata",
             "L'atto e' nullo se manca la dichiarazione di conformita' fra planimetria e stato di fatto. Non ogni difformita' pero' produce nullita': la Cassazione distingue le irregolarita' significative dai difetti minori.",
             "Agenzia delle Entrate, servizi catastali", "Acquirente o tecnico", "da fare", ""),
            ("Prima della proposta", "Conformita' urbanistica ed edilizia",
             "E' la corrispondenza fra lo stato di fatto e tutti i titoli edilizi della storia del fabbricato. E' cosa diversa dalla conformita' catastale e va verificata separatamente: e' la difformita' che blocca davvero la vendita e il mutuo.",
             "Accesso agli atti in Comune, titoli edilizi", "Tecnico di parte", "da fare", ""),
            ("Prima della proposta", "Stato legittimo e tolleranze costruttive",
             "Il decreto Salva Casa ha ampliato le tolleranze dell'articolo 34-bis del DPR 380/2001 per le difformita' realizzate prima del 24 maggio 2024, e ha dato valore probatorio alle dichiarazioni del tecnico. Sapere in quale regime ricade l'immobile cambia il costo della regolarizzazione.",
             "Relazione del tecnico, DL 69/2024 convertito in legge 105/2024", "Tecnico di parte", "da fare", ""),
            ("Prima della proposta", "Visura ipotecaria ventennale",
             "Rivela ipoteche, pignoramenti, sequestri, diritti di terzi, servitu' e trascrizioni pregiudizievoli. L'ipoteca del venditore va cancellata prima o contestualmente al rogito.",
             "Conservatoria dei registri immobiliari", "Notaio o visurista", "da fare", ""),
            ("Prima della proposta", "Atto di provenienza e continuita' delle trascrizioni",
             "Dice come il venditore e' diventato proprietario. Una provenienza per donazione e' un rischio concreto per il mutuo, perche' l'immobile e' aggredibile dai legittimari lesi.",
             "Atto notarile di acquisto o successione", "Notaio", "da fare", ""),
            ("Prima della proposta", "Quotazioni OMI della zona e comparabili reali",
             "Ancora il prezzo a un riferimento verificabile invece che alla richiesta dell'agenzia. Le quotazioni OMI sono semestrali, gratuite e pubbliche.",
             "Osservatorio del mercato immobiliare, Agenzia delle Entrate", "Acquirente", "da fare", ""),
            ("Prima della proposta", "Spese condominiali e liberatoria dell'amministratore",
             "L'acquirente risponde in solido con il venditore delle spese dell'anno in corso e di quello precedente. Vanno letti anche i verbali delle ultime assemblee, per i lavori deliberati e non ancora pagati.",
             "Consuntivi, riparti, verbali, regolamento", "Amministratore", "da fare", ""),
            ("Nella proposta", "Condizione sospensiva o risolutiva legata al mutuo",
             "Senza clausola, se la banca non delibera si perde la caparra e si deve comunque la provvigione. Con la sospensiva il contratto non produce effetti finche' la banca non eroga; con la risolutiva il contratto si scioglie se la condizione non si avvera entro il termine.",
             "Testo della proposta, articoli 1353 e seguenti del codice civile", "Acquirente e legale", "da fare", ""),
            ("Nella proposta", "Provvigione dell'agenzia legata all'avveramento della condizione",
             "La provvigione matura alla conclusione dell'affare. Se la condizione non si avvera e nulla e' stato pattuito, l'agenzia puo' comunque pretenderla: va escluso espressamente per iscritto.",
             "Testo della proposta", "Acquirente e legale", "da fare", ""),
            ("Nella proposta", "Termine per la stipula del definitivo",
             "Senza un termine, l'obbligo di concludere resta indeterminato e diventa difficile far valere l'inadempimento della controparte.",
             "Testo della proposta", "Acquirente e legale", "da fare", ""),
            ("Nella proposta", "Stato di fatto e di diritto e garanzia di libertà da gravami",
             "Va scritto che l'immobile e' trasferito libero da ipoteche, pesi, vincoli, pegni e da qualsivoglia gravame, e che il venditore garantisce la conformita' urbanistica e catastale.",
             "Testo della proposta", "Acquirente e legale", "da fare", ""),
            ("Nella proposta", "Natura delle somme versate, acconto o caparra confirmatoria",
             "La caparra confirmatoria da' diritto al doppio in caso di inadempimento del venditore; l'acconto no. La differenza va scritta, non lasciata implicita.",
             "Articolo 1385 del codice civile", "Acquirente e legale", "da fare", ""),
            ("Mutuo", "Farsi consegnare il PIES di ogni banca interpellata",
             "Il Prospetto Informativo Europeo Standardizzato e' il documento personalizzato che la banca deve consegnare gratuitamente prima che il cliente sia vincolato, ed e' l'unico modo per confrontare offerte diverse sulla stessa base. Contiene anche una tabella di ammortamento esemplificativa.",
             "Banca d'Italia, guida al mutuo ipotecario", "Acquirente", "da fare", ""),
            ("Mutuo", "Usare i sette giorni di riflessione sull'offerta vincolante",
             "Ricevuta l'offerta vincolante il consumatore ha diritto ad almeno sette giorni di riflessione, durante i quali l'offerta resta ferma per la banca e puo' essere accettata in qualsiasi momento. Sono giorni per confrontare, non per aspettare.",
             "Banca d'Italia, guida al mutuo ipotecario", "Acquirente", "da fare", ""),
            ("Mutuo", "Verificare che il tasso non sia usurario",
             "Al momento della firma il tasso non puo' superare la soglia d'usura, determinata sul tasso effettivo globale medio pubblicato trimestralmente. E' un controllo di un minuto che si fa una volta sola.",
             "Banca d'Italia, tassi effettivi globali medi", "Acquirente", "da fare", ""),
            ("Mutuo", "Confrontare la polizza della banca con il mercato",
             "La polizza incendio e scoppio e' obbligatoria ma il cliente puo' presentarne una reperita altrove, purche' di protezione equivalente, e la banca deve accettarla. Se si accetta quella proposta dalla banca, il cliente ha diritto di sapere quanta provvigione la compagnia paga alla banca stessa.",
             "Banca d'Italia, guida al mutuo ipotecario", "Acquirente", "da fare", ""),
            ("Mutuo", "Controllare la propria posizione in Centrale dei Rischi",
             "L'accesso ai propri dati e' gratuito e si fa online. Una segnalazione dimenticata o una pratica ancora aperta presso un mediatore creditizio pesa sulla delibera: l'incarico di mediazione si puo' revocare per iscritto, e con esso decade la richiesta in corso.",
             "Banca d'Italia, accesso alla Centrale dei Rischi", "Acquirente", "da fare", ""),
            ("Mutuo", "Sapere che la portabilita' e' gratuita per legge",
             "Trasferire il mutuo a un'altra banca, cioe' la surroga, e' per legge senza spese ne' penali, e non richiede il consenso della banca di partenza. In pratica se ne ottiene una nella vita del mutuo: le banche identificano il surrogatore seriale e negano la delibera, e le surroghe hanno spesso spread piu' alti proprio per questo.",
             "Banca d'Italia, guida al mutuo ipotecario", "Acquirente", "n.a.", ""),
            ("Prima del rogito", "Attestato di prestazione energetica",
             "E' obbligatorio allegarlo all'atto e va indicato negli annunci. Determina anche la classe da cui partire per ogni valutazione di adeguamento futuro.",
             "APE in corso di validita'", "Venditore", "da fare", ""),
            ("Prima del rogito", "Dichiarazione di conformita' o rispondenza degli impianti",
             "Per gli impianti realizzati dopo il 2008 serve la dichiarazione di conformita' ai sensi del DM 37/2008; per i piu' vecchi puo' bastare la dichiarazione di rispondenza rilasciata da un tecnico abilitato.",
             "Dichiarazione dell'installatore o del tecnico", "Venditore", "da fare", ""),
            ("Prima del rogito", "Trascrizione del preliminare se i tempi sono lunghi",
             "La trascrizione ai sensi dell'articolo 2645-bis protegge da ipoteche e pignoramenti iscritti dopo la firma e da' privilegio sul credito restitutorio. Ha un costo, e su tempi lunghi o su venditori a rischio lo vale.",
             "Preliminare in forma notarile", "Notaio", "da fare", ""),
            ("Prima del rogito", "Verifica dei requisiti prima casa e delle dichiarazioni in atto",
             "Residenza nel Comune o impegno a trasferirla entro diciotto mesi, assenza di altra abitazione nel Comune, assenza di altra prima casa agevolata in Italia salvo rivendita entro due anni. In comunione legale entrambi i coniugi devono intervenire in atto e rendere le dichiarazioni.",
             "Guida dell'Agenzia delle Entrate sulle agevolazioni prima casa", "Notaio", "da fare", ""),
            ("Prima del rogito", "Opzione prezzo-valore richiesta espressamente",
             "Va chiesta in atto e comporta la tassazione sul valore catastale, il blocco dell'accertamento di valore e la riduzione del trenta per cento dell'onorario notarile.",
             "Articolo 1 comma 497 legge 266/2005", "Notaio", "da fare", ""),
            ("Nuova costruzione", "Fideiussione a garanzia degli acconti",
             "Il decreto legislativo 122/2005 impone al costruttore di consegnare una fideiussione bancaria o assicurativa a garanzia di tutte le somme versate prima del trasferimento. La tutela non e' rinunciabile e ogni patto contrario e' nullo.",
             "Decreto legislativo 122/2005", "Notaio", "n.a.", ""),
            ("Nuova costruzione", "Polizza indennitaria decennale postuma",
             "Copre i danni materiali da rovina totale o parziale e da gravi difetti costruttivi per dieci anni dall'ultimazione. Va consegnata all'atto e gli estremi vanno indicati nel rogito.",
             "Decreto legislativo 122/2005, articolo 4", "Notaio", "n.a.", ""),
            ("Nuova costruzione", "Permesso di costruire, agibilita' e accatastamento",
             "La banca non delibera prima dell'accatastamento definitivo. Vanno verificati il titolo edilizio, il collaudo, l'agibilita' e la corrispondenza fra progetto approvato e stato realizzato.",
             "Titoli edilizi e certificato di agibilita'", "Tecnico di parte", "n.a.", ""),
            ("Nuova costruzione", "Capitolato, extracapitolato e cronoprogramma",
             "Distinguere cosa e' incluso nel prezzo da cosa e' extra evita la sorpresa piu' cara dell'acquisto sulla carta. Il cronoprogramma con le penali per il ritardo va scritto.",
             "Contratto di appalto e capitolato", "Acquirente", "n.a.", ""),
            ("Se si affitta", "Codice identificativo nazionale e adempimenti della locazione breve",
             "Dal 2026 il CIN va indicato in ogni annuncio e comunicazione. Servono inoltre la comunicazione alla questura degli alloggiati, i dispositivi di sicurezza obbligatori e il rispetto dei regolamenti comunali e condominiali.",
             "Ministero del turismo, banca dati strutture ricettive", "Proprietario", "n.a.", ""),
            ("Se si affitta", "Accordo territoriale e attestazione, se canone concordato",
             "Il canone concordato richiede l'attestazione di conformita' rilasciata da un'associazione firmataria dell'accordo territoriale del Comune, senza la quale i benefici fiscali decadono.",
             "Accordo territoriale del Comune", "Proprietario", "n.a.", ""),
            ("Se si affitta", "Verifica del regolamento condominiale",
             "Un regolamento contrattuale puo' vietare la locazione turistica o l'uso diverso dall'abitazione. Va letto prima di costruire un piano su affitti brevi.",
             "Regolamento condominiale trascritto", "Proprietario", "n.a.", ""),
        ]

        stato = DataValidation(type="list", formula1='"da fare,in corso,fatto,non applicabile,n.a."', allow_blank=True)
        ws.add_data_validation(stato)

        for fase, verifica, perche, fonte, chi, st, note in voci:
            ws.cell(row=r, column=1, value=fase).alignment = S.SINISTRA
            ws.cell(row=r, column=2, value=verifica).alignment = S.SINISTRA
            ws.cell(row=r, column=2).font = S.ETICHETTA_BOLD
            ws.cell(row=r, column=3, value=perche).alignment = S.SINISTRA
            ws.cell(row=r, column=4, value=fonte).alignment = S.SINISTRA
            ws.cell(row=r, column=5, value=chi).alignment = S.SINISTRA
            c = ws.cell(row=r, column=6, value=st)
            c.fill = S.FILL_INPUT
            c.alignment = S.CENTRO
            S.scelta(stato, c)
            ws.cell(row=r, column=7, value=note).fill = S.FILL_INPUT
            for col in range(1, 8):
                ws.cell(row=r, column=col).border = S.BORDO
            ws.row_dimensions[r].height = 46
            r += 1

        ultima = r - 1
        ws.auto_filter.ref = f"A{prima-1}:G{ultima}"
        ws.freeze_panes = ws.cell(row=prima, column=1)
        ws.conditional_formatting.add(
            f"F{prima}:F{ultima}",
            CellIsRule(operator="equal", formula=['"fatto"'], fill=S.FILL_RISULTATO),
        )
        ws.conditional_formatting.add(
            f"F{prima}:F{ultima}",
            CellIsRule(operator="equal", formula=['"da fare"'], fill=S.FILL_ATTENZIONE),
        )
        r += 1
        riga_aperte = r
        r = S.campo(ws, r, "Verifiche ancora aperte", f'=COUNTIF(F{prima}:F{ultima},"da fare")+COUNTIF(F{prima}:F{ultima},"in corso")', S.NUMERO, risultato=True)
        self.nome("verifiche_aperte", ws, f"B{riga_aperte}")

    # ------------------------------------------------------------------- asta
    def foglio_asta(self) -> None:
        """Costo reale di un'aggiudicazione all'asta, e confronto con il libero.

        Un'asta valutata con il modello del libero mercato da' un numero che
        sembra ottimo e non lo e', perche' le due operazioni differiscono in
        quattro punti che il modello ordinario non vede: non c'e' provvigione ma
        c'e' il compenso del delegato, il prezzo non si tratta ma si costruisce
        per rilanci, l'immobile puo' essere occupato e la liberazione ha tempi e
        costi, e non esiste garanzia per i vizi. Questo foglio li mette in conto.
        """
        ws = self.wb.create_sheet("Asta")
        ws.sheet_view.showGridLines = False
        S.larghezze_colonne(ws, {"A": 46, "B": 18, "C": 76, "D": 18})
        r = S.titolo(
            ws,
            1,
            "Acquisto all'asta giudiziaria",
            "Le celle gialle sono gli input, presi dall'avviso di vendita e dalla perizia del custode. "
            "Il confronto in fondo dice quanto sconto sul valore di mercato serve perche' l'operazione "
            "regga i suoi rischi specifici, che sono diversi da quelli di una compravendita ordinaria.",
            4,
        )

        r = S.sezione(ws, r, "L'asta", 4)
        riga_base = r
        r = S.campo(ws, r, "Prezzo base d'asta", 72_000, S.EURO, input_utente=True,
                    nota="Dall'avviso di vendita. Non e' il prezzo che pagherai: e' il punto di partenza.")
        self.nome("asta_base", ws, f"B{riga_base}")
        riga_off = r
        r = S.campo(ws, r, "Offerta minima ammessa", "=asta_base*(1-asta_ribasso_max)", S.EURO,
                    nota="L'offerta e' inefficace se inferiore di oltre un quarto al prezzo base, art. 571 c.p.c.")
        riga_rib = r
        r = S.campo(ws, r, "Ribasso massimo ammesso sull'offerta", 0.25, S.PERC, input_utente=True,
                    nota="Un quarto per legge. Si tocca solo se l'avviso di vendita dice altro.")
        self.nome("asta_ribasso_max", ws, f"B{riga_rib}")
        riga_agg = r
        r = S.campo(ws, r, "Prezzo di aggiudicazione ipotizzato", 84_000, S.EURO, input_utente=True,
                    nota="Con piu' offerenti si apre la gara: qui si mette il prezzo a cui si e' disposti a fermarsi, non la base.")
        self.nome("asta_aggiudicazione", ws, f"B{riga_agg}")
        riga_cauz = r
        r = S.campo(ws, r, "Cauzione da versare con l'offerta", "=asta_aggiudicazione*asta_cauzione_pct", S.EURO,
                    nota="Si perde a titolo di multa se non si versa il saldo nel termine, art. 587 c.p.c.")
        riga_cpct = r
        r = S.campo(ws, r, "Cauzione, quota dell'offerta", 0.10, S.PERC, input_utente=True,
                    nota="Di norma il dieci per cento. Lo dice l'avviso di vendita.")
        self.nome("asta_cauzione_pct", ws, f"B{riga_cpct}")
        riga_term = r
        r = S.campo(ws, r, "Termine per il saldo, giorni", 120, S.NUMERO, input_utente=True,
                    nota="Fissato dall'ordinanza di vendita, di norma centoventi giorni. E' il vincolo che decide se serve un mutuo gia' deliberato.")
        self.nome("asta_giorni_saldo", ws, f"B{riga_term}")
        r += 1

        r = S.sezione(ws, r, "Costi dell'aggiudicazione", 4)
        riga_imp = r
        r = S.campo(ws, r, "Imposte di trasferimento", "=MAX(reg_min,IF(asta_agevolata=\"SI\",asta_base_imponibile*reg_prima,asta_base_imponibile*reg_ord))+ipo_priv+cat_priv", S.EURO,
                    nota="Stessa disciplina del libero mercato: registro proporzionale piu' ipotecaria e catastale fisse. L'agevolazione prima casa si chiede nella domanda di partecipazione.")
        self.nome("asta_imposte", ws, f"B{riga_imp}")
        riga_agv = r
        r = S.campo(ws, r, "Agevolazione prima casa richiesta", "SI", input_utente=True,
                    nota="Si dichiara nell'offerta o subito dopo l'aggiudicazione: il decreto di trasferimento la recepisce.")
        self.nome("asta_agevolata", ws, f"B{riga_agv}")
        riga_bi = r
        r = S.campo(ws, r, "Base imponibile delle imposte",
                    "=IF(asta_prezzo_valore=\"SI\",rendita*riv_rendita*IF(asta_agevolata=\"SI\",molt_prima,molt_ord),asta_aggiudicazione)", S.EURO,
                    nota="Il prezzo-valore si applica anche alle vendite giudiziarie: lo ha stabilito la Corte costituzionale con la sentenza 6 del 2014, dichiarando illegittima l'esclusione dei trasferimenti coattivi.")
        self.nome("asta_base_imponibile", ws, f"B{riga_bi}")
        riga_pv = r
        r = S.campo(ws, r, "Opzione prezzo-valore", "SI", input_utente=True,
                    nota="Richiede la rendita catastale, che sta nella perizia. Va chiesta espressamente.")
        self.nome("asta_prezzo_valore", ws, f"B{riga_pv}")
        riga_del = r
        r = S.campo(ws, r, "Compenso del delegato a carico dell'aggiudicatario", 1_500, S.EURO, input_utente=True,
                    nota="Nelle vendite delegate ai professionisti, art. 591-bis c.p.c., una quota del compenso e' posta a carico dell'aggiudicatario. Lo dice l'avviso: va letto, non stimato.")
        self.nome("asta_delegato", ws, f"B{riga_del}")
        riga_canc = r
        r = S.campo(ws, r, "Spese di cancellazione dei gravami", 500, S.EURO, input_utente=True,
                    nota="Il decreto ordina la cancellazione di pignoramenti e ipoteche, art. 586 c.p.c., ma le formalita' hanno un costo che di norma resta all'aggiudicatario.")
        riga_lib = r
        r = S.campo(ws, r, "Costo stimato della liberazione", 0, S.EURO, input_utente=True,
                    nota="Zero se l'immobile e' libero. Se e' occupato, sono le spese della procedura di rilascio curata dal custode piu' l'eventuale ripristino.")
        self.nome("asta_liberazione", ws, f"B{riga_lib}")
        riga_tec = r
        r = S.campo(ws, r, "Verifica tecnica e visure di parte", 600, S.EURO, input_utente=True,
                    nota="La perizia agli atti e' del tribunale, non tua, ed e' spesso vecchia di anni. Una verifica propria costa poco rispetto a cio' che evita.")
        riga_tot = r
        r = S.campo(ws, r, "Costo totale dell'operazione",
                    f"=asta_aggiudicazione+asta_imposte+B{riga_del}+B{riga_canc}+B{riga_lib}+B{riga_tec}", S.EURO,
                    risultato=True, nota="Non c'e' provvigione di agenzia: e' il solo risparmio strutturale dell'asta rispetto al libero.")
        self.nome("asta_costo_totale", ws, f"B{riga_tot}")
        riga_inc = r
        r = S.campo(ws, r, "Incidenza dei costi sull'aggiudicazione",
                    "=IF(asta_aggiudicazione>0,asta_costo_totale/asta_aggiudicazione-1,0)", S.PERC,
                    nota="Nel libero mercato l'incidenza sta intorno al dieci per cento. Qui dipende soprattutto dalla liberazione.")
        r += 1

        r = S.sezione(ws, r, "Il confronto con il libero mercato", 4)
        riga_mkt = r
        r = S.campo(ws, r, "Valore di mercato dell'immobile", 120_000, S.EURO, input_utente=True,
                    nota="Dalla perizia, corretta con le quotazioni OMI della zona e con i comparabili reali. La perizia tende a essere prudente e a invecchiare.")
        self.nome("asta_valore_mercato", ws, f"B{riga_mkt}")
        riga_sc = r
        r = S.campo(ws, r, "Sconto effettivo sul valore di mercato",
                    "=IF(asta_valore_mercato>0,1-asta_costo_totale/asta_valore_mercato,0)", S.PERC,
                    risultato=True, nota="E' il numero che decide. Si confronta con la soglia sotto, non con zero.")
        self.nome("asta_sconto", ws, f"B{riga_sc}")
        riga_soglia = r
        r = S.campo(ws, r, "Sconto minimo che giustifica i rischi dell'asta", 0.20, S.PERC, input_utente=True,
                    nota="Venti per cento e' l'ordine di grandezza comunemente ritenuto minimo. Non e' una regola di legge: e' il prezzo dei quattro rischi elencati sotto.")
        self.nome("asta_soglia", ws, f"B{riga_soglia}")
        riga_esito = r
        r = S.campo(ws, r, "Esito",
                    "=IF(asta_sconto>=asta_soglia,\"lo sconto copre i rischi specifici dell'asta\",\"sconto insufficiente: al libero mercato compreresti meglio\")",
                    risultato=True)
        ws.cell(row=riga_esito, column=2).alignment = S.SINISTRA
        riga_max = r
        r = S.campo(ws, r, "Prezzo massimo a cui fermarsi in gara",
                    f"=MAX(0,asta_valore_mercato*(1-asta_soglia)-asta_imposte-B{riga_del}-B{riga_canc}-B{riga_lib}-B{riga_tec})", S.EURO,
                    risultato=True, nota="E' il numero da scriversi su un foglio prima di entrare in gara, perche' in gara non si ragiona.")
        r += 1

        for cella, regola in (
            (f"B{riga_sc}", CellIsRule(operator="lessThan", formula=["asta_soglia"], fill=S.FILL_ATTENZIONE)),
            (f"B{riga_inc}", CellIsRule(operator="greaterThan", formula=["0.15"], fill=S.FILL_ATTENZIONE)),
        ):
            ws.conditional_formatting.add(cella, regola)

        r = S.sezione(ws, r, "I quattro rischi che il prezzo deve pagare", 4)
        for testo in [
            "Nessuna garanzia per i vizi. L'articolo 2922 del codice civile esclude la garanzia per i vizi della cosa nella vendita forzata, e aggiunge che la vendita non puo' essere impugnata per causa di lesione. Si compra nello stato di fatto e di diritto in cui il bene si trova: un impianto da rifare, una difformita' edilizia o una superficie inferiore a quella indicata restano interamente a carico di chi aggiudica.",
            "L'occupazione. L'articolo 560 del codice di procedura civile lascia il debitore e i familiari conviventi nel possesso dell'immobile fino alla pronuncia del decreto di trasferimento. Se poi non escono, il rilascio lo cura il custode, ma servono mesi. L'articolo 2923 del codice civile rende inoltre opponibile all'acquirente la locazione con data certa anteriore al pignoramento: si eredita il contratto e il canone. Con l'eccezione, prevista dallo stesso articolo, del canone inferiore di un terzo al giusto prezzo.",
            "Il termine per il saldo. Il prezzo va versato nel termine fissato dall'ordinanza, e chi non lo rispetta decade e perde la cauzione a titolo di multa, articolo 587 del codice di procedura civile. Il mutuo va quindi istruito prima di offrire, non dopo: l'articolo 585 prevede il finanziamento con versamento diretto alla procedura e ipoteca di primo grado, ed e' la forma che le banche conoscono.",
            "La regolarizzazione edilizia. La perizia dice se l'immobile ha difformita' e se sono sanabili, e la sanatoria post aggiudicazione ha termini piu' larghi ma un costo che va stimato prima. Il decreto di trasferimento cancella pignoramenti e ipoteche, articolo 586, ma non sana nulla di urbanistico.",
        ]:
            r = S.nota_riga(ws, r, testo, 4)
        r += 1
        r = S.nota_riga(ws, r, "Il perimetro di questo foglio e' l'aggiudicazione, non la partecipazione: non modella la gara fra offerenti ne' le aste con incanto, ormai residuali. E la perizia agli atti va letta per intero, in particolare i capitoli sullo stato occupativo, sulla regolarita' edilizia e sulla provenienza: e' il documento piu' informativo di tutta l'operazione, ed e' gratuito.", 4)

    # ------------------------------------------------------- dossier tecnico
    def foglio_dossier(self) -> None:
        """Fascicolo dei documenti da farsi consegnare prima di impegnarsi.

        Sta separato dalla Checklist perche' risponde a una domanda diversa: la
        Checklist elenca verifiche da fare, questo elenca carte da avere in mano.
        Senza le carte le verifiche non si possono fare, e la richiesta va fatta
        in trattativa, quando si ha ancora potere negoziale, non dopo la proposta
        accettata, quando l'obbligo di comprare esiste gia'.
        """
        ws = self.wb.create_sheet("Dossier tecnico")
        ws.sheet_view.showGridLines = False
        r = S.titolo(
            ws,
            1,
            "Documentazione tecnica da richiedere in trattativa",
            "E' il fascicolo che un tecnico incaricato chiede all'agenzia o al venditore prima della proposta. "
            "Le colonne gialle si compilano man mano. Un documento marcato bloccante non e' un giudizio del modello: "
            "senza quello l'atto e' nullo, il mutuo non si delibera, oppure il costo di regolarizzazione resta ignoto.",
            11,
        )
        r = S.intestazioni(
            ws, r,
            ["Famiglia", "Documento", "Chi lo rilascia o lo detiene", "Riferimento normativo",
             "Che cosa prova, e il rischio se manca", "Peso", "Costo indicativo",
             "Stato", "Richiesto il", "Ricevuto il", "Note"],
            [24, 42, 24, 30, 66, 13, 16, 15, 13, 13, 26],
        )
        prima = r

        # Il peso vale bloccante quando l'assenza impedisce l'atto, il mutuo o la
        # quantificazione di un costo; importante quando incide su prezzo o rischio;
        # se ricorre quando dipende dalla situazione concreta dell'immobile.
        voci = [
            ("Identificazione e titolarita'", "Visura catastale storica per immobile",
             "Agenzia delle Entrate, catasto", "Art. 29 c. 1-bis legge 52/1985",
             "Da' categoria, rendita, consistenza e tutte le variazioni subite. La rendita e' la base di quasi tutte le imposte, e la sequenza storica delle variazioni e' il primo indizio di lavori mai dichiarati.",
             "bloccante", "1 EUR per unita'"),
            ("Identificazione e titolarita'", "Planimetria catastale depositata",
             "Agenzia delle Entrate, catasto", "Art. 29 c. 1-bis legge 52/1985",
             "E' il termine di paragone della dichiarazione di conformita' catastale che il venditore rende in atto a pena di nullita'. Va confrontata con lo stato di fatto, stanza per stanza, non solo guardata.",
             "bloccante", "circa 2 EUR"),
            ("Identificazione e titolarita'", "Elaborato planimetrico ed elenco dei subalterni",
             "Agenzia delle Entrate, catasto", "Prassi catastale",
             "Individua parti comuni e pertinenze, cantina, box e posto auto, e dice se sono censite autonomamente. E' il documento che rivela la pertinenza che l'annuncio dava per compresa e che catastalmente non lo e'.",
             "importante", "circa 2 EUR"),
            ("Identificazione e titolarita'", "Ispezione ipotecaria ventennale su immobile e venditore",
             "Conservatoria dei registri immobiliari", "Artt. 2643 e seguenti c.c.",
             "Rivela ipoteche, pignoramenti, sequestri, domande giudiziali, servitu' trascritte e diritti di terzi. L'ipoteca del venditore si cancella prima o contestualmente al rogito, con tempi e costi da mettere a calendario.",
             "bloccante", "80-200 EUR"),
            ("Identificazione e titolarita'", "Atto di provenienza e continuita' delle trascrizioni",
             "Venditore, notaio rogante", "Artt. 2643 e 2650 c.c.",
             "Dice come il venditore e' diventato proprietario. Una provenienza donativa e' aggredibile dai legittimari lesi e molte banche non la finanziano; una successoria richiede accettazione tacita trascritta e voltura.",
             "bloccante", "gratuito dal venditore"),
            ("Identificazione e titolarita'", "Dati del venditore: identita', stato civile, regime patrimoniale",
             "Venditore", "Artt. 177 e 179 c.c.",
             "In comunione legale l'atto richiede l'intervento di entrambi i coniugi. Se il venditore e' impresa servono visura camerale, poteri del firmatario e verifica di procedure concorsuali in corso.",
             "bloccante", "visura 5-25 EUR"),
            ("Identificazione e titolarita'", "Certificato di destinazione urbanistica",
             "Comune, ufficio urbanistica", "Art. 30 c. 2 DPR 380/2001",
             "Obbligatorio a pena di nullita' quando l'atto comprende terreni; non serve per l'area di pertinenza di un fabbricato censito se inferiore a cinquemila metri quadrati. Fuori da quel caso resta utile per sapere cosa si potra' fare dell'area.",
             "se ricorre", "bollo piu' diritti"),

            ("Legittimita' urbanistica", "Titolo edilizio originario con tutti gli elaborati grafici",
             "Comune, archivio edilizio", "Art. 9-bis c. 1-bis DPR 380/2001",
             "E' il fondamento dello stato legittimo: licenza, concessione o permesso che ha previsto la costruzione. Senza gli elaborati approvati non esiste termine di paragone con lo stato di fatto, e nessun tecnico puo' asseverare la conformita'.",
             "bloccante", "accesso agli atti"),
            ("Legittimita' urbanistica", "Tutti i titoli successivi: DIA, SCIA, CILA, varianti",
             "Comune, archivio edilizio", "Art. 9-bis c. 1-bis DPR 380/2001",
             "Lo stato legittimo e' il titolo originario integrato dagli eventuali titoli successivi che hanno abilitato interventi parziali. Un intervento eseguito e mai titolato interrompe la catena e rende l'immobile difforme.",
             "bloccante", "compreso nell'accesso"),
            ("Legittimita' urbanistica", "Dichiarazione sostitutiva per opere iniziate prima del 1 settembre 1967",
             "Venditore, atto notorio", "Art. 40 c. 3 legge 47/1985",
             "Per gli edifici iniziati prima di quella data, in luogo degli estremi della licenza si puo' produrre una dichiarazione sostitutiva di atto notorio che ne attesti l'inizio anteriore. E' la via che rende commerciabile un fabbricato antico privo di titolo.",
             "se ricorre", "gratuito"),
            ("Legittimita' urbanistica", "Condono: domanda, ricevute di oblazione e oneri, sanatoria",
             "Comune, venditore", "Leggi 47/1985, 724/1994, 326/2003",
             "Un condono chiesto e non concluso lascia l'immobile in sospeso: il titolo non c'e' ancora e l'esito non e' garantito. Vanno verificati il pagamento integrale e, se il provvedimento manca, lo stato dell'istruttoria.",
             "se ricorre", "accesso agli atti"),
            ("Legittimita' urbanistica", "Dichiarazione asseverata sulle tolleranze costruttive",
             "Tecnico abilitato", "Art. 34-bis c. 3 DPR 380/2001",
             "Le tolleranze non sono violazioni ma vanno dichiarate dal tecnico con atto asseverato da allegare al trasferimento. Per le opere entro il 24 maggio 2024 la soglia e' del 5 per cento sotto i cento metri quadrati, 4 fra cento e trecento, 3 fra trecento e cinquecento, 2 oltre.",
             "bloccante", "compresa nella perizia"),
            ("Legittimita' urbanistica", "Relazione di conformita' urbanistica e catastale di parte",
             "Tecnico incaricato dall'acquirente", "Art. 9-bis DPR 380/2001, art. 29 legge 52/1985",
             "Sintetizza tutto il resto e quantifica il costo di regolarizzazione di cio' che non torna. Va commissionata dall'acquirente: la relazione del tecnico del venditore non risponde verso di lui.",
             "bloccante", "400-900 EUR"),
            ("Legittimita' urbanistica", "Agibilita' o certificato storico di abitabilita'",
             "Comune, venditore", "Art. 24 DPR 380/2001",
             "Attesta sicurezza, igiene, salubrita', risparmio energetico e conformita' dell'opera al progetto. La sua assenza non impedisce l'atto ma e' un indice: quasi sempre significa che qualcosa non fu mai chiuso.",
             "importante", "accesso agli atti"),
            ("Legittimita' urbanistica", "Assenza di ordinanze, diffide e procedimenti sanzionatori",
             "Comune, edilizia privata", "Art. 22 legge 241/1990",
             "Un procedimento aperto o un'ordinanza di demolizione non trascritta non compare in visura ipotecaria e si scopre solo chiedendolo. Serve la delega del proprietario oppure la dimostrazione di un interesse qualificato.",
             "bloccante", "diritti di segreteria"),
            ("Legittimita' urbanistica", "Verifica della destinazione d'uso, dei frazionamenti e delle fusioni",
             "Tecnico, su titoli e catasto", "Art. 23-ter DPR 380/2001",
             "Un ufficio venduto come abitazione, o due unita' unite di fatto e non in catasto, cambiano imposte, agevolazione prima casa e possibilita' di locazione. Il confronto fra categoria catastale, titolo edilizio e stato di fatto lo rivela.",
             "importante", "compreso nella perizia"),

            ("Struttura e sismica", "Denuncia dei lavori e autorizzazione sismica",
             "Genio Civile, ufficio tecnico regionale", "Artt. 93 e 94 DPR 380/2001",
             "In zona sismica i lavori strutturali richiedono preavviso e, fuori dalla bassa sismicita', autorizzazione preventiva. La loro assenza su interventi gia' eseguiti e' una difformita' che non si sana con una semplice pratica edilizia.",
             "se ricorre", "accesso agli atti"),
            ("Struttura e sismica", "Certificato di collaudo statico",
             "Comune o Genio Civile", "Art. 67 DPR 380/2001",
             "Chiude il ciclo delle opere strutturali. Manca spesso negli edifici degli anni sessanta e settanta, e la sua assenza pesa quando si vuole intervenire sulle strutture o accedere a incentivi.",
             "importante", "accesso agli atti"),
            ("Struttura e sismica", "Documentazione degli interventi strutturali eseguiti",
             "Venditore, amministratore", "NTC 2018, art. 34-bis c. 3-bis DPR 380/2001",
             "Rinforzi, cordoli, cappotti sismici, sostituzione di solai. In zona sismica il tecnico deve attestare che anche le tolleranze rispettino le norme tecniche vigenti al tempo dell'intervento.",
             "importante", "gratuito dal venditore"),

            ("Vincoli e tutele", "Vincolo paesaggistico o monumentale e autorizzazioni rilasciate",
             "Soprintendenza, Comune", "D.lgs. 42/2004",
             "Sotto vincolo ogni intervento richiede autorizzazione preventiva, le tolleranze esecutive del comma 2 dell'articolo 34-bis non si applicano, e i tempi di qualunque lavoro futuro cambiano ordine di grandezza.",
             "se ricorre", "accesso agli atti"),
            ("Vincoli e tutele", "Se bene culturale: denuncia di trasferimento e prelazione",
             "Soprintendenza", "Artt. 59-62 d.lgs. 42/2004",
             "Il trasferimento va denunciato e lo Stato ha diritto di prelazione entro il termine di legge. Finche' il termine non decorre l'acquisto non e' definitivo, e la circostanza va scritta nella proposta.",
             "se ricorre", "gratuito"),
            ("Vincoli e tutele", "Vincolo idrogeologico, piano di assetto idrogeologico, usi civici",
             "Comune, Regione, autorita' di bacino", "Norme regionali e di piano",
             "Determinano cosa si puo' fare dell'area e, nel caso del rischio idraulico, incidono su assicurabilita' e valore. Un uso civico non estinto rende l'immobile inalienabile senza sdemanializzazione.",
             "se ricorre", "consultazione gratuita"),
            ("Vincoli e tutele", "Convenzioni urbanistiche ed edilizia convenzionata",
             "Comune, atto di provenienza", "Convenzione e norme di piano",
             "Un immobile in edilizia convenzionata puo' avere un prezzo massimo di cessione ancora vigente: pagare sopra quel prezzo espone alla ripetizione dell'eccedenza. Il vincolo si rimuove con atto oneroso, da quantificare prima di trattare.",
             "se ricorre", "accesso agli atti"),

            ("Impianti ed energia", "Dichiarazione di conformita' degli impianti",
             "Venditore, installatore", "DM 37/2008",
             "Elettrico, termico, gas, idrico e ricezione. Serve per l'agibilita' e per qualunque intervento futuro. La sua assenza non blocca l'atto ma sposta sull'acquirente il costo dell'adeguamento e della messa a norma.",
             "importante", "gratuito dal venditore"),
            ("Impianti ed energia", "Dichiarazione di rispondenza per impianti anteriori al 2008",
             "Tecnico abilitato", "Art. 7 c. 6 DM 37/2008",
             "Sostituisce la dichiarazione di conformita' quando questa non e' reperibile, ed e' rilasciata da un professionista o da un'impresa con i requisiti di legge dopo verifica dell'impianto. Ha un costo e va messo a preventivo.",
             "se ricorre", "300-600 EUR"),
            ("Impianti ed energia", "Libretto di impianto e ultimo rapporto di efficienza energetica",
             "Venditore, manutentore", "DPR 74/2013",
             "Dice eta', potenza e stato della caldaia e se le manutenzioni obbligatorie sono state fatte. Una caldaia a fine vita e' una spesa certa a breve, che va scontata dal prezzo e non scoperta a dicembre.",
             "importante", "gratuito dal venditore"),
            ("Impianti ed energia", "Attestato di prestazione energetica in corso di validita'",
             "Certificatore accreditato", "D.lgs. 192/2005",
             "Va allegato all'atto e indicato nell'annuncio. Determina la classe da cui parte ogni valutazione di adeguamento futuro e incide sul valore, sui costi di gestione e sull'accesso ad alcune agevolazioni.",
             "bloccante", "150-350 EUR"),
            ("Impianti ed energia", "Denuncia dell'impianto di terra e verifiche periodiche",
             "Amministratore, INAIL o organismo abilitato", "DPR 462/2001",
             "Riguarda le parti comuni e le pertinenze come autorimesse e locali tecnici. La periodicita' delle verifiche dipende dal tipo di luogo, e l'omissione e' una responsabilita' che passa alla proprieta'.",
             "se ricorre", "gratuito dall'amministratore"),
            ("Impianti ed energia", "Documentazione dell'ascensore: matricola, dichiarazioni, verifiche biennali",
             "Amministratore, manutentore", "DPR 162/1999",
             "Un ascensore fuori verifica o da adeguare e' una spesa straordinaria in arrivo, e la delibera puo' essere gia' stata presa. Va incrociata con i verbali di assemblea.",
             "se ricorre", "gratuito dall'amministratore"),
            ("Impianti ed energia", "Certificato di prevenzione incendi",
             "Comando provinciale dei vigili del fuoco", "DPR 151/2011",
             "Riguarda le attivita' soggette, tipicamente autorimesse oltre trecento metri quadrati e centrali termiche sopra una certa potenza. La sua mancanza espone il condominio a sanzioni e a lavori di adeguamento.",
             "se ricorre", "gratuito dall'amministratore"),
            ("Impianti ed energia", "Valutazione della presenza di amianto e del suo stato",
             "Tecnico, amministratore", "Legge 257/1992, DM 6 settembre 1994",
             "Negli edifici anteriori al 1994 e' frequente in coperture, canne fumarie e tubazioni. Non e' vietato di per se' se in buono stato e confinato, ma va censito, valutato e gestito, e la rimozione ha costi rilevanti.",
             "importante", "verifica 200-500 EUR"),
            ("Impianti ed energia", "Misurazione del gas radon dove prescritta",
             "Laboratorio riconosciuto", "D.lgs. 101/2020",
             "Riguarda in particolare locali interrati e seminterrati e alcune aree regionali a rischio. Interessa chi destina l'immobile a luogo di lavoro o a locazione turistica organizzata.",
             "se ricorre", "50-150 EUR"),
            ("Impianti ed energia", "Allacci, fognatura, serbatoi interrati e fosse settiche",
             "Venditore, gestore del servizio", "Regolamenti comunali e del gestore",
             "L'assenza di allaccio alla fognatura pubblica, o la presenza di un serbatoio interrato dismesso, sono costi certi e a volte vincoli ambientali. Vanno verificati prima, non alla prima bolletta.",
             "importante", "gratuito dal venditore"),

            ("Condominio", "Regolamento di condominio, con gli estremi di trascrizione se contrattuale",
             "Amministratore", "Artt. 1138 c.c. e 63 disp. att. c.c.",
             "Un regolamento contrattuale trascritto puo' vietare la locazione turistica, l'uso diverso dall'abitazione, gli animali o il cambio di destinazione. Va letto prima di costruire un piano di reddito su affitti brevi.",
             "bloccante", "gratuito"),
            ("Condominio", "Tabelle millesimali",
             "Amministratore", "Art. 68 disp. att. c.c.",
             "Determinano la quota di ogni spesa e il peso del voto in assemblea. Tabelle non aggiornate dopo lavori o frazionamenti sono una fonte ricorrente di contenzioso fra condomini.",
             "importante", "gratuito"),
            ("Condominio", "Consuntivi degli ultimi due esercizi e preventivo in corso",
             "Amministratore", "Art. 1130 c.c.",
             "Sono l'unica base attendibile per la voce spese condominiali del modello. La stima a voce dell'agenzia e' sistematicamente ottimistica, e la differenza su vent'anni non e' piccola.",
             "importante", "gratuito"),
            ("Condominio", "Verbali delle assemblee degli ultimi tre anni",
             "Amministratore", "Art. 1136 c.c.",
             "Rivelano i lavori deliberati e non ancora eseguiti, che sono un costo certo che arriva dopo il rogito, e le liti in corso. E' il documento che piu' spesso cambia il prezzo di una trattativa.",
             "bloccante", "gratuito"),
            ("Condominio", "Dichiarazione dell'amministratore su spese insolute e liti in corso",
             "Amministratore", "Art. 63 disp. att. c.c.",
             "L'acquirente risponde in solido con il venditore per le spese dell'anno in corso e di quello precedente. Senza liberatoria si compra un debito altrui senza conoscerne l'importo.",
             "bloccante", "gratuito o piccolo diritto"),
            ("Condominio", "Attestazione sui lavori straordinari deliberati e sul fondo speciale",
             "Amministratore", "Art. 1135 c.c.",
             "Per le opere di manutenzione straordinaria l'assemblea deve costituire un fondo di importo pari all'opera. Sapere se il fondo esiste e chi lo ha versato dice quale parte della spesa arrivera' all'acquirente.",
             "importante", "gratuito"),
            ("Condominio", "Documentazione tecnica delle parti comuni",
             "Amministratore", "Norme di settore e regolamenti locali",
             "Agibilita' dell'edificio, antincendio, impianto di terra, ascensore, e dove istituito il fascicolo del fabbricato. Serve a capire se l'edificio ha adempimenti aperti che diventeranno quote straordinarie.",
             "importante", "gratuito"),

            ("Nuova costruzione", "Fideiussione a garanzia delle somme versate",
             "Costruttore, banca o assicurazione", "Artt. 2 e 3 d.lgs. 122/2005",
             "Obbligatoria per ogni somma versata prima del trasferimento, a pena di nullita' relativa azionabile dal solo acquirente. La tutela non e' rinunciabile e ogni patto contrario e' nullo: non si versa un euro senza averla in mano.",
             "bloccante", "a carico del costruttore"),
            ("Nuova costruzione", "Polizza indennitaria decennale postuma",
             "Costruttore, assicurazione", "Art. 4 d.lgs. 122/2005",
             "Copre rovina totale o parziale e gravi difetti costruttivi per dieci anni dall'ultimazione. Va consegnata all'atto e i suoi estremi vanno indicati nel rogito: verificarne massimali ed esclusioni, non solo l'esistenza.",
             "bloccante", "a carico del costruttore"),
            ("Nuova costruzione", "Preliminare con il contenuto tipizzato di legge",
             "Costruttore, notaio", "Art. 6 d.lgs. 122/2005",
             "La legge impone un contenuto minimo: descrizione, elaborati, capitolato, termini di consegna, estremi della fideiussione, ipoteche esistenti e loro frazionamento o cancellazione prima del rogito.",
             "bloccante", "compreso nell'atto"),
            ("Nuova costruzione", "Permesso di costruire, varianti ed elaborati approvati",
             "Comune, costruttore", "DPR 380/2001",
             "Va confrontato con quanto si sta comprando: superfici, altezze, destinazione delle parti comuni, numero di posti auto. La differenza fra reso e approvato e' il rischio tipico dell'acquisto sulla carta.",
             "bloccante", "accesso agli atti"),
            ("Nuova costruzione", "Agibilita' e accatastamento definitivo",
             "Comune, costruttore", "Art. 24 DPR 380/2001",
             "La banca non delibera prima dell'accatastamento definitivo. Su un immobile in costruzione la tempistica di questi due adempimenti determina la data del rogito, e va scritta con le relative penali.",
             "bloccante", "a carico del costruttore"),
            ("Nuova costruzione", "Visura camerale del costruttore e verifica di procedure concorsuali",
             "Registro imprese, tribunale", "Art. 5 d.lgs. 122/2005",
             "La tutela della legge nasce proprio dal rischio di crisi dell'impresa. Bilanci, anzianita' e assenza di procedure in corso dicono quanto quel rischio sia teorico.",
             "importante", "5-25 EUR"),
            ("Nuova costruzione", "Capitolato, extracapitolato, cronoprogramma e penali",
             "Costruttore", "Contratto di appalto e preliminare",
             "Distinguere cosa e' compreso nel prezzo da cosa e' extra evita la sorpresa piu' cara dell'acquisto sulla carta. Il cronoprogramma senza penali per il ritardo non e' un impegno.",
             "importante", "compreso"),

            ("Occupazione e tributi", "Contratti di locazione o comodato in essere",
             "Venditore", "Legge 431/1998, art. 1599 c.c.",
             "La locazione con data certa anteriore e' opponibile all'acquirente, che subentra nel contratto: si compra un immobile gia' occupato, al canone gia' pattuito, fino alla scadenza. E' fra le prime cose da chiedere.",
             "bloccante", "gratuito dal venditore"),
            ("Occupazione e tributi", "Stato di occupazione e impegno alla liberazione",
             "Venditore", "Art. 1477 c.c.",
             "Vanno scritti chi occupa l'immobile e a che titolo, e il termine entro cui sara' libero. Un'occupazione senza titolo trasforma l'acquisto in un contenzioso di durata imprevedibile.",
             "bloccante", "gratuito"),
            ("Occupazione e tributi", "Ultime ricevute IMU e TARI, e situazione dei tributi locali",
             "Venditore", "Regolamenti comunali",
             "Servono a verificare l'aliquota effettivamente applicata e la rendita usata, e a scoprire una diversa destinazione dichiarata al Comune rispetto a quella catastale.",
             "importante", "gratuito"),
            ("Occupazione e tributi", "Utenze: intestazioni, letture, subentri e stato dei pagamenti",
             "Venditore, gestori", "Regolamenti dei gestori",
             "Un contatore rimosso o una fornitura cessata da anni comportano costi e tempi di riattivazione che sorprendono chi conta di entrare subito.",
             "importante", "gratuito"),
            ("Occupazione e tributi", "Se provenienza donativa: rinunce all'azione di restituzione o polizza",
             "Notaio, venditore", "Artt. 561 e 563 c.c.",
             "La donazione espone l'acquirente all'azione dei legittimari lesi per vent'anni dalla trascrizione. Le vie praticate sono la rinuncia degli aventi diritto o una polizza specifica: entrambe hanno costi e tempi.",
             "se ricorre", "polizza 0,5-1% del valore"),
            ("Occupazione e tributi", "Se provenienza successoria: successione, accettazione trascritta, voltura",
             "Venditore, notaio", "Artt. 2648 e 2650 c.c.",
             "Senza accettazione dell'eredita' trascritta la continuita' delle trascrizioni si interrompe e il notaio non roga. E' un adempimento che il venditore spesso non ha fatto e che richiede tempo.",
             "se ricorre", "a carico del venditore"),
            ("Occupazione e tributi", "Rilievo metrico e verifica delle superfici dichiarate",
             "Tecnico incaricato dall'acquirente", "DPR 138/1998, allegato C",
             "La superficie commerciale dell'annuncio non e' un dato normato e comprende quote di balconi e pertinenze secondo criteri variabili. Il prezzo al metro quadro con cui si confronta il mercato dipende da quale superficie si usa.",
             "importante", "compreso nella perizia"),
            ("Garanzie e dichiarazioni", "Relazione notarile preliminare ipocatastale ventennale",
             "Notaio incaricato", "Artt. 2643 e seguenti c.c.",
             "Non e' la visura ma la sua lettura professionale, con la responsabilita' del notaio dietro: ricostruisce la catena delle trascrizioni per vent'anni, verifica la continuita' e censisce iscrizioni e gravami. Va chiesta prima della proposta, non alla settimana del rogito, quando cio' che emerge non si puo' piu' usare per trattare.",
             "bloccante", "300-600 EUR"),
            ("Garanzie e dichiarazioni", "Dichiarazione del venditore di liberta' da ipoteche, pignoramenti e sequestri",
             "Venditore, in atto e nella proposta", "Art. 1482 c.c.",
             "La legge tutela il compratore solo se i gravami non erano dichiarati dal venditore e da lui ignorati: in quel caso puo' sospendere il prezzo, far fissare un termine per la liberazione e ottenere la risoluzione con il danno. Se invece li conosceva, resta solo la garanzia per evizione. La dichiarazione in atto e' cio' che tiene in vita il rimedio, e va anticipata nella proposta.",
             "bloccante", "gratuito"),
            ("Garanzie e dichiarazioni", "Dichiarazione su oneri e diritti di terzi non apparenti",
             "Venditore, in atto e nella proposta", "Art. 1489 c.c.",
             "E' la voce che nessuno chiede e che nessuna visura copre. Servitu' non apparenti, comodati, diritti personali di godimento e oneri reali non si trascrivono e non compaiono da nessuna parte: se non sono dichiarati nel contratto e il compratore non ne aveva conoscenza, si puo' domandare la risoluzione o la riduzione del prezzo, ma occorre poterlo provare.",
             "bloccante", "gratuito"),
            ("Garanzie e dichiarazioni", "Assenso alla cancellazione dell'ipoteca e verifica dell'avvenuta cancellazione",
             "Banca creditrice, conservatoria", "Art. 2882 c.c., art. 40-bis d.lgs. 385/1993",
             "Estinzione e cancellazione sono due cose diverse. Con la procedura semplificata la banca rilascia quietanza e trasmette la comunicazione al conservatore entro trenta giorni senza oneri, ma puo' anche comunicare che l'ipoteca permane per un giustificato motivo ostativo. Va quindi verificata la cancellazione nei registri, non la quietanza.",
             "bloccante", "gratuito o assenso notarile"),
            ("Garanzie e dichiarazioni", "Dichiarazione sostitutiva di atto di notorieta' del venditore",
             "Venditore", "Artt. 47 e 76 DPR 445/2000",
             "E' lo strumento con cui una dichiarazione privata acquista peso: riguarda stati, qualita' e fatti a diretta conoscenza del dichiarante, e chi rende dichiarazioni mendaci o forma atti falsi e' punito ai sensi del codice penale. E' la forma da chiedere per l'assenza di controversie, per lo stato di occupazione e per le opere anteriori al 1967.",
             "importante", "bollo se autenticata"),
            ("Garanzie e dichiarazioni", "Se provenienza successoria non divisa: rinuncia alla prelazione dei coeredi",
             "Coeredi, notaio", "Art. 732 c.c.",
             "Il coerede che vende la sua quota a un estraneo deve notificare la proposta agli altri, che hanno prelazione per due mesi. Senza notifica i coeredi possono riscattare la quota dall'acquirente e da ogni successivo avente causa, finche' dura la comunione ereditaria: e' un rischio che segue il bene e non si prescrive con il rogito.",
             "se ricorre", "gratuito"),
            ("Garanzie e dichiarazioni", "Vincoli di destinazione: fondo patrimoniale, trust, atti ex art. 2645-ter",
             "Conservatoria, atto di provenienza", "Artt. 167 e 2645-ter c.c.",
             "Sono vincoli trascritti che limitano la disponibilita' del bene e richiedono consensi o autorizzazioni ulteriori. Un immobile in fondo patrimoniale con figli minori puo' richiedere l'autorizzazione del giudice: e' un passaggio che allunga i tempi e va saputo prima di fissare la data del rogito.",
             "se ricorre", "compreso nella relazione"),
            ("Garanzie e dichiarazioni", "Se il venditore e' impresa: preliminare trascritto a giusto prezzo",
             "Notaio", "Art. 166 c. 3 d.lgs. 14/2019",
             "Non e' solo protezione da ipoteche successive. Le vendite e i preliminari trascritti ai sensi dell'articolo 2645-bis, conclusi a giusto prezzo e aventi ad oggetto immobili ad uso abitativo destinati ad abitazione principale dell'acquirente o di parenti entro il terzo grado, non sono soggetti all'azione revocatoria in caso di liquidazione giudiziale del venditore.",
             "se ricorre", "trascrizione e onorario"),
            ("Garanzie e dichiarazioni", "Solvibilita' del venditore e rischio di revocatoria ordinaria",
             "Registro imprese, tribunale", "Art. 2901 c.c.",
             "Chi compra da un debitore esposto rischia l'azione revocatoria del creditore, che nel termine di cinque anni puo' far dichiarare inefficace l'atto nei suoi confronti se l'acquirente era consapevole del pregiudizio. Su un venditore in difficolta' conclamata il prezzo di mercato e la tracciabilita' del pagamento sono la difesa.",
             "importante", "5-25 EUR"),
            ("Garanzie e dichiarazioni", "Capacita' e legittimazione delle parti",
             "Venditore, tribunale, registro imprese", "Artt. 320, 374 e 2384 c.c.",
             "Minori, interdetti e beneficiari di amministrazione di sostegno richiedono l'autorizzazione del giudice; una societa' richiede la verifica dei poteri del firmatario e il certificato di vigenza. Un atto stipulato da chi non poteva stipularlo e' un contenzioso che comincia dopo il pagamento.",
             "bloccante", "gratuito o vigenza 10 EUR"),
            ("Garanzie e dichiarazioni", "Verifica delle prelazioni legali applicabili",
             "Notaio, Comune, Soprintendenza", "Art. 732 c.c., d.lgs. 42/2004, legge 590/1965",
             "Coeredi, beni culturali, fondi agricoli confinanti e, per gli immobili urbani non abitativi, il conduttore. Una prelazione non rispettata da' al titolare il riscatto contro l'acquirente, quindi va esclusa per iscritto prima e non spiegata dopo.",
             "importante", "compreso nella relazione"),
            ("Garanzie e dichiarazioni", "Dichiarazione in atto su mediazione e modalita' di pagamento",
             "Acquirente e venditore, in atto", "Art. 35 c. 22 DL 223/2006",
             "Le parti devono dichiarare in atto le analitiche modalita' di pagamento, se si sono avvalse di un mediatore e con quali importi e mezzi di pagamento della provvigione. Non e' una formalita': la dichiarazione incompleta o mendace espone a sanzione e all'accertamento di valore, e la tracciabilita' e' anche la difesa dell'acquirente sul prezzo effettivamente pagato.",
             "bloccante", "gratuito"),
            ("Asta giudiziaria", "Perizia di stima del custode, integrale",
             "Portale delle vendite pubbliche, tribunale", "Art. 173-bis disp. att. c.p.c.",
             "E' il documento piu' informativo dell'intera operazione ed e' gratuito. Vanno letti per intero i capitoli sullo stato occupativo, sulla regolarita' edilizia e urbanistica, sulla provenienza e sulla stima del costo di regolarizzazione. Il riassunto dell'annuncio non basta, e la perizia e' del tribunale, non di chi compra: spesso e' vecchia di anni.",
             "bloccante", "gratuita"),
            ("Asta giudiziaria", "Avviso e ordinanza di vendita",
             "Delegato o cancelleria", "Artt. 570, 576 e 591-bis c.p.c.",
             "Portano i sei numeri che governano l'offerta: prezzo base, offerta minima, cauzione, termine per il saldo, quota del compenso del delegato a carico dell'aggiudicatario e modalita' di partecipazione. La quota del delegato varia e va letta, non stimata.",
             "bloccante", "gratuiti"),
            ("Asta giudiziaria", "Relazione del custode sullo stato di occupazione",
             "Custode giudiziario", "Artt. 559 e 560 c.p.c.",
             "Debitore e familiari conviventi non perdono il possesso fino al decreto di trasferimento. Sapere se l'immobile e' libero, occupato dal debitore, locato con contratto opponibile o occupato senza titolo cambia i tempi di disponibilita' e il costo della liberazione, che e' la voce piu' variabile dell'operazione.",
             "bloccante", "gratuita"),
            ("Asta giudiziaria", "Contratti di locazione opponibili e loro data certa",
             "Custode, perizia", "Art. 2923 c.c.",
             "La locazione con data certa anteriore al pignoramento e' opponibile all'acquirente, che eredita contratto e canone. L'eccezione, prevista dalla stessa norma, e' il canone inferiore di un terzo al giusto prezzo, che e' la difesa contro il contratto di comodo stipulato per svalutare il bene.",
             "se ricorre", "gratuiti"),
            ("Asta giudiziaria", "Delibera del mutuo con versamento diretto alla procedura",
             "Banca", "Art. 585 c.p.c.",
             "Il finanziamento deve prevedere il versamento diretto alle casse della procedura e l'ipoteca di primo grado sull'immobile venduto, e il decreto lo indica: il conservatore non trascrive il decreto se non insieme all'iscrizione dell'ipoteca. Non tutte le banche lo praticano, e va verificato prima di versare la cauzione, non dopo.",
             "bloccante", "istruttoria"),
            ("Asta giudiziaria", "Stima del costo di regolarizzazione delle difformita'",
             "Tecnico di parte, sulla perizia", "Art. 586 c.p.c., DPR 380/2001",
             "Il decreto di trasferimento cancella pignoramenti e ipoteche ma non sana nulla di urbanistico. La sanatoria dopo l'aggiudicazione ha termini piu' larghi e un costo che va stimato prima di offrire, perche' entra nel prezzo massimo a cui ci si puo' spingere.",
             "bloccante", "compresa nella verifica"),
            ("Asta giudiziaria", "Verifica tecnica di parte sull'immobile",
             "Tecnico incaricato dall'acquirente", "Art. 2922 c.c.",
             "Nella vendita forzata non ha luogo la garanzia per i vizi e la vendita non si puo' impugnare per lesione: tutto cio' che si scopre dopo resta a carico di chi aggiudica. E' la ragione per cui la visita con il custode va fatta, e va fatta con un tecnico.",
             "bloccante", "300-700 EUR"),
        ]

        stato = DataValidation(
            type="list",
            formula1='"da chiedere,richiesto,ricevuto,non applicabile"',
            allow_blank=True,
        )
        ws.add_data_validation(stato)

        for famiglia, documento, chi, norma, perche, peso, costo in voci:
            ws.cell(row=r, column=1, value=famiglia).alignment = S.SINISTRA
            c = ws.cell(row=r, column=2, value=documento)
            c.alignment = S.SINISTRA
            c.font = S.ETICHETTA_BOLD
            ws.cell(row=r, column=3, value=chi).alignment = S.SINISTRA
            ws.cell(row=r, column=4, value=norma).alignment = S.SINISTRA
            ws.cell(row=r, column=5, value=perche).alignment = S.SINISTRA
            ws.cell(row=r, column=6, value=peso).alignment = S.CENTRO
            ws.cell(row=r, column=7, value=costo).alignment = S.CENTRO
            s = ws.cell(row=r, column=8, value="da chiedere")
            s.fill = S.FILL_INPUT
            s.alignment = S.CENTRO
            S.scelta(stato, s)
            for col in (9, 10, 11):
                ws.cell(row=r, column=col).fill = S.FILL_INPUT
            for col in range(1, 12):
                ws.cell(row=r, column=col).border = S.BORDO
            ws.row_dimensions[r].height = 46
            r += 1

        ultima = r - 1
        ws.auto_filter.ref = f"A{prima-1}:K{ultima}"
        ws.freeze_panes = ws.cell(row=prima, column=1)
        ws.conditional_formatting.add(
            f"H{prima}:H{ultima}",
            CellIsRule(operator="equal", formula=['"ricevuto"'], fill=S.FILL_RISULTATO),
        )
        ws.conditional_formatting.add(
            f"H{prima}:H{ultima}",
            CellIsRule(operator="equal", formula=['"da chiedere"'], fill=S.FILL_ATTENZIONE),
        )
        ws.conditional_formatting.add(
            f"F{prima}:F{ultima}",
            CellIsRule(operator="equal", formula=['"bloccante"'], fill=S.FILL_ATTENZIONE),
        )
        r += 1

        r = S.sezione(ws, r, "Stato della raccolta", 11, secondaria=True)
        r = S.campo(ws, r, "Documenti in elenco", f"=COUNTA(B{prima}:B{ultima})", S.NUMERO)
        r = S.campo(ws, r, "Applicabili a questo immobile",
                    f'=COUNTA(B{prima}:B{ultima})-COUNTIF(H{prima}:H{ultima},"non applicabile")', S.NUMERO,
                    "Le voci marcate non applicabile escono dal denominatore: un immobile usato non ha le sette voci della nuova costruzione.")
        r = S.campo(ws, r, "Ricevuti", f'=COUNTIF(H{prima}:H{ultima},"ricevuto")', S.NUMERO)
        riga_pct = r
        r = S.campo(ws, r, "Completamento del fascicolo",
                    f'=IF(COUNTA(B{prima}:B{ultima})-COUNTIF(H{prima}:H{ultima},"non applicabile")=0,0,'
                    f'COUNTIF(H{prima}:H{ultima},"ricevuto")/(COUNTA(B{prima}:B{ultima})-COUNTIF(H{prima}:H{ultima},"non applicabile")))',
                    S.PERC, risultato=True)
        self.nome("documenti_completamento", ws, f"B{riga_pct}")
        riga_blocc = r
        r = S.campo(ws, r, "Documenti bloccanti ancora da avere",
                    f'=COUNTIFS(F{prima}:F{ultima},"bloccante",H{prima}:H{ultima},"da chiedere")'
                    f'+COUNTIFS(F{prima}:F{ultima},"bloccante",H{prima}:H{ultima},"richiesto")',
                    S.NUMERO, "Vanno a zero prima di firmare una proposta, oppure diventano condizioni scritte nella proposta stessa.",
                    risultato=True)
        self.nome("documenti_bloccanti_aperti", ws, f"B{riga_blocc}")
        ws.conditional_formatting.add(
            f"B{riga_blocc}",
            CellIsRule(operator="greaterThan", formula=["0"], fill=S.FILL_ATTENZIONE),
        )
        r += 1

        r = S.sezione(ws, r, "Come si usa questo foglio", 11, secondaria=True)
        for testo in [
            "La richiesta si fa per iscritto e in una volta sola, elencando i documenti con il loro riferimento normativo: una mail che chiede quindici cose motivate ottiene risposte diverse da quindici telefonate. Le voci marcate se ricorre si tolgono prima di inviare, marcandole non applicabile, cosi' la richiesta resta credibile.",
            "Cio' che l'agenzia non ha, di norma esiste comunque. I documenti catastali e ipotecari li prende un tecnico o un visurista in giornata. I titoli edilizi stanno nell'archivio del Comune e si ottengono con l'accesso agli atti, che pero' richiede la delega scritta del proprietario oppure la dimostrazione di un interesse qualificato, tipicamente la proposta gia' sottoscritta: e' la ragione per cui la proposta va condizionata invece che attesa.",
            "Un venditore che rifiuta di consegnare le carte, o un'agenzia che chiede di firmare prima, non sta necessariamente nascondendo qualcosa, ma sta chiedendo di assumersi un rischio che si puo' rifiutare. La via praticabile e' la proposta con condizione sospensiva legata all'esito della verifica tecnica, con termine breve e provvigione dovuta solo ad avveramento.",
            "Il costo complessivo della verifica preventiva sta fra le seicento e le millecinquecento euro, contando la relazione del tecnico, le visure e i diritti di accesso agli atti. E' fra lo zero virgola cinque e l'uno per cento del prezzo, ed e' l'unica spesa dell'operazione che serve a non farla, quando non va fatta.",
        ]:
            r = S.nota_riga(ws, r, testo, 11)

    # ----------------------------------------------------------------- annunci
    def foglio_annunci(self) -> None:
        ws = self.wb.create_sheet("Annunci")
        ws.sheet_view.showGridLines = False
        r = S.titolo(
            ws,
            1,
            "Registro degli immobili in valutazione",
            "Un immobile per riga. Le colonne calcolate danno prezzo al metro quadro, rendimento lordo e scarto rispetto alla quotazione OMI della zona, cosi' che il confronto sia immediato. Le due colonne in coda dichiarano il regime di acquisto della singola riga, prima casa e venditore impresa: lasciate vuote, la riga eredita il regime del foglio Immobile. Il file si popola anche dalla riga di comando con lo strumento annunci.",
            26,
        )
        # L'ordine delle colonne e' contrattuale: `annunci.esporta_in_excel` scrive
        # posizione per posizione e le tre colonne di formula non vanno mai toccate.
        colonne = [
            ("ID", 12), ("Data", 12), ("Stato", 16), ("Fonte", 18), ("Agenzia", 22),
            ("Contatto", 20), ("Link", 40), ("Comune", 20), ("Prov.", 8),
            ("Zona OMI", 12), ("Indirizzo", 28), ("Tipologia", 16),
            ("Destinazione d'uso", 22), ("Nuova costr.", 12), ("Data consegna", 14),
            ("Mq", 8), ("Prezzo richiesto", 16), ("Prezzo obiettivo", 16),
            ("Prezzo al mq", 14), ("Quotazione OMI min", 16), ("Quotazione OMI max", 16),
            ("Scarto su OMI", 14), ("Rendita catastale", 16), ("Categoria", 12),
            ("Piano", 8), ("Classe energetica", 14), ("Spese condominio anno", 18),
            ("Canone atteso mese", 16), ("Rendimento lordo", 14),
            ("Asta", 8), ("Base d'asta", 14), ("Data asta", 12),
            ("Tribunale e procedura", 24), ("Stato occupazione", 24),
            ("Punteggio", 10), ("Note", 46),
            ("Prima casa", 12), ("Venditore impresa", 16),
        ]
        CALCOLATE = (19, 22, 29)   # prezzo al mq, scarto su OMI, rendimento lordo
        # Le colonne con una tendina. Il riempimento generico delle righe, piu'
        # sotto, le salta: senza questa esclusione le colorava di giallo dopo che
        # `S.scelta` le aveva colorate di azzurro, e la cella tornava a somigliare
        # a una dove si digita. Il test lo ha trovato prima di me.
        A_SCELTA = (3, 13, 14, 37, 38)   # stato, destinazione d'uso, nuova, prima casa, impresa
        TOTALE = len(colonne)
        r = S.intestazioni(ws, r, [c[0] for c in colonne], [c[1] for c in colonne])
        prima = r

        # L'elenco arriva da `annunci.STATI_ANNUNCIO`, che e' la sorgente unica
        # condivisa con l'aiuto della riga di comando: due copie di un elenco di
        # valori ammessi divergono, e queste due erano gia' divergenti.
        stato = DataValidation(type="list", formula1='"' + ",".join(A.STATI_ANNUNCIO) + '"', allow_blank=True)
        ws.add_data_validation(stato)
        nuova = DataValidation(type="list", formula1='"SI,NO"', allow_blank=True)
        ws.add_data_validation(nuova)
        uso = DataValidation(type="list", formula1='"abitazione,ufficio,negozio,box,magazzino,altro"', allow_blank=True)
        ws.add_data_validation(uso)

        # Riga dimostrativa, con dati interamente di fantasia: mostra il formato
        # atteso di ogni colonna senza esporre l'annuncio, l'agenzia o il recapito di
        # nessuno. Il dominio `.invalid` e' riservato dalla RFC 2606 e non risolve.
        esempi = [
            ("house_1", date(2026, 1, 1), "da contattare", "portale.invalid", "Agenzia di esempio",
             "000 0000000", "https://portale.invalid/annuncio/1", "Comune di esempio",
             "XX", "D4", "via di esempio 1", "trilocale", "abitazione", "NO", "pronto",
             75, 89000, 82000, None, 1100, 1450, None, 420, "A/3", "1", "E", 900, 550, None, 7,
             "Riga di esempio: sovrascriverla o cancellarla al primo uso", "", ""),
        ]
        for e in esempi:
            for i, valore in enumerate(e, start=1):
                c = ws.cell(row=r, column=i, value=valore)
                c.border = S.BORDO
                c.fill = S.FILL_CALCOLO if i in CALCOLATE else S.FILL_INPUT
            r += 1

        for riga in range(prima, prima + 200):
            ws.cell(row=riga, column=2).number_format = S.DATA
            for colonna in (17, 18, 20, 21, 27, 28):
                ws.cell(row=riga, column=colonna).number_format = S.EURO
            ws.cell(row=riga, column=23).number_format = S.EURO_DEC
            # Prezzo al metro quadro, sul prezzo richiesto: e' quello confrontabile
            # con la quotazione di zona, mentre l'obiettivo e' una propria ipotesi.
            ws.cell(row=riga, column=19, value=f'=IF(N($P{riga})>0,$Q{riga}/$P{riga},"")').number_format = S.EURO
            ws.cell(
                row=riga, column=22,
                value=f'=IFERROR(IF(AND(N($T{riga})>0,N($U{riga})>0),$S{riga}/AVERAGE($T{riga},$U{riga})-1,""),"")',
            ).number_format = S.PERC
            ws.cell(
                row=riga, column=29,
                value=f'=IFERROR(IF(N($Q{riga})>0,$AB{riga}*12/$Q{riga},""),"")',
            ).number_format = S.PERC
            S.scelta(stato, ws.cell(row=riga, column=3))
            S.scelta(nuova, ws.cell(row=riga, column=14))
            S.scelta(uso, ws.cell(row=riga, column=13))
            # Il vuoto e' ammesso e significa "eredita dal foglio Immobile": la
            # validazione non lo vieta, percio' le due colonne restano a tre stati.
            S.scelta(nuova, ws.cell(row=riga, column=37))
            S.scelta(nuova, ws.cell(row=riga, column=38))
            for col in range(1, TOTALE + 1):
                ws.cell(row=riga, column=col).border = S.BORDO
                if col in A_SCELTA:
                    continue
                if col in CALCOLATE:
                    ws.cell(row=riga, column=col).fill = S.FILL_CALCOLO
                elif riga >= prima + len(esempi):
                    ws.cell(row=riga, column=col).fill = S.FILL_INPUT

        ultima = prima + 199
        self.riga_annunci = prima
        ws.auto_filter.ref = f"A{prima-1}:{get_column_letter(TOTALE)}{ultima}"
        ws.freeze_panes = ws.cell(row=prima, column=8)
        # Sopra la quotazione di zona il colore vira al rosso, sotto al verde.
        ws.conditional_formatting.add(
            f"V{prima}:V{ultima}",
            ColorScaleRule(start_type="min", start_color="C6E0B4", mid_type="num", mid_value=0, mid_color="FFF2CC", end_type="max", end_color="F8CBAD"),
        )
        ws.conditional_formatting.add(
            f"AC{prima}:AC{ultima}",
            ColorScaleRule(start_type="min", start_color="F8CBAD", end_type="max", end_color="C6E0B4"),
        )

    # ------------------------------------------------------ confronto immobili
    def foglio_confronto_immobili(self) -> None:
        """Applica il modello a ogni riga del registro annunci, una per riga.

        Il resto del workbook valuta un immobile alla volta, in profondita'. Questo
        foglio fa il movimento opposto: prende gli annunci gia' raccolti e li mette
        in fila con le stesse regole, per rispondere alla domanda che viene prima,
        cioe' quale dei candidati meriti la valutazione approfondita.

        Le colonne intermedie esistono apposta e non vanno compattate: ogni formula
        legge la colonna precedente invece di ricalcolare tutto da capo, il che
        rende ciascuna cella leggibile e ispezionabile quando un numero sorprende.
        """
        ws = self.wb.create_sheet("Confronto immobili")
        ws.sheet_view.showGridLines = False
        S.larghezze_colonne(ws, {"A": 34, "B": 18, "C": 56})
        r = S.titolo(
            ws,
            1,
            "Confronto fra gli immobili in valutazione",
            "Una riga per annuncio, alimentata dal foglio Annunci. Il prezzo usato e' l'obiettivo se compilato, altrimenti il richiesto.",
        )

        r = S.sezione(ws, r, "Assunzioni comuni a tutti gli immobili")
        riga_ltv = r
        r = S.campo(ws, r, "Loan to value applicato a ogni immobile", 0.75, S.PERC, input_utente=True, nota="Serve a confrontare a parita' di leva. Zero per confrontare gli acquisti in contanti.")
        self.nome("ltv_conf", ws, f"B{riga_ltv}")
        riga_aliq = r
        r = S.campo(ws, r, "Aliquota sul canone", "=ced_libero", S.PERC, nota="Predefinita alla cedolare secca a canone libero. Si puo' sovrascrivere con 10% per il concordato o con l'aliquota marginale per il regime ordinario.", input_utente=True)
        self.nome("aliquota_conf", ws, f"B{riga_aliq}")
        r = S.nota_riga(ws, r, "Tutto il resto arriva dagli altri fogli: tasso e durata dal foglio Mutuo, sfitto, morosita', manutenzione, assicurazione e aliquota IMU dal foglio Locazione, provvigione, notaio e altri costi dal foglio Immobile, soglia di rendimento dal foglio Scenari.")
        r = S.nota_riga(ws, r, "Il regime di acquisto si dichiara per riga, nelle due colonne in coda al foglio Annunci: prima casa e venditore impresa. Le colonne \"Prima casa\" e \"Da impresa\" di questo foglio mostrano il regime che la riga sta effettivamente usando, e sono quelle che le formule delle imposte e dei costi accessori leggono. Lasciare vuota la cella nel registro non e' un NO: significa che la riga eredita il regime impostato nel foglio Immobile, quindi un registro compilato senza toccare quelle colonne si comporta esattamente come prima.")
        r = S.nota_riga(ws, r, "Restano globali due assunzioni che il registro non porta: l'opzione prezzo-valore e la qualifica di immobile di lusso, cioe' la categoria A/1, A/8 o A/9, entrambe prese dal foglio Immobile. La prima e' una scelta che si esercita in atto e che conviene quasi sempre, la seconda riguarda un caso raro: se in lista c'e' un immobile di lusso accanto a immobili ordinari, quello va valutato a parte.")
        r = S.nota_riga(ws, r, "Una cosa che il foglio non puo' controllare per costruzione: l'agevolazione prima casa si usa una volta sola, mentre qui possono esserci piu' righe che la dichiarano. E' corretto cosi', perche' ogni riga e' un'alternativa all'altra e non un acquisto che si somma agli altri, ma la lettura giusta della graduatoria e' che il bonus andra' a una sola di quelle righe.")
        r += 1

        intest = [
            "ID", "Comune", "Mq", "Prezzo", "Prezzo al mq", "Rendita", "Canone annuo",
            "Spese cond.", "Imposte acq.", "Mutuo", "Costi accessori", "Costo totale",
            "Esborso", "Ricavo effettivo", "Costi operativi", "NOI", "Imposta canone",
            "Utile netto", "Rata annua", "Cash flow", "Rend. lordo", "Rend. netto",
            "Cap rate", "Cash on cash", "DSCR",
            "Prima casa", "Da impresa",
            "Zona OMI", "Quot. OMI min", "Quot. OMI max", "Scarto su OMI", "Esito",
        ]
        larghezze = [12, 20, 8, 14, 12, 12, 14, 12, 14, 14, 16, 14, 14, 16, 16, 14, 14,
                     14, 14, 14, 12, 12, 12, 12, 10, 11, 11, 12, 15, 15, 14, 16]
        r = S.intestazioni(ws, r, intest, larghezze)
        prima = r
        origine = self.riga_annunci

        for indice in range(60):
            s = origine + indice          # riga corrispondente nel foglio Annunci
            vuoto = f'$A{r}=""'
            ws.cell(row=r, column=1, value=f"=IF(Annunci!$A{s}=\"\",\"\",Annunci!$A{s})")
            ws.cell(row=r, column=2, value=f'=IF({vuoto},"",Annunci!$H{s})')
            ws.cell(row=r, column=3, value=f'=IF({vuoto},"",Annunci!$P{s})').number_format = S.NUMERO
            ws.cell(row=r, column=4, value=f'=IF({vuoto},"",IF(N(Annunci!$R{s})>0,Annunci!$R{s},Annunci!$Q{s}))').number_format = S.EURO
            ws.cell(row=r, column=5, value=f'=IF(OR({vuoto},N($C{r})=0),"",$D{r}/$C{r})').number_format = S.EURO
            ws.cell(row=r, column=6, value=f'=IF({vuoto},"",Annunci!$W{s})').number_format = S.EURO_DEC
            ws.cell(row=r, column=7, value=f'=IF({vuoto},"",Annunci!$AB{s}*12)').number_format = S.EURO
            ws.cell(row=r, column=8, value=f'=IF({vuoto},"",Annunci!$AA{s})').number_format = S.EURO
            ws.cell(
                row=r, column=9,
                value=(
                    f'=IF({vuoto},"",IF($AA{r}="SI",'
                    f'$D{r}*IF($Z{r}="SI",iva_prima,IF(di_lusso="SI",iva_lusso,iva_ord))+3*fisso_impresa,'
                    f'MAX(IF(AND(usa_prezzo_valore="SI",$F{r}>0),$F{r}*riv_rendita*IF($Z{r}="SI",molt_prima,molt_ord),$D{r})'
                    f'*IF($Z{r}="SI",reg_prima,reg_ord),reg_min)+ipo_priv+cat_priv))'
                ),
            ).number_format = S.EURO
            ws.cell(row=r, column=10, value=f'=IF({vuoto},"",$D{r}*ltv_conf)').number_format = S.EURO
            ws.cell(
                row=r, column=11,
                value=(
                    f'=IF({vuoto},"",$I{r}+$D{r}*provv_pct*(1+iva_provv)+notaio_cv+altri_costi'
                    f'+IF($J{r}>0,$J{r}*IF($Z{r}="SI",sost_prima,sost_ord)+istruttoria+perizia+notaio_mutuo,0))'
                ),
            ).number_format = S.EURO
            ws.cell(row=r, column=12, value=f'=IF({vuoto},"",$D{r}+$K{r})').number_format = S.EURO
            ws.cell(row=r, column=13, value=f'=IF({vuoto},"",$L{r}-$J{r})').number_format = S.EURO
            ws.cell(row=r, column=14, value=f'=IF({vuoto},"",($G{r}-Annunci!$AB{s}*mesi_sfitto)*(1-morosita_pct))').number_format = S.EURO
            ws.cell(
                row=r, column=15,
                value=(
                    f'=IF({vuoto},"",$H{r}*quota_condominio+$D{r}*manut_pct+assicurazione'
                    f'+$F{r}*riv_rendita*imu_molt*imu_aliquota+$D{r}*ristrutt_pct/ristrutt_anni)'
                ),
            ).number_format = S.EURO
            ws.cell(row=r, column=16, value=f'=IF({vuoto},"",$N{r}-$O{r})').number_format = S.EURO
            ws.cell(row=r, column=17, value=f'=IF({vuoto},"",$N{r}*aliquota_conf)').number_format = S.EURO
            ws.cell(row=r, column=18, value=f'=IF({vuoto},"",$P{r}-$Q{r})').number_format = S.EURO
            ws.cell(row=r, column=19, value=f'=IF({vuoto},"",IF($J{r}>0,PMT(tasso/12,durata*12,-$J{r})*12,0))').number_format = S.EURO
            ws.cell(row=r, column=20, value=f'=IF({vuoto},"",$R{r}-$S{r})').number_format = S.EURO
            ws.cell(row=r, column=21, value=f'=IF(OR({vuoto},N($D{r})=0),"",$G{r}/$D{r})').number_format = S.PERC
            ws.cell(row=r, column=22, value=f'=IF(OR({vuoto},N($L{r})=0),"",$R{r}/$L{r})').number_format = S.PERC
            ws.cell(row=r, column=23, value=f'=IF(OR({vuoto},N($L{r})=0),"",$P{r}/$L{r})').number_format = S.PERC
            ws.cell(row=r, column=24, value=f'=IF(OR({vuoto},N($M{r})=0),"",$T{r}/$M{r})').number_format = S.PERC
            ws.cell(row=r, column=25, value=f'=IF(OR({vuoto},N($S{r})=0),"",$P{r}/$S{r})').number_format = S.NUMERO_DEC
            # Il blocco OMI. La zona di riferimento e le due quotazioni arrivano dal
            # registro, lo scarto no: si ricalcola qui invece di leggere la colonna V
            # del foglio Annunci, perche' quella confronta la quotazione di zona con
            # il prezzo richiesto mentre questo foglio valuta il prezzo che usa
            # davvero, cioe' l'obiettivo quando e' compilato. Leggerla da la' avrebbe
            # messo in riga un solo numero riferito a un prezzo diverso da quello di
            # tutte le altre colonne, che e' il tipo di incoerenza che non si vede.
            # Il regime di acquisto della riga. Le due colonne non sono decorative:
            # sono le celle che le formule delle imposte e dei costi accessori
            # leggono, al posto dei nomi globali del foglio Immobile che leggevano
            # prima. Il registro puo' dichiararlo per riga, e il vuoto significa
            # eredita dal foglio Immobile, cioe' il comportamento precedente. Il
            # regime resta visibile accanto alle imposte proprio perche' una
            # graduatoria in cui una riga paga l'IVA e un'altra il registro va letta
            # sapendolo, non scoprendolo.
            ws.cell(row=r, column=26, value=f'=IF({vuoto},"",IF(Annunci!$AK{s}="",agevolata,Annunci!$AK{s}))')
            ws.cell(row=r, column=27, value=f'=IF({vuoto},"",IF(Annunci!$AL{s}="",da_impresa,Annunci!$AL{s}))')
            # La concatenazione con la stringa vuota non e' un vezzo: una cella vuota
            # nel registro, letta per riferimento diretto, arriva a video come zero,
            # e un codice di zona OMI che vale zero e' un dato falso, non un dato
            # assente. Lo stesso vale per le due quotazioni, che restano bianche
            # invece di dichiarare un valore di zero euro al metro quadro.
            ws.cell(row=r, column=28, value=f'=IF({vuoto},"",Annunci!$J{s}&"")')
            ws.cell(row=r, column=29, value=f'=IF(OR({vuoto},N(Annunci!$T{s})=0),"",Annunci!$T{s})').number_format = S.EURO
            ws.cell(row=r, column=30, value=f'=IF(OR({vuoto},N(Annunci!$U{s})=0),"",Annunci!$U{s})').number_format = S.EURO
            ws.cell(
                row=r, column=31,
                value=(
                    f'=IFERROR(IF(AND(N($AC{r})>0,N($AD{r})>0,N($E{r})>0),'
                    f'$E{r}/AVERAGE($AC{r},$AD{r})-1,""),"")'
                ),
            ).number_format = S.PERC
            ws.cell(
                row=r, column=32,
                value=(
                    f'=IF({vuoto},"",IF(NOT(ISNUMBER($V{r})),"",'
                    f'IF($V{r}>=rend_obiettivo,"sopra soglia","sotto soglia")))'
                ),
            )
            for col in range(1, 33):
                cella = ws.cell(row=r, column=col)
                cella.border = S.BORDO
                cella.fill = S.FILL_CALCOLO
            r += 1

        ultima = r - 1
        ws.freeze_panes = ws.cell(row=prima, column=3)
        ws.auto_filter.ref = f"A{prima-1}:AF{ultima}"
        for colonna, verso in (("V", "alto"), ("X", "alto"), ("T", "alto")):
            ws.conditional_formatting.add(
                f"{colonna}{prima}:{colonna}{ultima}",
                ColorScaleRule(start_type="min", start_color="F8CBAD", end_type="max", end_color="C6E0B4"),
            )
        ws.conditional_formatting.add(
            f"AE{prima}:AE{ultima}",
            ColorScaleRule(start_type="min", start_color="C6E0B4", mid_type="num", mid_value=0, mid_color="FFF2CC", end_type="max", end_color="F8CBAD"),
        )
        ws.conditional_formatting.add(
            f"AF{prima}:AF{ultima}",
            CellIsRule(operator="equal", formula=['"sopra soglia"'], fill=S.FILL_RISULTATO),
        )
        ws.conditional_formatting.add(
            f"AF{prima}:AF{ultima}",
            CellIsRule(operator="equal", formula=['"sotto soglia"'], fill=S.FILL_ATTENZIONE),
        )

        r += 1
        for testo in [
            "Le righe si popolano da sole man mano che il foglio Annunci si riempie: restano vuote finche' non c'e' un identificativo nella riga corrispondente. Gli annunci arrivano anche dalla riga di comando, con `python tools/valuta.py excel --con-annunci`.",
            "La colonna del cash flow e' quella che separa le operazioni sostenibili da quelle che assorbono cassa ogni mese, e non coincide quasi mai con l'ordine del rendimento lordo. Il debt service coverage ratio sotto uno dice la stessa cosa in forma di soglia.",
            "Le tre colonne del blocco OMI dicono in quale zona dell'Osservatorio cade l'immobile e fra quali quotazioni al metro quadro si muove quella zona per la tipologia, e vengono dal registro annunci: si popolano con `python tools/valuta.py omi cerca --comune ...` e finiscono qui attraverso il foglio Annunci. Lo scarto e' calcolato su questo foglio, contro il prezzo al mq della colonna E, quindi contro il prezzo obiettivo quando c'e': puo' percio' differire dallo scarto della colonna V del foglio Annunci, che confronta sempre il prezzo richiesto. La differenza fra i due numeri e' esattamente lo sconto che si sta chiedendo.",
            "Uno scarto negativo non e' di per se' un affare e uno positivo non e' di per se' un prezzo fuori mercato: la quotazione OMI e' un intervallo medio di zona per tipologia, non una stima dell'immobile, e non vede lo stato di conservazione, il piano, l'affaccio, la classe energetica ne' i lavori deliberati in condominio. Serve a segnalare le righe da capire, non a ordinarle.",
            "Questo foglio serve a scegliere quale immobile approfondire, non a decidere. Per l'immobile che sopravvive alla selezione si compila il foglio Immobile con i suoi dati reali, si verifica l'aliquota IMU nella delibera del Comune e le spese nel consuntivo condominiale, e si legge il foglio Metriche.",
        ]:
            r = S.nota_riga(ws, r, testo, 6)

    # ------------------------------------------------------------------- fonti
    def foglio_fonti(self) -> None:
        ws = self.wb.create_sheet("Fonti")
        ws.sheet_view.showGridLines = False
        r = S.titolo(
            ws,
            1,
            "Fonti",
            "Ogni numero del modello risale a una fonte verificabile. Le fonti istituzionali prevalgono sempre su quelle divulgative: queste ultime servono a orientarsi, non a decidere.",
            4,
        )
        r = S.intestazioni(ws, r, ["Categoria", "Fonte", "Cosa fornisce", "Link"], [22, 46, 62, 70])
        prima = r

        fonti = [
            ("Istituzionale", "Agenzia delle Entrate, l'acquisto della casa e le imposte", "Aliquote di registro, IVA, ipotecaria e catastale; regola prezzo-valore e moltiplicatori.", P.FONTI["imposte_acquisto"]),
            ("Istituzionale", "Agenzia delle Entrate, agevolazioni prima casa", "Requisiti, termini, decadenza e credito d'imposta per riacquisto.", P.FONTI["agevolazioni_prima_casa"]),
            ("Istituzionale", "Agenzia delle Entrate, locazioni brevi e cedolare secca", "Aliquote 21 e 26 per cento, soglia di due unita' dal 2026, obblighi degli intermediari. Guida aggiornata ad aprile 2026.", P.FONTI["locazioni_brevi"]),
            ("Istituzionale", "Agenzia delle Entrate, registrazione dei contratti di locazione", "Imposta di registro sui canoni, minimi e riduzione per il canone concordato.", P.FONTI["registrazione_locazione"]),
            ("Istituzionale", "Osservatorio del mercato immobiliare, quotazioni", "Prezzi al metro quadro di compravendita e locazione per zona omogenea, semestrali e gratuiti.", P.FONTI["quotazioni_omi"]),
            ("Dati aperti", "ondata, quotazioni immobiliari Agenzia Entrate", "Le stesse quotazioni OMI ripubblicate in CSV pronti all'uso, senza registrazione ai servizi telematici.", P.FONTI["omi_open_data"]),
            ("Dati aperti", "Attribuzione obbligatoria delle quotazioni", "Le condizioni della fornitura OMI impongono di citare la fonte quando i dati vengono usati: Agenzia Entrate - OMI. Vale per le colonne di quotazione del foglio Annunci e per lo scarto che ne deriva.", P.FONTI["quotazioni_omi"]),
            ("Istituzionale", "Banca d'Italia, guida al mutuo ipotecario", "Diritti del cliente in forma ufficiale: PIES, sette giorni di riflessione sull'offerta vincolante, portabilita' gratuita, verifica del tasso d'usura, liberta' di scelta della polizza, accesso alla Centrale dei Rischi.", "https://www.bancaditalia.it/pubblicazioni/guide-bi/guida-mutuo/"),
            ("Istituzionale", "Banca centrale europea, portale dati", "Tassi bancari armonizzati sulle nuove erogazioni per acquisto abitazione in Italia, ed Euribor. Sono la media di quello che le banche hanno davvero applicato, non un tasso pubblicitario. Interrogabili con python tools/valuta.py tassi.", "https://data.ecb.europa.eu/"),
            ("Istituzionale", "Consap, fondo di garanzia per la prima casa", "Garanzia statale fino all'ottanta per cento per under 36 con ISEE entro quarantamila euro, prorogata al 31 dicembre 2027.", P.FONTI["fondo_consap"]),
            ("Tecnica", "Carlo Pagliai, conformita' catastale e urbanistica nelle compravendite", "Differenza fra le due conformita', obblighi di legge, nullita' dell'atto e verifiche da fare.", P.FONTI["conformita_pagliai"]),
            ("Tecnica", "Carlo Pagliai, tolleranze costruttive dopo il Salva Casa", "Nuovo perimetro dell'articolo 34-bis del DPR 380/2001 e valore probatorio delle dichiarazioni del tecnico.", P.FONTI["salva_casa_tolleranze"]),
            ("Tecnica", "Carlo Pagliai, canale Telegram", "Aggiornamenti quotidiani su giurisprudenza urbanistica e commerciabilita' degli immobili.", "https://t.me/pagliaicarlo"),
            ("Notarile", "Tutele nell'acquisto di immobili da costruire", "Fideiussione obbligatoria e polizza decennale postuma del decreto legislativo 122/2005.", P.FONTI["immobili_da_costruire"]),
            ("Fiscale", "Fisco e Tasse, cedolare secca affitti brevi dal 2026", "Soglia di due appartamenti e conferma delle aliquote 21 e 26 per cento.", P.FONTI["cedolare_2026"]),
            ("Fiscale", "Fisco e Tasse, aliquote IRPEF 2026", "Scaglioni 23, 33 e 43 per cento dopo la riduzione della seconda aliquota.", P.FONTI["irpef_2026"]),
            ("Fiscale", "Fisco e Tasse, prima casa e termine di due anni", "Estensione da uno a due anni del termine per rivendere la precedente abitazione agevolata.", P.FONTI["prima_casa_due_anni"]),
            ("Fiscale", "BibLus, plusvalenza immobiliare infraquinquennale", "Perimetro dell'articolo 67 del TUIR, imposta sostitutiva del 26 per cento e finestra decennale post superbonus.", P.FONTI["plusvalenza"]),
            ("Fiscale", "MutuiOnline, detrazione degli interessi del mutuo prima casa 2026", "Massimale di quattromila euro, aliquota del diciannove per cento, requisito di residenza entro dodici mesi.", P.FONTI["detrazione_interessi"]),
            ("Divulgativa", "Paolo Coletti, foglio rendita immobiliare", "Impostazione del rendimento immobiliare su quarant'anni con sfitto, ristrutturazione periodica e confronto con l'indice azionario. File del 17 febbraio 2022: il metodo resta valido, le aliquote no.", P.FONTI["coletti_rendita"]),
            ("Divulgativa", "Paolo Coletti, foglio acquisto casa o affitto", "Confronto fra comprare e affittare con ammortamento dei costi di acquisto sugli anni di permanenza. File del 17 febbraio 2022.", P.FONTI["coletti_casa_o_affitto"]),
            ("Divulgativa", "Paolo Coletti, foglio mutuo con investimento", "Mutuo confrontato con l'investimento della liquidita' in BTP o azioni, con detrazione degli interessi modellata anno per anno. File del 29 settembre 2025.", P.FONTI["coletti_mutuo_investimento"]),
            ("Divulgativa", "Paolo Coletti, indice dei materiali", "Elenco completo dei fogli di calcolo e dei notebook pubblicati.", "https://www.paolocoletti.com/youtube/"),
            ("Divulgativa", "Bite Salad, gestione di una casa in affitto con Excel", "Modello di tracciamento dei movimenti reali di un immobile locato e calcolo del ROI anno per anno, rivisto dallo stesso Coletti.", "https://www.bitesalad.com/2024/12/affitto-excel-coletti/"),
            ("Community", "Reddit r/ItaliaPersonalFinance, comprare casa for dummies", "Guida della community ai passaggi dell'acquisto.", "https://www.reddit.com/r/ItaliaPersonalFinance/comments/17i7wfz/comprare_casa_for_dummies/"),
            ("Community", "Reddit r/ItaliaPersonalFinance, perche' comprare per affittare e' un pessimo investimento", "Tesi critica sul rendimento reale della locazione residenziale.", "https://www.reddit.com/r/ItaliaPersonalFinance/comments/1k9dfbv/"),
            ("Community", "Reddit r/ItaliaPersonalFinance, acquistare per affitti brevi non e' investire", "Distinzione fra investimento e attivita' imprenditoriale nella locazione turistica.", "https://www.reddit.com/r/ItaliaPersonalFinance/comments/1r7zufg/acquistando_un_immobile_per_affitti_brevi_non/"),
            ("Community", "Reddit r/ItaliaPersonalFinance, il simulatore mutui piu' preciso del web", "Confronto fra simulatori di ammortamento.", "https://www.reddit.com/r/ItaliaPersonalFinance/comments/1jg5mp8/"),
            ("Community", "Wiki open source di Italia Personal Finance", "Raccolta di contenuti e fogli di calcolo della community italiana.", "https://github.com/emish89/italiapersonalfinance"),
            ("Software", "Sossoldi", "Applicazione open source di tracciamento del patrimonio netto, utile a inquadrare l'immobile dentro il patrimonio complessivo.", "https://github.com/napitek/sossoldi"),
            ("Software", "ai-realestate-claude", "Motore di analisi immobiliare con punteggio della proprieta', flussi di cassa e rapporti in PDF. Mercato statunitense.", "https://github.com/zubair-trabzada/ai-realestate-claude"),
            ("Software", "RealEstateInvestmentCal", "Calcolatore in Streamlit con ROI annuo e TIR con e senza leva finanziaria.", "https://github.com/moonnstarr/RealEstateInvestmentCal"),
            ("Software", "rental-property-calculator", "Analisi di annunci con cash flow, cap rate e cash on cash.", "https://github.com/gmlesher/rental-property-calculator"),
            ("Software", "Rental Property Deal Analyzer", "Oltre venti metriche e punteggio del contratto, con estrazione dei dati dall'annuncio.", "https://github.com/berkcankapusuzoglu/Rental-Property-Deal-Analyzer"),
            ("Software", "awesome-real-estate-investing", "Raccolta curata di strumenti, piattaforme e progetti per l'investimento immobiliare.", "https://github.com/Deal-Scale/awesome-real-estate-investing"),
            ("Software", "awesome-real-estate", "Seconda raccolta curata, orientata ai progetti open source del settore.", "https://github.com/etewiah/awesome-real-estate"),
            ("Software", "Idealista Immobiliare Scraper", "Estrazione di annunci dai due portali italiani. Da usare solo nel rispetto dei termini di servizio e del file robots.txt.", "https://github.com/martapanc/Idealista-Immobiliare-Scraper"),
            ("Software", "immobiscraper", "Scraper per immobiliare.it che restituisce un DataFrame.", "https://github.com/Stemanz/immobiscraper"),
            ("Software", "ammortamento", "Generatore di piani di ammortamento alla francese in Python.", "https://github.com/ferromauro/ammortamento"),
        ]
        for categoria, nome, cosa, link in fonti:
            ws.cell(row=r, column=1, value=categoria).alignment = S.SINISTRA
            c = ws.cell(row=r, column=2, value=nome)
            c.alignment = S.SINISTRA
            c.font = S.ETICHETTA_BOLD
            ws.cell(row=r, column=3, value=cosa).alignment = S.SINISTRA
            l = ws.cell(row=r, column=4, value=link)
            l.hyperlink = link
            l.font = S.LINK
            l.alignment = S.SINISTRA
            for col in range(1, 5):
                ws.cell(row=r, column=col).border = S.BORDO
            ws.row_dimensions[r].height = 32
            r += 1
        ws.auto_filter.ref = f"A{prima-1}:D{r-1}"
        ws.freeze_panes = ws.cell(row=prima, column=1)


def genera(percorso: str) -> str:
    """Genera il workbook e restituisce il percorso scritto."""
    c = Costruttore()
    c.costruisci()
    c.salva(percorso)
    return percorso
